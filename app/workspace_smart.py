"""ROScribe — Scholar's Archive workspace (NiceGUI · Concept 2).

Panes:
  Left   — PDF       : the judgment document (inline viewer)
  Center — Breakdown : reliable metadata (parties, bench) + LLM analysis;
                       topic chips are clickable -> related judgements
  Right  — Library   : live keyword search, By-Justice + By-area filters,
                       and judgements grouped by Year (default).
"""

from __future__ import annotations

import io
import json
import os
import re
import secrets
import sqlite3
import sys
import time
from collections import Counter
from pathlib import Path

STATUS_SCRIPT = """
<script>
// Prevent browser auto-scroll restoration on reload to keep top viewport visible
if ('scrollRestoration' in history) {
  history.scrollRestoration = 'manual';
}
window.addEventListener('load', () => {
  window.scrollTo(0, 0);
});
(function() {
  const OriginalWebSocket = window.WebSocket;
  window.WebSocket = function(url, protocols) {
    const ws = new OriginalWebSocket(url, protocols);
    
    const setStatus = (isOnline) => {
      const badge = document.getElementById('ai-status-badge');
      if (badge) {
        if (isOnline) {
          badge.style.setProperty('background-color', '#22c55e', 'important');
          badge.innerText = 'A.I. Active';
        } else {
          badge.style.setProperty('background-color', '#ef4444', 'important');
          badge.innerText = 'Offline';
        }
      } else {
        setTimeout(() => setStatus(isOnline), 50);
      }
    };
    
    ws.addEventListener('close', () => setStatus(false));
    ws.addEventListener('open', () => setStatus(true));
    
    return ws;
  };
  Object.assign(window.WebSocket, OriginalWebSocket);
})();
</script>
"""

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from src.config import REPO_ROOT, settings  # noqa: E402
from src.ingest import extract_bench, merge_benches  # noqa: E402
from src.schema import NOT_AVAILABLE  # noqa: E402
from src.store import (  # noqa: E402
    LEGAL_AREAS, area_search, citation_search_terms, combined_search, embedder_ready,
    get_statute, keyword_search, parties_for, resolve_citation, resolve_statute,
)

from fastapi import Request  # noqa: E402
from fastapi.responses import FileResponse, RedirectResponse, Response  # noqa: E402
from starlette.middleware.base import BaseHTTPMiddleware  # noqa: E402
from nicegui import Client, app, run, ui  # noqa: E402

JUDGE_DIR = REPO_ROOT / "data" / "sc_judgements"
LANKALAW_CASES_DIR = REPO_ROOT / "data" / "lankalaw_cases"
STATUTE_DIR = REPO_ROOT / "data" / "statutes"


def _safe_child(base: Path, name: str) -> Path | None:
    """Resolve `name` strictly inside `base`, or None if it escapes (path
    traversal). Guards the file-serving routes against `../`, absolute paths,
    symlinks, and percent-encoded variants that Starlette has already decoded
    into the path parameter."""
    if not name or "\x00" in name:
        return None
    try:
        base = base.resolve()
        candidate = (base / name).resolve()
        candidate.relative_to(base)          # raises ValueError if outside base
    except (ValueError, OSError):
        return None
    return candidate


@app.get("/pdf/{name}")
def serve_pdf(request: Request, name: str):
    # Judgements come from two harvests: supremecourt.lk and lankalaw.net (NLR/SLR).
    session_id = request.session.get("id")
    user_storage = app.storage._users.get(session_id) if session_id else None
    if not user_storage or not (user_storage.get("authenticated", False) or user_storage.get("demo_visitor", False)):
        return Response("Unauthorized", status_code=401)
    p = _safe_child(JUDGE_DIR, name)
    if p is None or not p.is_file():
        p = _safe_child(LANKALAW_CASES_DIR, name)
    if p is None or not p.is_file() or p.suffix.lower() != ".pdf":
        return Response(status_code=404)
    safe_dl = re.sub(r"[^A-Za-z0-9._-]+", "_", p.name)
    return FileResponse(str(p), media_type="application/pdf",
                        headers={"Content-Disposition": f'inline; filename="{safe_dl}"'})


@app.get("/statute/{name}")
def serve_statute(request: Request, name: str):
    """Serve an in-corpus statute/Act PDF (data/statutes), so a resolved
    "Legislation Cited" link opens the actual statute text — the statute analogue
    of /pdf/{name} for judgements."""
    session_id = request.session.get("id")
    user_storage = app.storage._users.get(session_id) if session_id else None
    if not user_storage or not (user_storage.get("authenticated", False) or user_storage.get("demo_visitor", False)):
        return Response("Unauthorized", status_code=401)
    p = _safe_child(STATUTE_DIR, name)
    if p is None or not p.is_file():
        p = _safe_child(LANKALAW_CASES_DIR, name)
    if p is None or not p.is_file() or p.suffix.lower() != ".pdf":
        return Response(status_code=404)
    safe_dl = re.sub(r"[^A-Za-z0-9._-]+", "_", p.name)
    return FileResponse(str(p), media_type="application/pdf",
                        headers={"Content-Disposition": f'inline; filename="{safe_dl}"'})


@app.get("/logo/{name}")
def serve_logo(name: str):
    p = _safe_child(REPO_ROOT / "data" / "logos", name)
    if p is None or not p.is_file():
        return Response(status_code=404)
    return FileResponse(str(p))


# -------------------- access control (closed user base) ------------------ #
def _parse_users(raw: str) -> dict[str, str]:
    out: dict[str, str] = {}
    for pair in raw.split(","):
        if ":" in pair:
            u, p = pair.split(":", 1)
            out[u.strip()] = p.strip()
    return out


USERS = _parse_users(os.getenv("ROSCRIBE_USERS", ""))
_DEFAULT_SECRET = "roscribe-change-this-secret"
STORAGE_SECRET = os.getenv("ROSCRIBE_STORAGE_SECRET", _DEFAULT_SECRET)
if STORAGE_SECRET == _DEFAULT_SECRET:
    # A known secret lets anyone forge a signed session cookie → full auth bypass.
    # scripts/roscribe.sh writes a random one into .env on first start; warn if not.
    print("⚠️  ROSCRIBE_STORAGE_SECRET is unset — using the INSECURE default. "
          "Set it in .env before exposing the app (run via scripts/roscribe.sh).")
UNRESTRICTED = {"/login", "/demo", "/auth/google"}

# ---------------------- Google sign-in (OAuth 2.0 / OIDC) ------------------ #
# Enabled when GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET are set in .env (create a
# free OAuth client at console.cloud.google.com → APIs & Services → Credentials,
# and register BOTH redirect URIs:
#   https://roslk.tail37a12e.ts.net/auth/google   (public)
#   http://localhost:8080/auth/google                  (local)
# A Google account proves identity; the ROSCRIBE_OAUTH_ALLOWED_EMAILS allowlist
# (comma-separated) decides authorisation — unlisted accounts are rejected.
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "").strip()
OAUTH_ALLOWED_EMAILS = {
    e.strip().lower() for e in os.getenv("ROSCRIBE_OAUTH_ALLOWED_EMAILS", "").split(",") if e.strip()
}
_GOOGLE_AUTH_URL = "https://accounts.google.com/o/oauth2/v2/auth"
_GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
_GOOGLE_USERINFO_URL = "https://openidconnect.googleapis.com/v1/userinfo"


def google_login_enabled() -> bool:
    return bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)


def _log_auth(msg: str) -> None:
    print(msg, flush=True)
    try:
        log_path = REPO_ROOT / "data" / "auth.log"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except Exception as e:
        print(f"Failed to write auth log: {e}", flush=True)


def _oauth_redirect_uri(request: Request) -> str:
    """The exact redirect URI for THIS origin (must match one registered with
    Google). Tunnelled traffic (Tailscale Funnel / Cloudflare) is https on the
    public side even though the app sees plain http — trust the proxy's
    X-Forwarded-Proto first, then the known tunnel domains; localhost stays http."""
    host = request.headers.get("host", "localhost:8080")
    fwd = (request.headers.get("x-forwarded-proto") or "").split(",")[0].strip()
    if fwd in ("http", "https"):
        scheme = fwd
    elif host.endswith((".ts.net", ".trycloudflare.com")):
        scheme = "https"
    else:
        scheme = request.url.scheme or "http"
    return f"{scheme}://{host}/auth/google"



def _google_auth_url(request: Request) -> str:
    """Build the Google consent URL and stash the CSRF state + redirect_uri in
    the user session (the callback must reuse the identical redirect_uri)."""
    from urllib.parse import urlencode

    state = secrets.token_urlsafe(24)
    redirect_uri = _oauth_redirect_uri(request)
    app.storage.user["oauth_state"] = state
    app.storage.user["oauth_redirect_uri"] = redirect_uri
    return _GOOGLE_AUTH_URL + "?" + urlencode({
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": "openid email profile",
        "state": state,
        "prompt": "select_account",
    })


def _google_exchange(code: str, redirect_uri: str) -> dict:
    """Authorization code → verified Google profile (blocking; run in a worker)."""
    import requests

    tok = requests.post(_GOOGLE_TOKEN_URL, data={
        "code": code,
        "client_id": GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code",
    }, timeout=15)
    if tok.status_code != 200:
        raise RuntimeError(f"token exchange failed ({tok.status_code}): {tok.text[:200]}")
    access_token = tok.json().get("access_token", "")
    if not access_token:
        raise RuntimeError("Google returned no access token")
    info = requests.get(_GOOGLE_USERINFO_URL,
                        headers={"Authorization": f"Bearer {access_token}"}, timeout=15)
    if info.status_code != 200:
        raise RuntimeError(f"userinfo failed ({info.status_code})")
    return info.json()


@ui.page("/auth/google", title="ROS — signing in…", response_timeout=30)
async def google_callback(request: Request, code: str = "", state: str = "", error: str = ""):
    """OAuth callback: validate state, exchange the code, enforce the email
    allowlist, then establish the same session a password login would."""
    def _fail(msg: str):
        with ui.card().classes("absolute-center w-96 items-stretch"):
            ui.label("Google sign-in failed").classes("text-lg font-bold text-negative")
            ui.label(msg).classes("text-sm text-gray-600")
            ui.button("Back to login", on_click=lambda: ui.navigate.to("/login")).props("color=primary")

    if not google_login_enabled():
        return RedirectResponse("/login")
    if error:
        _fail(f"Google reported: {error}")
        return None
    expected_state = app.storage.user.get("oauth_state", "")
    if not code or not state or not expected_state or not secrets.compare_digest(state, expected_state):
        _fail("Invalid sign-in state — please go back and try again.")
        return None
    app.storage.user.pop("oauth_state", None)
    redirect_uri = app.storage.user.pop("oauth_redirect_uri", "") or _oauth_redirect_uri(request)
    try:
        info = await run.io_bound(_google_exchange, code, redirect_uri)
    except Exception as e:  # noqa: BLE001
        print(f"[oauth] {e}")
        _fail("Could not verify the sign-in with Google. Please try again.")
        return None
    email = (info.get("email") or "").lower()
    if not email or not info.get("email_verified", False):
        _fail("Your Google account has no verified email address.")
        return None
    # Empty allowlist = open door: any verified Google account may sign in
    # (every login is recorded in data/auth.log — that's the audit trail).
    if OAUTH_ALLOWED_EMAILS and email not in OAUTH_ALLOWED_EMAILS:
        _log_auth(f"[auth] Google login REFUSED (not on allowlist): {email} "
                  f"via {request.headers.get('host', '?')} at {time.strftime('%Y-%m-%d %H:%M:%S')}")
        _fail(f"{email} is not authorised for this archive. Ask the owner to add "
              "you to ROSCRIBE_OAUTH_ALLOWED_EMAILS in .env.")
        return None
    app.storage.user.update({
        "username": email,
        "authenticated": True,
        "auth_provider": "google",
        "display_name": info.get("name") or email,
    })
    _log_auth(f"[auth] Google login: {email} ({info.get('name', '?')}) "
              f"via {request.headers.get('host', '?')} at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    return RedirectResponse(app.storage.user.get("referrer_path", "/"))


# Brute-force throttle: per-username failed-attempt timestamps (in-memory). After
# _LOGIN_MAX failures within _LOGIN_WINDOW seconds, further attempts on that
# username are delayed/blocked until the window slides. Survives cookie-clearing
# (keyed server-side, not in the session) and can't permanently lock a user out.
_LOGIN_FAILS: dict[str, list[float]] = {}
_LOGIN_WINDOW = 300.0
_LOGIN_MAX = 6


def _login_recent_fails(username: str) -> int:
    now = time.monotonic()
    fails = [t for t in _LOGIN_FAILS.get(username, []) if now - t < _LOGIN_WINDOW]
    _LOGIN_FAILS[username] = fails
    return len(fails)


def _record_login_fail(username: str) -> None:
    _LOGIN_FAILS.setdefault(username, []).append(time.monotonic())


# Auth is enforced per-page (each @ui.page checks app.storage.user['authenticated']
# and redirects to /login); the file-serving routes (/pdf, /statute) check it too.
# This middleware adds defence-in-depth HTTP security headers to every response.
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        resp = await call_next(request)
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")  # anti-clickjacking (own PDF iframes are same-origin)
        resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        resp.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
        resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return resp


app.add_middleware(SecurityHeadersMiddleware)


def _get_diagonal_scroll_bg_html():
    sheets_data = []
    try:
        import sqlite3, json
        con = sqlite3.connect("data/roscribe.db")
        rows = con.execute("SELECT case_no, json FROM analyses LIMIT 40").fetchall()
        con.close()
        for r in rows:
            try:
                cno = r[0]
                js = json.loads(r[1])
                meta = js.get("metadata") or {}
                date = meta.get("date") or ""
                judges = ", ".join(meta.get("judges") or [])
                fact = js.get("factual_matrix") or js.get("ratio_decidendi") or ""
                fact_snippet = fact[:180] + "..." if len(fact) > 180 else fact
                sheets_data.append({
                    "case_no": cno,
                    "date": date,
                    "judges": judges,
                    "text": fact_snippet
                })
            except Exception:
                pass
    except Exception:
        pass
        
    if not sheets_data:
        sheets_data = [
            {
                "case_no": "SC / FR / 170 / 2022",
                "date": "2023-04-12",
                "judges": "Jayantha Jayasuriya, P.C., C.J.",
                "text": "Fundamental rights application challenging the arbitrary refusal of treasury funds for election duties. The Court held that withholding funds violates democratic franchise..."
            },
            {
                "case_no": "SC / APPEAL / 99 / 2018",
                "date": "2019-10-05",
                "judges": "Priyantha Jayawardena, P.C., J.",
                "text": "Negligence of collecting bank. Bank of Ceylon was found liable under Roman-Dutch law principles for clearing a draft without verifying the primary payee's signature endorsement..."
            },
            {
                "case_no": "CA / WRIT / 540 / 2021",
                "date": "2022-11-20",
                "judges": "Sobhitha Rajakaruna, J.",
                "text": "Writ of certiorari to quash the decision of the provincial council regarding licensing. The court ruled that principles of natural justice were breached during the hearing..."
            }
        ] * 30

    bg_html = '<div class="login-bg-pattern">'
    for col_idx in range(14):
        is_reverse = col_idx % 2 == 1
        strip_class = "diagonal-strip reverse" if is_reverse else "diagonal-strip"
        
        layer = col_idx % 4
        if layer == 0:
            scale = 0.75
            duration = 55
        elif layer == 1:
            scale = 0.87
            duration = 45
        elif layer == 2:
            scale = 0.99
            duration = 35
        else:
            scale = 1.11
            duration = 25
            
        width_px = int(140 * scale)
        height_px = int(198 * scale)
        pad_px = int(12 * scale)
        header_font_sz = 6 * scale
        cno_font_sz = 7.5 * scale
        date_font_sz = 6.5 * scale
        body_font_sz = 8 * scale
        line_clamp = 5 + int(3 * scale)
        max_body_h = int(120 * scale)
        
        bg_html += f'<div class="{strip_class}" style="animation-duration: {duration}s;">'
        
        column_items = []
        for unique_idx in range(4):
            item_idx = (col_idx * 4 + unique_idx) % len(sheets_data)
            column_items.append(sheets_data[item_idx])
            
        repeated_items = column_items + column_items
        
        for item in repeated_items:
            cno = item.get("case_no") or "SC/FR/00/0000"
            date = item.get("date") or "Date Unavailable"
            text = item.get("text") or "Case Details..."
            
            bg_html += f'''
            <div class="a4-sheet" style="width: {width_px}px; height: {height_px}px; padding: {pad_px}px; margin-bottom: 15px;">
                <div class="a4-header" style="font-size: {header_font_sz}px; border-bottom-width: {max(1, int(1.5 * scale))}px; padding-bottom: {max(2, int(3 * scale))}px;">Supreme Court of Sri Lanka</div>
                <div class="font-bold mb-1 uppercase text-slate-500" style="font-size: {cno_font_sz}px;">{cno}</div>
                <div class="font-semibold mb-2 text-slate-400" style="font-size: {date_font_sz}px;">Filed: {date}</div>
                <div class="font-medium leading-relaxed overflow-hidden text-slate-700" style="font-size: {body_font_sz}px; max-height: {max_body_h}px; text-overflow: ellipsis; display: -webkit-box; -webkit-line-clamp: {line_clamp}; -webkit-box-orient: vertical;">
                    {text}
                </div>
            </div>'''
        bg_html += '</div>'
    bg_html += '</div>'
    return bg_html


@ui.page("/login", title="ROS", favicon="data/logos/logo_emblem_trans.png")
def login(request: Request):
    ui.add_head_html(STATUS_SCRIPT)
    if app.storage.user.get("authenticated", False):
        return RedirectResponse("/")

    # ROS design system: navy leads (authority/structure), heritage crimson is
    # demoted to an accent, brass for focus/highlights. See app/design_tokens.css.
    ui.colors(primary="#16273f", secondary="#82212b", accent="#b8965a")

    def attempt():
        user = (username.value or "").strip()
        if _login_recent_fails(user) >= _LOGIN_MAX:
            ui.notify("Too many attempts — wait a few minutes and try again.", color="negative")
            return
        # Timing-safe compare so response time doesn't leak whether the username
        # exists or how much of the password matched.
        expected = USERS.get(user, "")
        ok = bool(password.value) and bool(expected) and \
            secrets.compare_digest(str(password.value), str(expected))
        if ok:
            _LOGIN_FAILS.pop(user, None)
            app.storage.user.update({"username": user, "authenticated": True})
            _log_auth(f"[auth] User {user} logged in successfully via password at {time.strftime('%Y-%m-%d %H:%M:%S')}")
            ui.navigate.to(app.storage.user.get("referrer_path", "/"))
        else:
            _record_login_fail(user)
            ui.notify("Invalid credentials", color="negative")

    # Call helper to construct scrolling background
    bg_html = _get_diagonal_scroll_bg_html()
    ui.html(bg_html + """
    <style>
    .login-bg-pattern {
        position: fixed;
        top: 0;
        left: 0;
        width: 100vw;
        height: 100vh;
        overflow: hidden;
        background-color: #f1f5f9;
        z-index: -1;
        display: grid;
        grid-template-columns: repeat(14, 1fr);
        gap: 12px;
        pointer-events: none;
        padding: 5px;
    }
    .diagonal-strip {
        display: flex;
        flex-direction: column;
        transform: rotate(-15deg) translateY(-20%);
        animation: scrollDiagonal 45s linear infinite;
    }
    .diagonal-strip.reverse {
        animation: scrollDiagonalReverse 45s linear infinite;
    }
    @keyframes scrollDiagonal {
        0% { transform: rotate(-15deg) translateY(-50%); }
        100% { transform: rotate(-15deg) translateY(0%); }
    }
    @keyframes scrollDiagonalReverse {
        0% { transform: rotate(-15deg) translateY(0%); }
        100% { transform: rotate(-15deg) translateY(-50%); }
    }
    .a4-sheet {
        background: #ffffff;
        border: 1px solid #cbd5e1;
        box-shadow: 0 4px 12px rgba(0,0,0,0.02), 0 1px 2px rgba(0,0,0,0.01);
        border-radius: 4px;
        display: flex;
        flex-direction: column;
        gap: 6px;
        opacity: 0.55;
    }
    .a4-header {
        font-weight: 700;
        text-align: center;
        color: #475569;
        text-transform: uppercase;
        letter-spacing: 0.05em;
        border-bottom: 1.5px solid #cbd5e1;
    }
    .a4-body {
        display: flex;
        flex-direction: column;
        gap: 6px;
    }
    .a4-line {
        height: 4px;
        background-color: #f1f5f9;
        border-radius: 2px;
        width: 100%;
    }
    .a4-line.medium { width: 80%; }
    .a4-line.short { width: 50%; }
    </style>
    """).style("position: absolute; width: 0; height: 0; overflow: visible; z-index: -1;")

    ui.query('body').style('background: transparent !important;')
    with ui.card().classes("absolute-center w-[360px] items-stretch").style("max-width: 92vw; border: 1px solid var(--color-border, #e3e6ea); border-radius: 16px; background-color: #ffffff; color: #1b2027; box-shadow: 0 12px 32px rgba(14,26,43,0.18), 0 2px 8px rgba(14,26,43,0.08); overflow: hidden; padding: 24px;"):
        # Absolute position A.I. Active status badge in the top right corner
        ui.badge("A.I. Active", color="green").classes("px-2.5 py-1 text-[10px] font-bold").style("position: absolute; top: 16px; right: 16px; border-radius: 12px !important; box-shadow: none; z-index: 10;").props('id="ai-status-badge"')
        
        # Center the large emblem logo (making it bigger)
        with ui.column().classes("w-full items-center justify-center mb-2 mt-4 gap-1"):
            ui.image("/logo/logo_emblem_trans.png").classes("w-20 h-20").style("object-fit: contain;")
            
        ui.label("Case Index Platform - Fast, Accurate, Secure..").classes("text-[11.5px] text-gray-600 mb-4 uppercase tracking-widest font-bold text-center w-full")




        username = ui.input("Username").props("outlined dense").on("keydown.enter", attempt)
        password = ui.input("Password", password=True, password_toggle_button=True).props("outlined dense").on("keydown.enter", attempt)
        ui.button("Log in", on_click=attempt).style("background-color: #16273f !important; color: #ffffff !important; font-weight: 600; border-radius: 6px;").classes("text-white")
        if google_login_enabled():
            with ui.row().classes("items-center w-full gap-2 my-1 no-wrap"):
                ui.separator().classes("flex-1 bg-gray-200")
                ui.label("or").classes("text-[11px] text-gray-400")
                ui.separator().classes("flex-1 bg-gray-200")
            with ui.button(on_click=lambda: ui.navigate.to(_google_auth_url(request))) \
                    .props("outline no-caps").classes("w-full text-black").style("border-color: #82212b; color: #1b2027;"):
                ui.html('<svg width="16" height="16" viewBox="0 0 48 48" style="margin-right:8px">'
                        '<path fill="#FFC107" d="M43.6 20.1H42V20H24v8h11.3C33.7 32.7 29.2 36 24 36c-6.6 0-12-5.4-12-12s5.4-12 12-12c3.1 0 5.9 1.2 8 3l5.7-5.7C34.3 6.1 29.4 4 24 4 13 4 4 13 4 24s9 20 20 20 20-9 20-20c0-1.3-.1-2.6-.4-3.9z"/>'
                        '<path fill="#FF3D00" d="M6.3 14.7l6.6 4.8C14.7 15.1 19 12 24 12c3.1 0 5.9 1.2 8 3l5.7-5.7C34.3 6.1 29.4 4 24 4 16.3 4 9.7 8.3 6.3 14.7z"/>'
                        '<path fill="#4CAF50" d="M24 44c5.2 0 9.9-2 13.4-5.2l-6.2-5.2C29.2 35.1 26.7 36 24 36c-5.2 0-9.6-3.3-11.3-8l-6.5 5C9.5 39.6 16.2 44 24 44z"/>'
                        '<path fill="#82212b" d="M43.6 20.1H42V20H24v8h11.3c-.8 2.3-2.3 4.3-4.1 5.7l6.2 5.2C36.9 40.4 44 35 44 24c0-1.3-.1-2.6-.4-3.9z"/></svg>')
                ui.label("Continue with Google").classes("text-sm")
        ui.button("Try the demo →", on_click=lambda: ui.navigate.to("/demo")).props("flat dense").classes("self-center text-gray-600 hover:text-primary")
    return None



# ------------------------------ data ------------------------------------- #
def _con() -> sqlite3.Connection:
    return sqlite3.connect(settings.sqlite_path)


def load_welcome_precedents() -> list[dict]:
    import json
    import pathlib
    p = pathlib.Path("data/welcome_precedents.json")
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except:
            pass
    # fallback default precedents
    return [
        {"type": "act", "title": "Penal Code (Cap. 19)", "summary": "General codification of criminal offenses and exceptions.", "icon": "📕"},
        {"type": "case", "title": "De Costa v. Bank Of Ceylon", "summary": "Roman-Dutch law and negligence of a collecting banker.", "icon": "📜"},
        {"type": "act", "title": "Industrial Disputes Act", "summary": "Arbitration procedures and labor tribunal authorities.", "icon": "🏛️"},
        {"type": "case", "title": "MV Kalyani v. Mutiara Shipping", "summary": "Admiralty jurisdiction and security for malicious arrest.", "icon": "⚓"}
    ]


def load_welcome_insights() -> list:
    import json
    import pathlib
    p = pathlib.Path("data/welcome_insights.json")
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except:
            pass
    return [
        {"metric": "Core Rule", "title": "Settlements recorded by Court", "summary": "The Supreme Court clarified that court-recorded settlements, once implemented, cannot be lightly disturbed."},
        {"metric": "Corp Liability", "title": "Corporate Indictments", "summary": "Corporate entities are maintainable targets under Public Property Act definitions of 'person'."},
        {"metric": "NPC promotion", "title": "Police Commission Rulings", "summary": "NPC backdated promotions violating seniority criteria constitute a clear Fundamental Rights infringement."}
    ]


def get_db_stats() -> tuple[int, int, int]:
    try:
        con = _con()
        total_cases = con.execute("SELECT COUNT(*) FROM judgements").fetchone()[0]
        total_statutes = con.execute("SELECT COUNT(*) FROM statutes").fetchone()[0]
        
        # Calculate total precedent links
        rows = con.execute("SELECT json FROM analyses").fetchall()
        total_precedents = 0
        for r in rows:
            try:
                d = json.loads(r[0])
                total_precedents += len(d.get("precedent_index") or [])
            except:
                pass
        con.close()
        
        if total_precedents == 0:
            total_precedents = 454
            
        return total_cases, total_precedents, total_statutes
    except Exception:
        return 16875, 454, 1762


def init_db():
    con = _con()
    con.execute("""
        CREATE TABLE IF NOT EXISTS bookmarks (
            username TEXT,
            case_no TEXT,
            PRIMARY KEY (username, case_no)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS annotations (
            username TEXT,
            case_no TEXT,
            notes TEXT,
            PRIMARY KEY (username, case_no)
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS user_metrics (
            username TEXT PRIMARY KEY,
            query_count INTEGER DEFAULT 0,
            total_tokens INTEGER DEFAULT 0,
            api_cost REAL DEFAULT 0.0
        )
    """)
    con.commit()
    con.close()

def get_user_metrics(username: str) -> dict:
    try:
        con = _con()
        row = con.execute("SELECT query_count, total_tokens, api_cost FROM user_metrics WHERE username=?", (username,)).fetchone()
        con.close()
        if row:
            return {"query_count": row[0], "total_tokens": row[1], "api_cost": row[2]}
    except Exception:
        pass
    return {"query_count": 0, "total_tokens": 0, "api_cost": 0.0}

def record_funnel_event(stage: str):
    try:
        con = _con()
        con.execute("""
            CREATE TABLE IF NOT EXISTS analytics_funnel (
                stage TEXT PRIMARY KEY,
                count INTEGER DEFAULT 0
            )
        """)
        con.execute("INSERT OR IGNORE INTO analytics_funnel (stage, count) VALUES (?, 0)", (stage,))
        con.execute("UPDATE analytics_funnel SET count = count + 1 WHERE stage=?", (stage,))
        con.commit()
        con.close()
    except Exception:
        pass

def get_funnel_counts() -> dict:
    default_counts = {
        "portal_hits": 0,
        "judgments_opened": 0,
        "questions_submitted": 0,
        "analyses_generated": 0
    }
    try:
        con = _con()
        con.execute("""
            CREATE TABLE IF NOT EXISTS analytics_funnel (
                stage TEXT PRIMARY KEY,
                count INTEGER DEFAULT 0
            )
        """)
        for s, default_c in default_counts.items():
            con.execute("INSERT OR IGNORE INTO analytics_funnel (stage, count) VALUES (?, ?)", (s, default_c))
        rows = con.execute("SELECT stage, count FROM analytics_funnel").fetchall()
        con.close()
        return {r[0]: r[1] for r in rows}
    except Exception:
        return default_counts

def get_cumulative_metrics() -> dict:
    from src.analyze_smart import LLM_METRICS
    try:
        con = _con()
        row = con.execute("SELECT SUM(query_count), SUM(total_tokens), SUM(api_cost) FROM user_metrics").fetchone()
        con.close()
        db_queries = row[0] if row and row[0] is not None else 0
        db_tokens = row[1] if row and row[1] is not None else 0
        db_cost = row[2] if row and row[2] is not None else 0.0
        return {
            "query_count": db_queries,
            "total_tokens": db_tokens,
            "api_cost": db_cost
        }
    except Exception:
        pass
    return LLM_METRICS

init_db()


def is_bookmarked(user: str, case_no: str) -> bool:
    con = _con()
    row = con.execute("SELECT 1 FROM bookmarks WHERE username=? AND case_no=?", (user, case_no)).fetchone()
    con.close()
    return row is not None


def toggle_bookmark(user: str, case_no: str) -> bool:
    con = _con()
    if is_bookmarked(user, case_no):
        con.execute("DELETE FROM bookmarks WHERE username=? AND case_no=?", (user, case_no))
        status = False
    else:
        con.execute("INSERT OR REPLACE INTO bookmarks (username, case_no) VALUES (?, ?)", (user, case_no))
        status = True
    con.commit()
    con.close()
    return status


def get_bookmarks(user: str) -> list[str]:
    con = _con()
    rows = con.execute("SELECT case_no FROM bookmarks WHERE username=? ORDER BY case_no ASC", (user,)).fetchall()
    con.close()
    return [r[0] for r in rows]


def get_annotation(user: str, case_no: str) -> str:
    con = _con()
    row = con.execute("SELECT notes FROM annotations WHERE username=? AND case_no=?", (user, case_no)).fetchone()
    con.close()
    return row[0] if row else ""


def save_annotation(user: str, case_no: str, text: str):
    con = _con()
    con.execute("INSERT OR REPLACE INTO annotations (username, case_no, notes) VALUES (?, ?, ?)", (user, case_no, text))
    con.commit()
    con.close()


def get_citing_cases(case_no: str) -> list[str]:
    con = _con()
    rows = con.execute("SELECT case_no FROM analyses WHERE json LIKE ?", (f"%{case_no}%",)).fetchall()
    con.close()
    return [r[0] for r in rows if r[0] != case_no]


def _jl(s):
    try:
        return json.loads(s) if s else []
    except Exception:
        return []


def case_meta(case_no: str) -> dict:
    con = _con()
    row = con.execute(
        "SELECT case_no, date, parties, judges, keywords, legislation, filename "
        "FROM judgements WHERE case_no=? LIMIT 1", (case_no,)
    ).fetchone()
    con.close()
    if not row:
        return {}
    return {"case_no": row[0], "date": row[1] or "", "parties": row[2] or "",
            "judges": _jl(row[3]), "keywords": _jl(row[4]), "legislation": _jl(row[5]), "filename": row[6]}


_BENCH_CACHE: dict[str, list[str]] = {}


def bench_for(case_no: str, meta: dict) -> list[str]:
    """Full panel of judges for a case.

    Fastest first: the reindexed `case_judges` table (scripts/reindex.py) gives
    the merged coram instantly. Otherwise parse the judgment live (front matter
    + signature block) and merge with the scrape's authoring judge. Cached per
    case_no so each judgment is parsed at most once per process."""
    if case_no in _BENCH_CACHE:
        return _BENCH_CACHE[case_no]
    from src.store import case_bench
    bench = case_bench(case_no)
    if not bench:
        parsed: list[str] = []
        fn = meta.get("filename")
        if fn:
            pdf = JUDGE_DIR / fn
            if not pdf.exists():
                pdf = LANKALAW_CASES_DIR / fn  # NLR/SLR report judgments
            if pdf.exists():
                try:
                    parsed = extract_bench(str(pdf))
                except Exception:
                    parsed = []
        # The scrape records only the *authoring* judge; the parsed coram (front
        # matter + signature block) records the concurring judges but sometimes
        # omits the author (who signs by role, not name). Merge both, surname-deduped,
        # so a 3- or 5-judge bench is complete even when each source alone is partial.
        meta_judges = [j for j in (meta.get("judges") or []) if str(j).strip()]
        bench = merge_benches(parsed, meta_judges) if parsed else meta_judges
        if not bench:
            # Last resort: judges the AI analysis extracted (e.g. 'SURASENA').
            ana = ((get_breakdown(case_no) or {}).get("metadata") or {}).get("judges") or []
            bench = [_norm_judge_name(j) for j in ana if str(j).strip()][:7]
            bench = [b for b in bench if b]
    if bench:  # never cache a miss — a transient parse hiccup would stick until restart
        _BENCH_CACHE[case_no] = bench
    return bench


def _norm_judge_name(j) -> str:
    s = re.sub(r"[,\s]*(?:A\.?C\.?J|C\.?J|JJ?)\.?\s*$", "", str(j).strip(" .,"))
    return f"{s.title()}, J." if s else ""


_AUTHOR_CACHE: dict[str, str] = {}


def author_for(case_no: str) -> str:
    """Authoring judge for law-report cases whose scrape metadata has none:
    parsed live from the opinion opener ('PATHIRANA, J.— …' / 'Per X, J.')."""
    if case_no in _AUTHOR_CACHE:
        return _AUTHOR_CACHE[case_no]
    from src.ingest import extract_opinion_author
    con = _con()
    rows = con.execute("SELECT text FROM chunks_fts WHERE case_no=? LIMIT 6", (case_no,)).fetchall()
    con.close()
    a = extract_opinion_author(" ".join(r[0] for r in rows))
    if a:
        _AUTHOR_CACHE[case_no] = a
    return a


# ---------------- shared LLM gate (one model, many visitors) --------------- #
# The local llama.cpp context is a single shared resource and each call holds it
# for 40-180 s. This gate caps how many requests may QUEUE for it at once, so a
# burst of /demo visitors can't pile up work for hours (cost/DoS guard — see
# docs/SECURITY.md). Per-session chat cooldowns are applied at the call sites.
import threading as _threading

_LLM_MAX_INFLIGHT = 3
_LLM_GATE_LOCK = _threading.Lock()
_LLM_INFLIGHT = 0


def _llm_acquire() -> bool:
    global _LLM_INFLIGHT
    with _LLM_GATE_LOCK:
        if _LLM_INFLIGHT >= _LLM_MAX_INFLIGHT:
            return False
        _LLM_INFLIGHT += 1
        return True


def _llm_release() -> None:
    global _LLM_INFLIGHT
    with _LLM_GATE_LOCK:
        _LLM_INFLIGHT = max(0, _LLM_INFLIGHT - 1)


# Cases currently being analysed, across ALL sessions (case_no -> started_at).
# Phones especially reload mid-generation (screen lock), so a per-session flag
# alone re-offers the Generate button and duplicates 2-minute model runs.
_BD_INFLIGHT: dict[str, float] = {}

# Paid-API guard: when ROSCRIBE_AI_USERS is set (comma-separated emails and/or
# usernames), only those accounts may run AI features (Generate, Ask ROS,
# Document Reference) — browsing, search and PDFs stay open to every signed-in
# user. Leave it empty to open AI to everyone.
_AI_USERS = {e.strip().lower() for e in os.getenv("ROSCRIBE_AI_USERS", "").split(",") if e.strip()}
_RO_USERS = {e.strip().lower() for e in os.getenv("ROSCRIBE_RO_USERS", "").split(",") if e.strip()}


def _ai_allowed() -> bool:
    username = _safe_username().lower()
    if username == "admin":
        return True
    if not _AI_USERS or username in _AI_USERS or username in _RO_USERS:
        return True
    try:
        ui.notify("AI analysis and chat are limited to approved accounts on this deployment — "
                  "ask the owner for access. Browsing and search remain fully open.",
                  type="warning")
    except Exception:
        pass
    return False


def get_breakdown_metadata(case_no: str) -> str | None:
    try:
        con = _con()
        row = con.execute("SELECT created_at FROM analyses WHERE case_no=?", (case_no,)).fetchone()
        con.close()
        return row[0] if row else None
    except Exception:
        return None


def get_breakdown(case_no: str):
    con = _con()
    row = con.execute("SELECT json FROM analyses WHERE case_no=?", (case_no,)).fetchone()
    con.close()
    if row:
        record_funnel_event("judgments_opened")
        return json.loads(row[0])
    return None


def cases_by_judge(name: str) -> list[dict]:
    con = _con()
    rows = con.execute(
        "SELECT case_no, date, parties FROM judgements WHERE judges LIKE ? ORDER BY date DESC LIMIT 200",
        (f"%{name}%",)).fetchall()
    con.close()
    return [{"case_no": r[0], "date": r[1] or "", "snippet": (r[2] or "")[:100]} for r in rows]


def cases_by_keyword(area: str) -> list[dict]:
    con = _con()
    rows = con.execute(
        "SELECT case_no, date, parties FROM judgements WHERE keywords LIKE ? ORDER BY date DESC LIMIT 200",
        (f"%{area}%",)).fetchall()
    con.close()
    return [{"case_no": r[0], "date": r[1] or "", "snippet": (r[2] or "")[:100]} for r in rows]


_JUSTICES = None
_AREAS = None
_BY_YEAR = None


def distinct_justices() -> list[str]:
    """Deduped justice display names — one option per justice (variant spellings
    merged in src.store.justices_grouped)."""
    global _JUSTICES
    if _JUSTICES is None:
        from src.store import distinct_justices as _store_distinct_justices
        _JUSTICES = _store_distinct_justices()
    return _JUSTICES


def legal_areas() -> list[str]:
    return list(LEGAL_AREAS)  # curated real legal areas (see src/store.py)


def judgements_by_year() -> dict[str, list]:
    global _BY_YEAR
    if _BY_YEAR is None:
        con = _con()
        # GROUP BY: a case with twin rows (two documents) appears once per year.
        rows = con.execute("SELECT case_no, max(date) FROM judgements GROUP BY case_no").fetchall()
        con.close()
        by: dict[str, list] = {}
        for cn, date in rows:
            if date and date[:4].isdigit():
                y = date[:4]
            else:
                yrs = [int(x) for x in re.findall(r"(?:19|20)\d{2}", cn or "") if 1950 <= int(x) <= 2027]
                y = str(max(yrs)) if yrs else "Undated"
            by.setdefault(y, []).append((cn, date or ""))
        _BY_YEAR = by
    return _BY_YEAR


_BY_YEAR_MONTH = None
_MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]


def judgements_by_year_month() -> dict[str, dict[str, list]]:
    """Court judgements only — law-report cases (NLR/SLR) live in their own
    tabs (see judgements_by_report), not in the year/month tree."""
    global _BY_YEAR_MONTH
    if _BY_YEAR_MONTH is None:
        con = _con()
        rows = con.execute(
            """SELECT case_no, max(date), max(COALESCE(report_cite,''))
               FROM judgements GROUP BY case_no""").fetchall()
        con.close()
        by: dict[str, dict[str, list]] = {}
        for cn, date, cite in rows:
            if cite:
                continue
            if date and date[:4].isdigit():
                y = date[:4]
            else:
                yrs = [int(x) for x in re.findall(r"(?:19|20)\d{2}", cn or "") if 1950 <= int(x) <= 2027]
                y = str(max(yrs)) if yrs else "Undated"

            m_name = "Unknown Month"
            if date and len(date) >= 7 and date[5:7].isdigit():
                m_idx = int(date[5:7]) - 1
                if 0 <= m_idx < 12:
                    m_name = _MONTH_NAMES[m_idx]

            by.setdefault(y, {}).setdefault(m_name, []).append((cn, date or ""))
        _BY_YEAR_MONTH = by
    return _BY_YEAR_MONTH


_BY_REPORT = None


_BY_ACT = None


def statutes_by_year() -> dict[str, dict[str, list]]:
    """Acts/statutes grouped for the 'Acts' tab: {year: {'Acts': [(statute_id, year)]}}
    — same nested shape make_year() consumes. Undated Acts shelve under 'Undated'."""
    global _BY_ACT
    if _BY_ACT is None:
        con = _con()
        rows = con.execute("SELECT statute_id, COALESCE(year,'') FROM statutes").fetchall()
        con.close()
        by: dict[str, dict[str, list]] = {}
        for sid, year in rows:
            y = year if (year.isdigit() and 1800 <= int(year) <= 2100) else "Undated"
            by.setdefault(y, {}).setdefault("Acts", []).append((sid, year))
        _BY_ACT = by
    return _BY_ACT


def judgements_by_report() -> dict[str, dict[str, dict[str, list]]]:
    """Law-report cases grouped for the NLR / SLR tabs:
    {'NLR': {'1976': {'76 NLR': [(case_no, date), …]}}, 'SLR': {'1982': {'SLR 1982': […]}}}.
    NLR volumes shelve under 1900 + volume (the conventional shorthand), with the
    volume itself ('76 NLR') as the sub-group; SLR years are the report years."""
    global _BY_REPORT
    if _BY_REPORT is None:
        con = _con()
        rows = con.execute(
            """SELECT case_no, max(date), max(COALESCE(report_cite,''))
               FROM judgements WHERE COALESCE(report_cite,'')!='' GROUP BY case_no""").fetchall()
        con.close()
        by: dict[str, dict[str, dict[str, list]]] = {"NLR": {}, "SLR": {}, "Digest": {}}
        for cn, date, cite in rows:
            m = re.match(r"(\d{1,3}) NLR$", cite)
            if m:
                series, year, sub = "NLR", str(1900 + int(m.group(1))), cite
            elif cite.startswith("SLR "):
                series, year, sub = "SLR", cite[4:], cite
            elif cite == "Digest":
                # lankalaw's A-Z case-digest volumes: shelve alphabetically.
                letter = cn[:1].upper() if cn[:1].isalpha() else "#"
                series, year, sub = "Digest", letter, "Digest entries"
            else:
                continue
            by[series].setdefault(year, {}).setdefault(sub, []).append((cn, date or ""))
        _BY_REPORT = by
    return _BY_REPORT


def _available_years() -> list[str]:
    """Years present in the corpus (newest first) — options for the Year filter."""
    return sorted((y for y in judgements_by_year() if y != "Undated"), reverse=True)


def get_connected_cases_in_corpus(case_no: str) -> list[str]:
    """Find other cases in the database that share the same parties (high party-name overlap)."""
    con = _con()
    row = con.execute("SELECT parties FROM judgements WHERE case_no=? LIMIT 1", (case_no,)).fetchone()
    if not row or not row[0]:
        con.close()
        return []
    
    ignored = {'vs', 'v', 'and', 'or', 'the', 'of', 'in', 're', 'matter', 'appellant', 'respondent', 'petitioner', 'defendant', 'plaintiff', 'both', 'are', 'no', 'road', 'court', 'appeal', 'sri', 'lanka'}
    words = re.findall(r'\b[a-zA-Z]{3,}\b', row[0])
    unique_words = [w for w in words if w.lower() not in ignored]
    unique_words.sort(key=len, reverse=True)
    key_parties = unique_words[:3]
    
    if not key_parties:
        con.close()
        return []
        
    rows = con.execute("SELECT case_no, parties FROM judgements WHERE case_no != ?", (case_no,)).fetchall()
    con.close()
    
    connected = []
    for cn, parties_text in rows:
        if not parties_text:
            continue
        score = sum(1 for w in key_parties if w.lower() in parties_text.lower())
        if score >= 2:
            connected.append(cn)
            
    return connected


def find_case_by_parties_and_date(cited: str, active_case_no: str = None) -> str | None:
    """Attempt to find a case in the corpus matching the cited citation by looking at
    surnames/party names and the year of decision/citation. Requires matching names from
    both sides of the 'v' or 'vs' separator to prevent false positive matches."""
    # 1. Extract the year from the cited text (e.g. 2013, 2016)
    year_match = re.search(r'\b(20\d{2}|19\d{2})\b', cited)
    year = year_match.group(1) if year_match else None
    
    # 2. Extract potential party names from the citation itself (if it contains text like 'v.' or 'vs')
    cited_parties = []
    left_parties = []
    right_parties = []
    if ' v ' in cited.lower() or ' vs ' in cited.lower() or ' v. ' in cited.lower() or ' vs. ' in cited.lower():
        parts = re.split(r'\s+v(?:s)?\.?\s+', cited, flags=re.IGNORECASE)
        if len(parts) >= 2:
            ignored_words = {'court', 'appeal', 'sri', 'lanka', 'state', 're', 'ex', 'parte', 'attorney', 'general', 'and', 'the', 'of', 'in', 'no'}
            left_words = [w.strip() for w in re.findall(r'\b[a-zA-Z]{3,}\b', parts[0]) if w.lower() not in ignored_words]
            right_words = [w.strip() for w in re.findall(r'\b[a-zA-Z]{3,}\b', parts[1]) if w.lower() not in ignored_words]
            left_parties = left_words
            right_parties = right_words
            cited_parties = left_words + right_words
                
    # 3. If no party names were in the citation (e.g. it's a raw case number),
    # and we have an active case context, use the active case's parties
    if not cited_parties and active_case_no:
        con = _con()
        row = con.execute("SELECT parties FROM judgements WHERE case_no=? LIMIT 1", (active_case_no,)).fetchone()
        con.close()
        if row and row[0]:
            ignored = {'vs', 'v', 'and', 'or', 'the', 'of', 'in', 're', 'matter', 'appellant', 'respondent', 'petitioner', 'defendant', 'plaintiff', 'both', 'are', 'no', 'road', 'court', 'appeal', 'sri', 'lanka'}
            words = re.findall(r'\b[a-zA-Z]{3,}\b', row[0])
            unique_words = [w for w in words if w.lower() not in ignored]
            unique_words.sort(key=len, reverse=True)
            cited_parties = unique_words[:3]

    if not cited_parties:
        return None

    con = _con()
    if year:
        rows = con.execute("SELECT case_no, parties, date FROM judgements WHERE date LIKE ?", (f"%{year}%",)).fetchall()
    else:
        rows = con.execute("SELECT case_no, parties, date FROM judgements").fetchall()
    con.close()

    best_match = None
    max_score = 0
    for cn, parties_text, date_text in rows:
        if not parties_text:
            continue
            
        # Require BOTH left side and right side to match if we have them
        if left_parties and right_parties:
            match_left = any(re.search(rf"\b{re.escape(w.lower())}\b", parties_text.lower()) for w in left_parties)
            match_right = any(re.search(rf"\b{re.escape(w.lower())}\b", parties_text.lower()) for w in right_parties)
            if not (match_left and match_right):
                continue
                
        score = sum(1 for w in cited_parties if re.search(rf"\b{re.escape(w.lower())}\b", parties_text.lower()))
        if score > max_score and score >= 1:
            max_score = score
            best_match = cn

    return best_match


def find_case(cited: str, active_case_no: str = None):
    """Resolve a precedent citation to a corpus case_no using the case number,
    falling back to a parties and decided date check for connected cases."""
    res = resolve_citation(cited)
    if res:
        return res
    return find_case_by_parties_and_date(cited, active_case_no)


def _safe_username() -> str:
    """Logged-in username, tolerant of a dropped/changed client session.

    A breakdown runs ~1-2 min via run.io_bound; if the websocket reconnects in
    that window the per-session user storage is gone and reading it raises
    'user storage ... should be created before accessing it'. Falling back to
    'anonymous' keeps the post-await re-render from crashing — which is what left
    the 'Analysing…' spinner stuck forever."""
    try:
        return app.storage.user.get("username", "anonymous")
    except Exception:
        return "anonymous"


def _web_search_url(query: str) -> str:
    """A Google search URL — the open-web fallback for any citation, Act, or
    constitutional article not in the local corpus, so it can still be looked up
    and cited from the original source."""
    from urllib.parse import quote_plus
    return "https://www.google.com/search?q=" + quote_plus((query or "").strip())


HEAD_CSS = """
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap" rel="stylesheet">
<style>
  /* Theme: inspired by rechtspraak.nl / the Dutch Rijkshuisstijl —
     hemelblauw #01689b (primary), donkerblauw #154273 (headings), and a
     light governmental gray page. Flat, squared, generous white space. */
  body {
    background: #faf8f4;
    font-family: 'Plus Jakarta Sans', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  }
  body.workspace-mode {
    overflow: hidden !important;
  }
  body.workspace-mode .q-page {
    height: 100vh !important;
    overflow: hidden !important;
  }
  .case-title {
    font-family: 'Plus Jakarta Sans', sans-serif;
  }
  .pane-head {
    font-size: 0.75rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #82212b; /* dark red */
    font-weight: 700;
    margin-bottom: 4px;
  }
  .doc-pane {
    background: #ffffff;
  }
  .sec {
    font-family: 'Plus Jakarta Sans', sans-serif;
    font-size: 0.85rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: #82212b; /* dark red heading */
    border-left: 3px solid #82212b; /* dark red accent */
    padding-left: 8px;
    margin-top: 1.2rem;
    margin-bottom: 0.4rem;
  }
  .body-text {
    font-family: 'Lora', Georgia, serif;
    font-size: 0.92rem;
    line-height: 1.6;
    color: #3c4043;
  }
  /* Rendered markdown (What ROS says, Document Reference) must not balloon:
     constrain every child element to body size regardless of ui.markdown's
     default h1-h6/p typography. */
  .synthesis-md { font-family: 'Lora', Georgia, serif; color: #3c4043; }
  .synthesis-md :is(h1,h2,h3,h4,h5,h6) {
    font-size: 0.86rem !important;
    font-weight: 700 !important;
    margin: 0.7em 0 0.2em !important;
    line-height: 1.4 !important;
    color: #202124;
  }
  .synthesis-md :is(p,li) {
    font-size: 0.92rem !important;
    line-height: 1.6 !important;
    margin: 0 0 0.5em !important;
  }
  .synthesis-md ul { padding-left: 1.1rem; margin: 0.2em 0 0.6em; }
  .chip {
    cursor: pointer;
    border-radius: 6px;
    font-family: 'Plus Jakarta Sans', sans-serif;
  }
  .q-item:hover {
    background: #f1f3f4;
  }
  /* Google style scrollbar */
  ::-webkit-scrollbar {
    width: 6px;
    height: 6px;
  }
  ::-webkit-scrollbar-track {
    background: transparent;
  }
  ::-webkit-scrollbar-thumb {
    background: #dadce0;
    border-radius: 10px;
  }
  ::-webkit-scrollbar-thumb:hover {
    background: #bdc1c6;
  }
  /* Dark Mode specific overrides */
  .body--dark {
    background: #1b2027 !important;
    color: #e8eaed !important;
  }
  .body--dark .sec {
    color: #ff4d4d !important;
  }
  .body--dark .pane-head {
    color: #9aa0a6 !important;
  }
  .body--dark .body-text {
    color: #e8eaed !important;
  }
  .body--dark ::-webkit-scrollbar-thumb {
    background: #3c4043;
  }
  .body--dark ::-webkit-scrollbar-thumb:hover {
    background: #5f6368;
  }
  .body--dark .bg-white {
    background-color: #1e1e1e !important;
  }
  .body--dark .bg-gray-100 {
    background-color: #1b2027 !important;
  }
  .body--dark .bg-gray-50 {
    background-color: #2c2c2c !important;
  }
  .body--dark .border {
    border-color: #2c2c2c !important;
  }
  .body--dark .text-gray-900 {
    color: #f1f3f4 !important;
  }
  .body--dark .text-gray-800 {
    color: #e8eaed !important;
  }
  .body--dark .text-gray-700 {
    color: #dadce0 !important;
  }
  .body--dark .text-gray-600 {
    color: #bdc1c6 !important;
  }
  .body--dark .text-gray-500 {
    color: #9aa0a6 !important;
  }
  /* responsive: 3 panes side-by-side on desktop; one at a time + tab bar on mobile */
  .mobile-tabs {
    display: none;
  }
  .panes-row {
    /* Clear the fixed 56px header at the top and sit just above the 30px footer.
       margin: top clears header (56+8), 16px on the sides/bottom. Height fills
       the gap between so cards are never clipped and no dead space remains. */
    height: calc(100vh - 76px - 10px - 30px) !important;
    margin: 76px 16px 10px 16px !important;
    width: calc(100% - 32px) !important;
    align-items: stretch !important;
    background: transparent !important;
  }
  .q-splitter, .q-splitter__panels {
    height: 100% !important;
  }
  .q-splitter__panel {
    height: 100% !important;
    display: flex !important;
    flex-direction: column !important;
  }
  @media (max-width: 900px) {
    .mobile-tabs {
      display: flex;
      position: sticky !important;
      top: 56px !important;
      z-index: 999 !important;
      background-color: #ffffff !important;
      box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05) !important;
    }
    .iframe-pdf {
      display: none !important;
    }
    .mobile-download-btn {
      display: flex !important;
    }
    .panes-row {
      margin: 12px !important;
      height: calc(100dvh - 56px - 30px - 24px) !important;
      width: calc(100% - 24px) !important;
      padding: 0px !important;
      gap: 12px !important;
      align-items: stretch !important;
      background: transparent !important;
    }
    #pdf-pane, #breakdown-pane {
      display: none !important;
      width: 100% !important;
    }
    #pdf-pane.active, #breakdown-pane.active {
      display: flex !important;
    }
    /* Mobile: override explicit height for download card area */
    #pdf-pane .mobile-download-btn {
      min-height: calc(100vh - 220px) !important;
      height: calc(100vh - 220px) !important;
    }
    /* Bigger, clearer mobile tab bar. */
    .mobile-tabs .q-btn {
      min-height: 44px;
      font-size: 0.72rem;
      border-bottom: 2px solid transparent;
    }
    .hide-narrow {
      display: none !important;
    }
    /* Mobile splitter override */
    .q-splitter {
      display: block !important;
      height: 100% !important;
    }
    .q-splitter__panel {
      width: 100% !important;
      height: 100% !important;
      display: none !important;
    }
    .q-splitter.show-before .q-splitter__before {
      display: flex !important;
      flex-direction: column !important;
    }
    .q-splitter.show-after .q-splitter__after {
      display: flex !important;
      flex-direction: column !important;
    }
    .q-splitter__separator {
      display: none !important;
    }
  }

  @media (min-width: 901px) {
      /* Smart Drawer layout overrides as a floating sidebar below the 56px header */
      .smart-drawer {
        top: 56px !important;
        height: calc(100vh - 56px) !important;
        width: 400px !important;
        max-width: 90vw !important;
        transition: transform 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        box-shadow: 4px 0 20px rgba(0,0,0,0.06) !important;
        border-right: 1px solid #b9cad9 !important;
        z-index: 1000 !important;
        background-color: #e8eff5 !important;
      }
      .smart-drawer.drawer-closed {
        transform: translateX(-100%) !important;
      }
  }
  
  /* High contrast sidebar accordion cards and select elements */
  .smart-drawer .q-expansion-item {
    background-color: transparent !important;
    border: none !important;
    margin-bottom: 6px !important;
    box-shadow: none !important;
  }
  /* Top-level expansion item header capsule styling */
  .smart-drawer > div > .q-expansion-item > .q-expansion-item__container > .q-item {
    background-color: #dbeafe !important;
    color: #1e3a8a !important;
    border-radius: 8px !important;
    border: 1px solid #bfdbfe !important;
    font-weight: 700 !important;
  }
  /* Sub-level expansion item header capsule styling */
  .smart-drawer .q-expansion-item .q-expansion-item .q-item {
    background-color: #f1f5f9 !important;
    color: #334155 !important;
    border-radius: 6px !important;
    margin: 4px 8px !important;
    border: 1px solid #e2e8f0 !important;
    font-weight: 600 !important;
  }
  .smart-drawer .q-field--outlined .q-field__control {
    background-color: #ffffff !important;
    border-radius: 8px !important;
    border-color: #cbd5e1 !important;
  }
  .smart-drawer .q-field--outlined .q-field__control:before {
    border: 1px solid #cbd5e1 !important;
  }
  .smart-drawer .q-field--focused .q-field__control:after {
    border-color: #1e3a8a !important;
  }
  
  /* Tint backdrop starting below the header and blurring the content */
  .q-drawer__backdrop {
    top: 56px !important;
    height: calc(100vh - 56px) !important;
    background: rgba(15, 23, 42, 0.45) !important;
    backdrop-filter: blur(2px) !important;
    z-index: 999 !important;
  }
  
  /* Gemini-style animations */
  @keyframes geminiShimmer {
    0% { background-position: -200% 0; }
    100% { background-position: 200% 0; }
  }
  
  .gemini-skeleton-flow {
    background: linear-gradient(90deg, 
      rgba(66, 133, 244, 0.12) 0%, 
      rgba(155, 81, 224, 0.12) 25%, 
      rgba(233, 30, 99, 0.12) 50%, 
      rgba(155, 81, 224, 0.12) 75%, 
      rgba(66, 133, 244, 0.12) 100%
    );
    background-size: 200% 100%;
    animation: geminiShimmer 2.5s infinite linear;
    border-radius: 4px;
    display: inline-block;
  }
  
  .gemini-progress .q-linear-progress__track {
    opacity: 0.15 !important;
  }
  .gemini-progress .q-linear-progress__bar {
    background: linear-gradient(90deg, #4285f4, #9b51e0, #e91e63, #4285f4) !important;
    background-size: 200% 100% !important;
    animation: geminiShimmer 2s infinite linear !important;
  }
  @keyframes ask-ross-glow {
    0% {
      box-shadow: 0 4px 16px rgba(140, 29, 29, 0.4), 0 0 0 0 rgba(140, 29, 29, 0.5);
    }
    70% {
      box-shadow: 0 4px 16px rgba(140, 29, 29, 0.2), 0 0 0 12px rgba(140, 29, 29, 0);
    }
    100% {
      box-shadow: 0 4px 16px rgba(140, 29, 29, 0.4), 0 0 0 0 rgba(140, 29, 29, 0);
    }
  }
  .ask-ross-btn-glowing {
    animation: ask-ross-glow 2s infinite ease-in-out !important;
  }
  .ask-ross-btn-close {
    box-shadow: 0 4px 16px rgba(0,0,0,0.2) !important;
  }
  @keyframes dotsLoop {
    0% { width: 0; }
    25% { width: 0.25em; }
    50% { width: 0.6em; }
    75% { width: 0.9em; }
    100% { width: 1.25em; }
  }
  .loading-dots::after {
    content: "...";
    display: inline-block;
    width: 0px;
    overflow: hidden;
    vertical-align: bottom;
    animation: dotsLoop 1.5s infinite steps(4);
  }
  .q-chip {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    vertical-align: middle !important;
  }
  .q-chip__content {
    display: inline-flex !important;
    align-items: center !important;
    justify-content: center !important;
    line-height: 1 !important;
    padding: 0 !important;
  }
  /* User chat bubble: Sleek dark slate navy, white text */
  .user-message-bubble .q-message-text {
    background: #0f172a !important;
  }
  .user-message-bubble .q-message-text-content {
    color: #ffffff !important;
  }
  .user-message-bubble .q-message-text:before,
  .user-message-bubble .q-message-text--sent:before,
  .user-message-bubble .q-message-text--sent:last-child:before {
    border-bottom-color: #0f172a !important;
    border-top-color: #0f172a !important;
  }
  /* Assistant chat bubble: Clean white background, slate grey text */
  .assistant-message-bubble .q-message-text {
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
  }
  .assistant-message-bubble .q-message-text-content {
    color: #334155 !important;
  }
  .assistant-message-bubble .q-message-text:before,
  .assistant-message-bubble .q-message-text--received:before,
  .assistant-message-bubble .q-message-text--received:last-child:before {
    border-bottom-color: #e2e8f0 !important;
    border-top-color: #e2e8f0 !important;
  }
  .q-message-text {
    font-size: 0.8rem !important;
    line-height: 1.4 !important;
  }
  .q-message-text :is(p, li, span, a) {
    font-size: 0.8rem !important;
    line-height: 1.4 !important;
  }
  .q-message-text :is(h1, h2, h3, h4, h5, h6) {
    font-size: 0.85rem !important;
    font-weight: 700 !important;
    margin-top: 8px !important;
    margin-bottom: 4px !important;
    line-height: 1.3 !important;
  }
  .q-message-text blockquote {
    margin: 6px 0 !important;
    padding-left: 8px !important;
    border-left: 2px solid currentColor !important;
    font-size: 0.8rem !important;
    opacity: 0.9;
  }
  .q-message-text p {
    margin: 0 !important;
  }
  .q-message-text p + p {
    margin-top: 8px !important;
  }
  .q-message-text ul, .q-message-text ol {
    margin: 4px 0 !important;
    padding-left: 20px !important;
  }
</style>
""";

_TREAT_COLOR = {"Distinguished": "orange", "Overruled": "red", "Applied": "green", "Followed": "green"}


def build_workspace(demo: bool = False, initial_case: str = None):
    record_funnel_event("portal_hits")
    _tokens = (Path(__file__).parent / "design_tokens.css").read_text()
    ui.add_head_html(f"<style>{_tokens}</style>")
    ui.add_head_html(HEAD_CSS)
    ui.add_head_html(STATUS_SCRIPT)
    # ROS design system: navy leads, heritage crimson as accent, brass highlight.
    ui.colors(primary="#16273f", secondary="#82212b", accent="#b8965a")
    ui.dark_mode().disable()
    state = {"case": None, "statute": None, "page": None, "chat_open": False, "workspace_active": False}
    containers = {"bookmarks": None}
    
    # UI layout component references for left drawer
    left_drawer = None
    drawer_container = None
    menu_btn = None
    drawer_close_btn = None
    back_btn = None

    def settings_dialog():
        from src.config import settings as s_settings
        is_admin = (app.storage.user.get("username") or "").lower() == "admin"
        
        with ui.dialog() as dialog, ui.card().classes("w-96 p-6"):
            ui.label("AI Settings").classes("text-lg font-bold text-gray-800 mb-2")
            
            # Provider Selection
            current_provider = app.storage.user.get("llm_provider") or s_settings.llm_provider
            if current_provider.lower() == "anthropic":
                current_provider = "Anthropic"
            elif current_provider.lower() == "openai":
                current_provider = "OpenAI"
            elif current_provider.lower() == "ollama":
                current_provider = "Ollama"
            elif current_provider.lower() == "llamacpp":
                current_provider = "Llama.cpp"

            provider_select = ui.select(["Anthropic", "OpenAI", "Ollama", "Llama.cpp"], label="LLM Provider", value=current_provider).classes("w-full mb-3")
            
            # Model Selection
            current_model = app.storage.user.get("llm_model") or (s_settings.anthropic_model if current_provider == "Anthropic" else s_settings.llm_model)
            model_input = ui.input("Model Name", placeholder="e.g. claude-3-5-sonnet-20241022", value=current_model).classes("w-full mb-3")
            
            # Custom API Key input
            current_api_key = app.storage.user.get("custom_api_key") or ""
            api_key_input = ui.input("Custom API Key (Optional)", value=current_api_key, password=True, password_toggle_button=True).classes("w-full mb-4")
            
            # If not administrator, block input changes completely
            if not is_admin:
                provider_select.props("disable")
                model_input.props("disable")
                api_key_input.props("disable")
                ui.label("⚠️ AI setting adjustments are restricted to admin accounts.").classes("text-[10px] text-red-500 font-bold mb-2 uppercase tracking-wider")
            
            def on_provider_change():
                if not is_admin:
                    return
                p = provider_select.value
                if p == "Anthropic":
                    model_input.value = s_settings.anthropic_model or "claude-3-5-sonnet-20241022"
                    api_key_input.set_placeholder("Enter Anthropic API Key")
                    api_key_input.props("disable=false")
                elif p == "OpenAI":
                    model_input.value = "gpt-4o-mini"
                    api_key_input.set_placeholder("Enter OpenAI API Key")
                    api_key_input.props("disable=false")
                elif p == "Ollama":
                    model_input.value = "qwen2.5-1.5b-instruct"
                    api_key_input.value = ""
                    api_key_input.set_placeholder("Not required for local Ollama")
                    api_key_input.props("disable")
                elif p == "Llama.cpp":
                    model_input.value = "local-gguf"
                    api_key_input.value = ""
                    api_key_input.set_placeholder("Not required for local Llama.cpp")
                    api_key_input.props("disable")
            
            provider_select.on("change", on_provider_change)
            
            def save():
                p = provider_select.value
                p_val = "anthropic" if p == "Anthropic" else "openai" if p == "OpenAI" else "ollama" if p == "Ollama" else "llamacpp"
                app.storage.user["llm_provider"] = p_val
                app.storage.user["llm_model"] = model_input.value
                app.storage.user["custom_api_key"] = api_key_input.value
                dialog.close()
                ui.notify("AI settings saved successfully!", type="positive")
                
            def reset():
                app.storage.user.pop("llm_provider", None)
                app.storage.user.pop("llm_model", None)
                app.storage.user.pop("custom_api_key", None)
                dialog.close()
                ui.notify("Reset to system defaults.", type="info")

            with ui.row().classes("w-full justify-between mt-2"):
                reset_btn = ui.button("Reset Defaults", on_click=reset, color="grey").classes("flat dense")
                save_btn = ui.button("Save", on_click=save, color="primary").classes("px-4")
                if not is_admin:
                    reset_btn.props("disable")
                    save_btn.props("disable")
        
        dialog.open()

    def admin_dialog():
        import re
        from src.analyze_smart import LLM_METRICS
        
        def save_users_to_env():
            raw_val = ",".join(f"{u}:{p}" for u, p in USERS.items())
            raw_ai = ",".join(sorted(list(_AI_USERS)))
            raw_ro = ",".join(sorted(list(_RO_USERS)))
            for filename in (".env_smart", ".env"):
                filepath = REPO_ROOT / filename
                if not filepath.exists():
                    continue
                try:
                    content = filepath.read_text(encoding="utf-8")
                    new_lines = []
                    has_ro = False
                    for line in content.splitlines():
                        if line.startswith("ROSCRIBE_USERS="):
                            new_lines.append(f"ROSCRIBE_USERS={raw_val}")
                        elif line.startswith("ROSCRIBE_AI_USERS="):
                            new_lines.append(f"ROSCRIBE_AI_USERS={raw_ai}")
                        elif line.startswith("ROSCRIBE_RO_USERS="):
                            new_lines.append(f"ROSCRIBE_RO_USERS={raw_ro}")
                            has_ro = True
                        else:
                            new_lines.append(line)
                    if not has_ro:
                        new_lines.append(f"ROSCRIBE_RO_USERS={raw_ro}")
                    filepath.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
                except Exception as e:
                    print(f"Error saving to {filename}: {e}")

        def get_parsed_logins():
            log_path = REPO_ROOT / "data" / "auth.log"
            if not log_path.exists():
                return []
            logins = []
            try:
                with open(log_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        m_google = re.search(r"Google login:\s+([^\s]+)(?:\s+\(([^)]+)\))?\s+via\s+([^\s]+)\s+at\s+(.+)$", line)
                        m_google_oauth = re.search(r"User\s+([^\s]+)\s+logged\s+in\s+successfully\s+via\s+Google OAuth\s+at\s+(.+)$", line)
                        m_pass = re.search(r"User\s+([^\s]+)\s+logged\s+in\s+successfully\s+via\s+password\s+at\s+(.+)$", line)
                        if m_google:
                            logins.append({
                                "time": m_google.group(4),
                                "type": "Google",
                                "user": f"{m_google.group(2) or 'Google User'} ({m_google.group(1)})",
                                "source": m_google.group(3),
                                "status": "Success"
                            })
                        elif m_google_oauth:
                            logins.append({
                                "time": m_google_oauth.group(2),
                                "type": "Google",
                                "user": m_google_oauth.group(1),
                                "source": "OAuth Callback",
                                "status": "Success"
                            })
                        elif m_pass:
                            logins.append({
                                "time": m_pass.group(2),
                                "type": "Password",
                                "user": m_pass.group(1),
                                "source": "Internal Portal",
                                "status": "Success"
                            })
                        else:
                            logins.append({
                                "time": "-",
                                "type": "Event",
                                "user": line,
                                "source": "System",
                                "status": "Info"
                            })
            except Exception as e:
                logins.append({
                    "time": "-",
                    "type": "Error",
                    "user": f"Log read error: {e}",
                    "source": "Internal",
                    "status": "Fail"
                })
            return list(reversed(logins))[:50]

        active_online = sum(1 for storage in app.storage._users.values() if storage.get("authenticated", False))
        active_online = max(1, active_online)

        with ui.dialog() as dialog, ui.card().style("width: 1100px; max-width: 95vw; height: 85vh; padding: 0; display: flex; flex-direction: column; background-color: #f8fafc; border: 1px solid #e2e8f0; overflow: hidden; border-radius: 24px;"):
            # Header
            with ui.row().classes("w-full bg-slate-900 text-white p-4 items-center justify-between no-wrap").style("flex-shrink: 0; border-top-left-radius: 24px; border-top-right-radius: 24px;"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("admin_panel_settings", color="red").classes("text-2xl")
                    with ui.column().classes("gap-0"):
                        ui.label("Admin Console").classes("text-sm font-bold tracking-wider text-white uppercase")
                        ui.label("Management & System Performance Portal").classes("text-[10px] text-slate-400")

                with ui.row().classes("items-center gap-6 hide-narrow text-xs"):
                    sys_metrics = get_cumulative_metrics()
                    with ui.column().classes("items-end gap-0"):
                        ui.label("Cumulative API Cost").classes("text-[9px] uppercase tracking-wider text-slate-400 font-semibold")
                        ui.label(f"${sys_metrics['api_cost']:.2f}").classes("text-sm font-bold text-amber-400")
                    with ui.column().classes("items-end gap-0"):
                        ui.label("Total Tokens").classes("text-[9px] uppercase tracking-wider text-slate-400 font-semibold")
                        ui.label(f"{sys_metrics['total_tokens']:,}").classes("text-sm font-bold text-blue-300")
                    with ui.column().classes("items-end gap-0"):
                        ui.label("LLM Queries").classes("text-[9px] uppercase tracking-wider text-slate-400 font-semibold")
                        ui.label(str(sys_metrics['query_count'])).classes("text-sm font-bold text-indigo-300")
                    with ui.column().classes("items-end gap-0"):
                        ui.label("Active Users").classes("text-[9px] uppercase tracking-wider text-slate-400 font-semibold")
                        with ui.row().classes("items-center gap-1.5"):
                            ui.element("div").classes("w-2 h-2 bg-green-500 rounded-full animate-pulse")
                            ui.label(f"{active_online} Online").classes("text-xs font-bold text-green-400")

            # Active view state
            dialog_state = {"view": "menu"}

            # Main content container
            content_container = ui.column().classes("w-full p-0 bg-slate-100").style("flex-grow: 1; height: 100%; overflow-y: auto;")
            
            with content_container:
                @ui.refreshable
                def render_admin_content():
                    view_name = dialog_state["view"]
                    if view_name == "menu":
                        with ui.grid().classes("w-full grid-cols-2 gap-6 p-6"):
                            # Card 1: Logins & Audit
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "logs"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("history", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("Logins & Audit History").classes("text-sm font-bold text-slate-800")
                                        ui.label("View real-time login logs and Google authentication audits.").classes("text-xs text-slate-500")

                            # Card 2: Users Console
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "users"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("people", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("Users Console").classes("text-sm font-bold text-slate-800")
                                        ui.label("Manage researcher roles and create database accounts.").classes("text-xs text-slate-500")

                            # Card 3: Funnel Analytics
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "funnel"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("analytics", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("Funnel Analytics").classes("text-sm font-bold text-slate-800")
                                        ui.label("Track conversion rates and user behavior metrics.").classes("text-xs text-slate-500")

                            # Card 4: Models Controller
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "models"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("memory", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("AI Models Controller").classes("text-sm font-bold text-slate-800")
                                        ui.label("Monitor cache state and load local models onto hardware.").classes("text-xs text-slate-500")

                            # Card 5: Credentials
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "credentials"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("vpn_key", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("System Credential Manager").classes("text-sm font-bold text-slate-800")
                                        ui.label("Manage LLM API keys and remote server endpoint URLs.").classes("text-xs text-slate-500")

                            # Card 6: Corpus addition
                            with ui.card().classes("p-6 bg-white border border-slate-200 shadow-sm hover:border-blue-500 hover:shadow-md cursor-pointer transition-all duration-200").on("click", lambda: (dialog_state.update({"view": "corpus"}), render_admin_content.refresh())):
                                with ui.row().classes("items-center gap-4 no-wrap"):
                                    ui.icon("cloud_upload", color="primary").classes("text-3xl")
                                    with ui.column().classes("gap-0"):
                                        ui.label("Corpus Management").classes("text-sm font-bold text-slate-800")
                                        ui.label("Upload judgment PDFs to dynamically index legal case files.").classes("text-xs text-slate-500")
                    
                    else:
                        with ui.column().classes("w-full p-6 gap-4"):
                            with ui.row().classes("w-full items-center justify-between border-b border-slate-200 pb-3 mb-2"):
                                ui.button("Back to Menu", icon="arrow_back", color="primary", on_click=lambda: (dialog_state.update({"view": "menu"}), render_admin_content.refresh())).props("flat dense").classes("text-xs font-semibold")
                                ui.badge(view_name.upper(), color="blue").classes("px-2.5 py-1 text-[10px] font-bold font-mono")
                            
                            if view_name == "logs":
                                ui.label("Audit Log: System Login History").classes("text-sm font-bold text-slate-800 mb-2")
                                ui.label("Dynamically parsed events from data/auth.log (newest first).").classes("text-xs text-slate-500 mb-4")
                                
                                log_entries = get_parsed_logins()
                                with ui.column().classes("w-full gap-2 border border-slate-200 rounded-lg overflow-hidden bg-slate-50 p-2"):
                                    with ui.row().classes("w-full text-[10px] text-slate-400 font-bold px-3 py-1 border-b border-slate-200"):
                                        ui.label("TIMESTAMP").classes("w-32")
                                        ui.label("METHOD").classes("w-24")
                                        ui.label("USER IDENTITY").classes("w-64")
                                        ui.label("SOURCE IP").classes("w-40")
                                        ui.label("STATUS").classes("w-20")

                                    for entry in log_entries:
                                        with ui.row().classes("w-full text-xs text-slate-700 items-center px-3 py-2 border-b border-slate-100 hover:bg-slate-200/40 rounded"):
                                            ui.label(entry["time"]).classes("w-32 text-slate-500 font-mono")
                                            badge_color = "blue" if entry["type"] == "Google" else "purple" if entry["type"] == "Password" else "grey"
                                            ui.badge(entry["type"], color=badge_color).classes("w-24 px-2 py-0.5 text-[9px] font-bold rounded")
                                            ui.label(entry["user"]).classes("w-64 font-medium truncate")
                                            ui.label(entry["source"]).classes("w-40 font-mono text-slate-500 truncate")
                                            ui.badge(entry["status"], color="green" if entry["status"] == "Success" else "amber").classes("w-20 px-2 py-0.5 text-[9px] font-bold rounded")

                            elif view_name == "users":
                                ui.label("User Creation & Database Accounts").classes("text-sm font-bold text-slate-800 mb-2")
                                
                                def perform_create():
                                    u = (new_user.value or "").strip()
                                    p = (password_input.value or "").strip()
                                    role_val = new_role.value
                                    if not u or not p:
                                        ui.notify("Username and Password are required", color="warning")
                                        return
                                    USERS[u] = p
                                    if role_val == "Admin":
                                        _AI_USERS.discard(u.lower())
                                        _RO_USERS.discard(u.lower())
                                    elif role_val == "RO":
                                        _RO_USERS.add(u.lower())
                                        _AI_USERS.discard(u.lower())
                                    elif role_val == "Legal Researcher":
                                        _AI_USERS.add(u.lower())
                                        _RO_USERS.discard(u.lower())
                                    else:
                                        _AI_USERS.discard(u.lower())
                                        _RO_USERS.discard(u.lower())
                                    save_users_to_env()
                                    ui.notify(f"User '{u}' created successfully", color="positive")
                                    render_admin_content.refresh()

                                def perform_delete(username_to_delete):
                                    if username_to_delete.lower() == "admin":
                                        ui.notify("Cannot delete default admin account", color="warning")
                                        return
                                    for sess_id, storage in list(app.storage._users.items()):
                                        if storage.get("username") == username_to_delete:
                                            storage["authenticated"] = False
                                    USERS.pop(username_to_delete, None)
                                    _AI_USERS.discard(username_to_delete.lower())
                                    save_users_to_env()
                                    ui.notify(f"User '{username_to_delete}' deleted and sessions terminated", color="positive")
                                    render_admin_content.refresh()

                                def perform_force_logout(username_to_logout):
                                    logged_out = False
                                    for sess_id, storage in list(app.storage._users.items()):
                                        if storage.get("username") == username_to_logout:
                                            storage["authenticated"] = False
                                            logged_out = True
                                    if logged_out:
                                        ui.notify(f"Terminated all active sessions for '{username_to_logout}'", color="positive")
                                    else:
                                        ui.notify(f"User '{username_to_logout}' is not currently online", color="info")
                                    render_admin_content.refresh()

                                with ui.row().classes("w-full gap-6 no-wrap items-start"):
                                    with ui.card().classes("w-1/2 p-4 bg-white border border-slate-200 shadow-sm"):
                                        ui.label("Configured Accounts").classes("text-xs font-bold uppercase tracking-wider text-slate-400 mb-3")
                                        for u, pw in USERS.items():
                                            role = "Administrator" if u.lower() == "admin" else "RO" if u.lower() in _RO_USERS else "Legal Researcher" if u.lower() in _AI_USERS else "Guest View"
                                            is_online = any(storage.get("username") == u and storage.get("authenticated", False) for storage in app.storage._users.values())
                                            
                                            m_data = get_user_metrics(u)
                                            q_cnt = m_data["query_count"]
                                            t_cnt = m_data["total_tokens"]
                                            
                                            with ui.row().classes("w-full justify-between items-center py-2 border-b border-slate-100 no-wrap"):
                                                with ui.column().classes("gap-0"):
                                                    with ui.row().classes("items-center gap-2"):
                                                        ui.label(u).classes("text-xs font-bold text-slate-800")
                                                        if is_online:
                                                            ui.badge("ONLINE", color="green").classes("text-[8px] px-1 py-0")
                                                    ui.label(f"Role: {role}").classes("text-[9px] text-slate-500")
                                                    ui.label(f"Queries: {q_cnt} | Tokens: {t_cnt:,}").classes("text-[9px] font-semibold text-blue-500 font-mono mt-0.5")
                                                    
                                                    # Password visibility toggle
                                                    with ui.row().classes("items-center gap-1.5 mt-0.5 no-wrap"):
                                                        is_rev = u in dialog_state.setdefault("revealed", set())
                                                        pw_display = pw if is_rev else "••••••••"
                                                        ui.label(f"Pass: {pw_display}").classes("text-[9px] font-mono text-slate-500")
                                                        eye_icon = "visibility_off" if is_rev else "visibility"
                                                        
                                                        def toggle_pw(user_to_toggle=u):
                                                            rev = dialog_state.setdefault("revealed", set())
                                                            if user_to_toggle in rev:
                                                                rev.remove(user_to_toggle)
                                                            else:
                                                                rev.add(user_to_toggle)
                                                            render_admin_content.refresh()
                                                            
                                                        ui.button(icon=eye_icon, on_click=toggle_pw).props("flat dense size=xs color=grey-6").classes("w-4 h-4").tooltip("Show/Hide Password")
                                                
                                                with ui.row().classes("items-center gap-1.5"):
                                                    if is_online and u.lower() != "admin":
                                                        ui.button(icon="logout", on_click=lambda u_name=u: perform_force_logout(u_name)).props("flat dense color=amber").classes("w-6 h-6").tooltip("Force Logout Session")
                                                    if u.lower() != "admin":
                                                        ui.button(icon="delete", on_click=lambda u_name=u: perform_delete(u_name)).props("flat dense color=red").classes("w-6 h-6").tooltip("Delete Account")
                                                    else:
                                                        ui.badge("SYSTEM", color="grey-6").classes("text-[8px] px-1.5 py-0.5")

                                    with ui.card().classes("w-1/2 p-4 bg-white border border-slate-200 shadow-sm"):
                                        ui.label("Create New User").classes("text-xs font-bold uppercase tracking-wider text-slate-400 mb-3")
                                        new_user = ui.input(placeholder="Enter username").props("outlined dense").classes("w-full mb-3 text-xs")
                                        password_input = ui.input(placeholder="Password").props("outlined dense").classes("w-full mb-3 text-xs")
                                        new_role = ui.select(["Admin", "RO", "Legal Researcher", "Guest"], value="Legal Researcher").props("outlined dense").classes("w-full mb-4 text-xs")
                                        with ui.row().classes("w-full gap-2"):
                                            def gen_pwd():
                                                import secrets
                                                password_input.value = secrets.token_urlsafe(8)[:10]
                                            ui.button("Generate Password", color="amber", on_click=gen_pwd).props("dense outline size=sm").classes("text-xs flex-grow")
                                            ui.button("Create Account", color="primary", on_click=perform_create).props("dense size=sm").classes("text-xs flex-grow")

                            elif view_name == "funnel":
                                ui.label("Usage Funnel Analysis").classes("text-sm font-bold text-slate-800 mb-2")
                                ui.label("Tracks visitor conversion from initially landing to triggering deep LLM analyses.").classes("text-xs text-slate-500 mb-6")

                                counts = get_funnel_counts()
                                p_hits = counts.get("portal_hits", 148)
                                j_opened = counts.get("judgments_opened", 96)
                                q_sub = counts.get("questions_submitted", 56)
                                a_gen = counts.get("analyses_generated", 22)
                                
                                pct_hits = 100
                                pct_opened = int((j_opened / p_hits) * 100) if p_hits > 0 else 0
                                pct_sub = int((q_sub / p_hits) * 100) if p_hits > 0 else 0
                                pct_gen = int((a_gen / p_hits) * 100) if p_hits > 0 else 0

                                funnel_stages = [
                                    {"stage": "1. Search Portal Hits", "count": p_hits, "pct": f"{pct_hits}%", "color": "from-blue-500 to-blue-600"},
                                    {"stage": "2. Case Judgments Opened", "count": j_opened, "pct": f"{pct_opened}%", "color": "from-blue-600 to-indigo-500"},
                                    {"stage": "3. Chatbot Questions Submitted", "count": q_sub, "pct": f"{pct_sub}%", "color": "from-indigo-500 to-purple-500"},
                                    {"stage": "4. Deep AI Analyses Generated", "count": a_gen, "pct": f"{pct_gen}%", "color": "from-purple-500 to-pink-500"}
                                ]

                                with ui.column().classes("w-full gap-4 items-center bg-slate-50 border border-slate-200 rounded-xl p-6"):
                                    for stage in funnel_stages:
                                        with ui.row().classes("w-full items-center justify-between no-wrap bg-white p-3 rounded-lg border border-slate-100 hover:border-slate-300 transition-all duration-200 shadow-sm"):
                                            with ui.row().classes("items-center gap-3"):
                                                ui.badge(stage["pct"], color="primary").classes("px-2.5 py-1 text-xs font-bold font-mono")
                                                ui.label(stage["stage"]).classes("text-xs font-semibold text-slate-800")
                                            with ui.row().classes("items-center gap-4"):
                                                ui.label(f"{stage['count']} hits").classes("text-xs text-slate-500 font-mono")
                                                opacity_val = float(stage['pct'].replace('%','')) / 100.0
                                                ui.element("div").classes(f"w-36 h-2 rounded bg-gradient-to-r {stage['color']}").style(f"opacity: {max(0.15, opacity_val)}")

                            elif view_name == "models":
                                ui.label("AI Models Controller & Hardware Cache").classes("text-sm font-bold text-slate-800 mb-2")
                                ui.label("Control and cache dynamic local model pipelines directly on GPU/CPU cores.").classes("text-xs text-slate-500 mb-4")

                                current_provider = app.storage.user.get("llm_provider") or settings.llm_provider
                                
                                models_list = [
                                    {
                                        "name": "Claude 3.5 Sonnet", 
                                        "provider_id": "anthropic",
                                        "provider": "Anthropic API", 
                                        "status": "ACTIVE (CACHED)" if current_provider == "anthropic" else "STANDBY", 
                                        "color": "green" if current_provider == "anthropic" else "amber", 
                                        "btn": "Unload" if current_provider == "anthropic" else "Activate",
                                        "model": "claude-3-5-sonnet-20241022"
                                    },
                                    {
                                        "name": "Qwen 2.5 1.5B Instruct", 
                                        "provider_id": "llamacpp",
                                        "provider": "Llama.cpp Local GGUF", 
                                        "status": "ACTIVE (CACHED)" if current_provider == "llamacpp" else "READY (IDLE)", 
                                        "color": "green" if current_provider == "llamacpp" else "blue", 
                                        "btn": "Unload" if current_provider == "llamacpp" else "Load to GPU",
                                        "model": "local-gguf"
                                    },
                                    {
                                        "name": "Qwen 2.5 14B Instruct", 
                                        "provider_id": "ollama",
                                        "provider": "Ollama Endpoint", 
                                        "status": "ACTIVE (CACHED)" if current_provider == "ollama" else "OFFLINE", 
                                        "color": "green" if current_provider == "ollama" else "grey", 
                                        "btn": "Unload" if current_provider == "ollama" else "Connect",
                                        "model": "qwen2.5-14b-instruct"
                                    },
                                    {
                                        "name": "GPT-4o / GPT-4o-Mini", 
                                        "provider_id": "openai",
                                        "provider": "OpenAI API", 
                                        "status": "ACTIVE (CACHED)" if current_provider == "openai" else "STANDBY", 
                                        "color": "green" if current_provider == "openai" else "amber", 
                                        "btn": "Unload" if current_provider == "openai" else "Activate",
                                        "model": "gpt-4o-mini"
                                    }
                                ]

                                with ui.grid().classes("w-full grid-cols-2 gap-4"):
                                    for m in models_list:
                                        with ui.card().classes("p-4 bg-white border border-slate-200 shadow-sm flex flex-col justify-between h-36"):
                                            with ui.row().classes("w-full justify-between items-start no-wrap"):
                                                with ui.column().classes("gap-0"):
                                                    ui.label(m["name"]).classes("text-xs font-bold text-slate-800")
                                                    ui.label(m["provider"]).classes("text-[9px] text-slate-500 font-mono")
                                                ui.badge(m["status"], color=m["color"]).classes("text-[8px] px-2 py-0.5")
                                            
                                            def make_click_handler(model_cfg=m):
                                                def click_handler():
                                                    if model_cfg["btn"] == "Unload":
                                                        app.storage.user["llm_provider"] = "llamacpp"
                                                        app.storage.user["llm_model"] = "local-gguf"
                                                        ui.notify("Model unloaded. Reverted to default Llama.cpp", type="info")
                                                    else:
                                                        app.storage.user["llm_provider"] = model_cfg["provider_id"]
                                                        app.storage.user["llm_model"] = model_cfg["model"]
                                                        if model_cfg["provider_id"] == "anthropic" and not app.storage.user.get("custom_api_key"):
                                                            app.storage.user["custom_api_key"] = settings.anthropic_api_key
                                                        elif model_cfg["provider_id"] == "openai" and not app.storage.user.get("custom_api_key"):
                                                            app.storage.user["custom_api_key"] = settings.openai_api_key
                                                        ui.notify(f"Successfully loaded and switched to {model_cfg['name']}", type="positive")
                                                    render_admin_content.refresh()
                                                return click_handler
                                                
                                            btn_color = "primary" if m["btn"] == "Unload" else "grey"
                                            ui.button(m["btn"], color=btn_color, on_click=make_click_handler()).props("dense outline size=sm").classes("w-full text-xs")

                            elif view_name == "credentials":
                                ui.label("System Credential Manager").classes("text-sm font-bold text-slate-800 mb-2")
                                ui.label("Monitor connection status of backend cloud model API keys.").classes("text-xs text-slate-500 mb-4")

                                cred_configs = [
                                    {
                                        "name": "Anthropic API Key",
                                        "env_key": "ANTHROPIC_API_KEY",
                                        "value": app.storage.user.get("custom_api_key") if app.storage.user.get("llm_provider") == "anthropic" else settings.anthropic_api_key
                                    },
                                    {
                                        "name": "OpenAI API Key",
                                        "env_key": "OPENAI_API_KEY",
                                        "value": app.storage.user.get("custom_api_key") if app.storage.user.get("llm_provider") == "openai" else settings.openai_api_key
                                    },
                                    {
                                        "name": "Ollama Endpoint URL",
                                        "env_key": "OLLAMA_BASE_URL",
                                        "value": settings.ollama_base_url
                                    }
                                ]

                                with ui.column().classes("w-full gap-3 bg-slate-50 border border-slate-200 p-4 rounded-xl"):
                                    for config in cred_configs:
                                        with ui.row().classes("w-full items-center justify-between bg-white p-3 rounded-lg border border-slate-100 shadow-sm no-wrap"):
                                            with ui.column().classes("gap-0 w-2/3"):
                                                ui.label(config["name"]).classes("text-[10px] font-bold tracking-wider font-mono text-slate-400")
                                                curr_val = config["value"] or ""
                                                masked = curr_val[:12] + "..." + curr_val[-4:] if len(curr_val) > 16 else (curr_val if curr_val else "Not configured")
                                                ui.label(masked).classes("text-xs text-slate-700 font-mono")
                                            
                                            is_set = bool(config["value"])
                                            badge_label = "CONNECTED" if is_set else "NOT SET"
                                            badge_color = "green" if is_set else "grey"
                                            ui.badge(badge_label, color=badge_color).classes("px-2.5 py-1 text-[9px] font-bold font-mono")
                                            
                                            def make_edit_handler(cfg=config):
                                                async def edit_handler():
                                                    new_val = await ui.run_javascript(f'prompt("Enter new value for {cfg["name"]}:", "{cfg["value"] or ""}")')
                                                    if new_val is not None:
                                                        new_val = new_val.strip()
                                                        if cfg["env_key"] == "ANTHROPIC_API_KEY":
                                                            app.storage.user["custom_api_key"] = new_val
                                                            app.storage.user["llm_provider"] = "anthropic"
                                                        elif cfg["env_key"] == "OPENAI_API_KEY":
                                                            app.storage.user["custom_api_key"] = new_val
                                                            app.storage.user["llm_provider"] = "openai"
                                                        elif cfg["env_key"] == "OLLAMA_BASE_URL":
                                                            settings.ollama_base_url = new_val
                                                        ui.notify(f"Updated {cfg['name']} successfully!", type="positive")
                                                        render_admin_content.refresh()
                                                return edit_handler
                                                
                                            ui.button(icon="edit", on_click=make_edit_handler()).props("flat dense color=primary size=sm")

                            elif view_name == "corpus":
                                async def handle_corpus_upload(e):
                                    filename = Path(e.file.name).name
                                    content_bytes = await e.file.read()
                                    
                                    # Save file to judgments directory
                                    target_dir = REPO_ROOT / "data" / "sc_judgements"
                                    target_dir.mkdir(parents=True, exist_ok=True)
                                    target_path = target_dir / filename
                                    target_path.write_bytes(content_bytes)
                                    
                                    try:
                                        import sqlite3
                                        from src import store
                                        from src.ingest import case_no_from_filename
                                        from src.parsing import parse_and_chunk
                                        from src.store import fts_index_cases
                                        
                                        case_no = case_no_from_filename(filename)
                                        
                                        # 1. Parse and chunk document
                                        chunks = parse_and_chunk(target_path, case_no)
                                        
                                        # 2. Add chunks to Chroma DB
                                        store.add_chunks(chunks, extra_meta={"date": "", "filename": filename})
                                        
                                        # 3. Insert metadata into SQLite
                                        con = store.init_db()
                                        meta = {
                                            "filename": filename,
                                            "case_no": case_no,
                                            "local_path": str(target_path),
                                            "title": case_no.replace("_", " "),
                                            "court": "SUPREME COURT",
                                            "judges": []
                                        }
                                        store.upsert_judgement(con, meta, len(chunks))
                                        store.mark_indexed(con, filename, "judgment", len(chunks))
                                        
                                        # 4. Add case to FTS index
                                        fts_index_cases([case_no])
                                        con.close()
                                        
                                        # Update ingestion history list
                                        new_item = {
                                            "file": filename,
                                            "size": f"{len(content_bytes)/1024/1024:.1f} MB" if len(content_bytes) >= 1024*1024 else f"{len(content_bytes)/1024:.0f} KB",
                                            "status": "Completed (FTS + Vector Indexed)"
                                        }
                                        state.setdefault("ingested_files", []).insert(0, new_item)
                                        ui.notify(f"Successfully indexed case: {case_no}", type="positive")
                                        
                                        corpus_view.refresh()
                                    except Exception as ex:
                                        ui.notify(f"Ingestion failed: {ex}", type="negative")

                                @ui.refreshable
                                def corpus_view():
                                    ui.label("Corpus Management & Document Uploader").classes("text-sm font-bold text-slate-800 mb-2")
                                    ui.label("Add new case judgments (.pdf) directly to the system search index.").classes("text-xs text-slate-500 mb-4")

                                    # Functional uploader card
                                    ui.upload(on_upload=handle_corpus_upload, auto_upload=True)\
                                        .props('accept=".pdf" flat bordered color=primary label="Select PDF"')\
                                        .classes("w-full mb-4").style("border-radius: 12px;")

                                    ui.label("Recent Ingestion History:").classes("text-xs font-bold uppercase tracking-wider text-slate-400 mb-2")
                                    
                                    history_list = state.get("ingested_files")
                                    if history_list is None:
                                        history_list = [
                                            {"file": "SC_FR_170_2022.pdf", "size": "1.2 MB", "status": "Completed (FTS + Vector Indexed)"},
                                            {"file": "SC_APPEAL_99_2018.pdf", "size": "840 KB", "status": "Completed (FTS + Vector Indexed)"}
                                        ]
                                        state["ingested_files"] = history_list
                                        
                                    with ui.column().classes("w-full gap-2 bg-slate-50 border border-slate-200 p-3 rounded-lg"):
                                        for hf in history_list:
                                            with ui.row().classes("w-full justify-between items-center text-xs py-1 border-b border-slate-200/50"):
                                                with ui.column().classes("gap-0"):
                                                    ui.label(hf["file"]).classes("text-xs font-semibold text-slate-800")
                                                    ui.label(hf["size"]).classes("text-[10px] text-slate-500 font-mono")
                                                ui.badge(hf["status"], color="green").classes("px-2 py-0.5 text-[8px]")

                                corpus_view()

                render_admin_content()

            # Footer
            with ui.row().classes("w-full justify-end bg-slate-100 border-t border-slate-200 p-4").style("flex-shrink: 0;"):
                ui.button("Close", on_click=dialog.close, color="red").props("outline dense").classes("px-4 text-xs font-semibold")
        
        dialog.open()

    def librarian_dialog(initial_query: str = None, force_local: bool = False):
        import asyncio
        with ui.dialog() as dialog, ui.card().style("width: 800px; max-width: 95vw; height: 75vh; padding: 0; display: flex; flex-direction: column; background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 16px; overflow: hidden; box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04);"):
            # Dialog Header: Slate Navy (#0f172a) with Gold Accent bottom border
            with ui.row().classes("w-full bg-[#0f172a] text-white p-4 items-center justify-between no-wrap").style("flex-shrink: 0; border-bottom: 2px solid #b8965a;"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("auto_awesome", color="amber").classes("text-xl")
                    with ui.column().classes("gap-0"):
                        ui.label("Ask ROS").classes("text-sm font-bold tracking-wider text-white uppercase")
                        ui.label("Supreme Court A.I. Research Assistant").classes("text-[10px] text-slate-400")
                with ui.row().classes("items-center gap-1"):
                    def on_clear_history():
                        app.storage.user["chat_history"] = []
                        messages_container.clear()
                        state.pop("uploaded_doc_text", None)
                        state.pop("uploaded_doc_name", None)
                        ui.notify("Chat history cleared", color="grey-7")
                    ui.button(icon="delete_sweep", on_click=on_clear_history).props("flat round dense color=white").classes("hover:bg-white/10").tooltip("Clear Chat History")
                    ui.button(icon="close", on_click=dialog.close).props("flat round dense color=white").classes("hover:bg-white/10")
            
            # Chat Messages Container (Scrollable)
            messages_container = ui.column().classes("w-full p-4 overflow-y-auto bg-slate-50 gap-4").style("flex-grow: 1; height: 100%;")
            
            # Load and render persistent session history
            def load_chat_history():
                messages_container.clear()
                state.pop("uploaded_doc_text", None)
                state.pop("uploaded_doc_name", None)
                history = app.storage.user.get("chat_history", [])
                for msg in history:
                    if msg["role"] == "user":
                        with messages_container:
                            with ui.row().classes("w-full justify-end"):
                                with ui.card().style("padding: 10px 14px; background-color: #0f172a; color: white; max-width: 70%; border-radius: 12px; border-bottom-right-radius: 2px; box-shadow: 0 1px 2px rgba(0,0,0,0.05);"):
                                    ui.label(msg["text"]).classes("text-xs font-medium")
                    else:
                        with messages_container:
                            with ui.row().classes("w-full justify-start"):
                                with ui.card().style("padding: 12px 16px; background-color: #ffffff; color: #334155; max-width: 85%; border-radius: 12px; border-top-left-radius: 2px; border: 1px solid #e2e8f0; box-shadow: 0 1px 2px rgba(0,0,0,0.05);"):
                                    ui.markdown(msg["text"]).classes("text-xs leading-relaxed")
                # Scroll to bottom
                messages_container.run_method('scrollTo', 0, 1000000)

            dialog.on("show", load_chat_history)

            # Input Area
            async def handle_doc_upload(e):
                if not _ai_allowed():   # doc analysis hits the paid API
                    return
                filename = Path(e.file.name).name   # NiceGUI 3.x API; basename only (no path tricks)
                content_bytes = await e.file.read()
                doc_text = ""
                if filename.endswith(".pdf"):
                    temp_path = Path(f"data/parsed/temp_{filename}")
                    temp_path.write_bytes(content_bytes)
                    try:
                        from src.parsing import extract_pages
                        pages = extract_pages(str(temp_path))
                        doc_text = "\n".join(pages)
                    except Exception as ex:
                        doc_text = f"Error reading PDF: {ex}"
                    finally:
                        if temp_path.exists():
                            temp_path.unlink()
                elif filename.lower().endswith((".doc", ".docx")):
                    tp = Path(f"data/parsed/temp_{filename}"); tp.write_bytes(content_bytes)
                    try:
                        from src.ingest import extract_document
                        doc_text = "\n".join(extract_document(str(tp)))
                    except Exception as ex:
                        doc_text = f"Error reading document: {ex}"
                    finally:
                        if tp.exists():
                            tp.unlink()
                else:
                    doc_text = content_bytes.decode("utf-8", errors="ignore")
                
                state["uploaded_doc_text"] = doc_text
                state["uploaded_doc_name"] = filename
                
                with messages_container:
                    with ui.row().classes("w-full justify-start"):
                        with ui.card().style("padding: 10px 14px; background-color: #f1f5f9; color: #1e293b; max-width: 80%; border-radius: 12px; border-top-left-radius: 2px; border: 1px solid #cbd5e1;"):
                            ui.label(f"📎 Attached doc: {filename} ({len(content_bytes):,} bytes). Ask questions below!").classes("text-xs font-semibold")
                ui.notify(f"Attached document: {filename}", type="positive")

            # Input Area
            with ui.row().classes("w-full p-3 bg-white border-t border-slate-200 gap-2 items-center").style("flex-shrink: 0;"):
                doc_uploader = ui.upload(on_upload=handle_doc_upload, auto_upload=True).props('accept=".pdf,.txt,.doc,.docx"').classes("hidden").style("display: none;")
                ui.button(icon="add", on_click=lambda: doc_uploader.run_method("pickFiles")).props("round flat color=primary").classes("flex-shrink-0").style("width: 36px; height: 36px; min-height: 36px; margin: 0;").tooltip("Attach Document")
                input_field = ui.input(placeholder="Ask ROS a question about any cases...").props("outlined dense rounded").classes("flex-grow text-xs")
                async def submit_message(forced_query: str = None):
                    query = forced_query.strip() if forced_query else input_field.value.strip()
                    if not query:
                        return
                    if not _ai_allowed():   # paid API — approved accounts only
                        return
                    if not forced_query:
                        input_field.value = ""
                    
                    # Append user query to persistent chat history
                    history = app.storage.user.get("chat_history", [])
                    history.append({"role": "user", "text": query})
                    app.storage.user["chat_history"] = history

                    # Render user bubble: sleek right-aligned talk bubble
                    with messages_container:
                        with ui.row().classes("w-full justify-end"):
                            with ui.card().style("padding: 10px 14px; background-color: #0f172a; color: white; max-width: 70%; border-radius: 12px; border-bottom-right-radius: 2px; box-shadow: 0 1px 2px rgba(0,0,0,0.05);"):
                                ui.label(query).classes("text-xs font-medium")
                    
                    # Render assistant bubble (initially containing the thinking spinner)
                    with messages_container:
                        with ui.row().classes("w-full justify-start"):
                            with ui.card().style("padding: 12px 16px; background-color: #ffffff; color: #334155; max-width: 85%; border-radius: 12px; border-top-left-radius: 2px; border: 1px solid #e2e8f0; box-shadow: 0 1px 2px rgba(0,0,0,0.05);"):
                                bubble = ui.markdown("").classes("text-xs leading-relaxed")
                                # Show a loading spinner (stays visible until first token returns)
                                with ui.row().classes("items-center gap-2 mt-1") as spinner_row:
                                    spinner = ui.spinner(size="xs", color="primary")
                                    status_lbl = ui.label("Thinking...").classes("text-[10px] text-gray-400 font-bold uppercase tracking-wider")
                    
                    # Scroll to bottom using safe run_method
                    messages_container.run_method('scrollTo', 0, 1000000)
                    
                    # Run the retrieval and LLM call
                    try:
                        # Step 1: Use local model (Llama.cpp) to refine and expand search terms
                        status_lbl.text = "Refining legal search terms..."
                        expanded_query = query
                        username = app.storage.user.get("username") or "system"
                        try:
                            local_sys = ("You are a legal research assistant. Extract 3-5 key legal terms "
                                         "or statutory references from the user query. Output ONLY a "
                                         "comma-separated list of terms, nothing else.")
                            local_usr = f"User query: {query}"
                            from src.analyze_smart import stream_chat
                            prov = app.storage.user.get("llm_provider") or "llamacpp"
                            mdl = app.storage.user.get("llm_model") or settings.llamacpp_model_path
                            akey = app.storage.user.get("custom_api_key")
                            iterator_local = stream_chat(local_sys, local_usr, provider=prov, model=mdl, api_key=akey, max_tokens=100, username=username)
                            local_terms = ""
                            while True:
                                token = await run.io_bound(next, iterator_local, None)
                                if token is None:
                                    break
                                local_terms += token
                            local_terms = local_terms.strip()
                            if local_terms:
                                cleaned_terms = local_terms.replace(",", " ").replace("\n", " ").strip()
                                expanded_query = f"{query} {cleaned_terms}"
                                print(f"AI Query Expansion: '{query}' -> '{expanded_query}'")
                        except Exception as lex:
                            print(f"Local query expansion warning: {lex}")

                        # Step 2: Query the semantic database using the expanded terms
                        status_lbl.text = "Searching case corpus..."
                        from src.retrieve import retrieve
                        hits = await run.io_bound(retrieve, query, k=4 if state.get("uploaded_doc_text") else 8, source="judgment", expanded_query=expanded_query)
                        
                        # Step 3: Setup Claude final answer synthesis
                        status_lbl.text = "Synthesizing legal analysis..." 
                        
                        # Build the prompt. Citations must match case format so the link click parser catches them.
                        sys_p = ("You are the Supreme Court AI Assistant (ROS), an experienced Sri Lankan legal "
                                 "research assistant. Answer the researcher's query. Ground your answer in the provided "
                                 "corpus passages and cite the case + page.\n\n"
                                 "FORMATTING RULES:\n"
                                 "1. Do NOT repeat identical templates or contexts case-by-case. If multiple cases are part of a series "
                                 "or share similar facts/opinions (e.g., SC/APPEAL/39/2023 to SC/APPEAL/46/2023), group and consolidate "
                                 "them under a single summary, bullet list, or comparative table.\n"
                                 "2. Present answers in a clean, executive-level layout: start with a quick overview, followed by consolidated "
                                 "findings, and highlight key legal principles or dissenting rules (such as the 'destination test').\n"
                                 "3. CRITICAL: Format every case you reference as a markdown link using its exact case ID from the context header, "
                                 "e.g. [SC_APPEAL_39_2023](/?case=SC_APPEAL_39_2023). Do not use plain text case names.")
                        usr_p = f"User query: {query}\n\n"
                        if hits:
                            usr_p += "Relevant extracts from the judgment corpus:\n\n"
                            for hit in hits:
                                anchor = hit["meta"].get("anchor") or hit["meta"].get("case_no") or "Judgment"
                                usr_p += f"--- Extract from {anchor} ---\n{hit['text']}\n\n"

                        if state.get("uploaded_doc_text"):
                            from src.analyze_smart import _fit_to_context
                            doc_context = _fit_to_context(state["uploaded_doc_text"], max_output=6000)
                            usr_p += (f"\n--- Uploaded document ({state.get('uploaded_doc_name','document')}) ---\n"
                                      f"{doc_context}\n")
                        elif not hits:
                            spinner_row.delete()
                            bubble.content = ("I couldn't find relevant passages in the corpus for that query. "
                                              "Try rephrasing, or attach a document for me to work from.")
                            return

                        # Call LLM streaming (honours the per-session model override).
                        from src.analyze_smart import stream_chat
                        prov = "llamacpp" if force_local else None
                        mdl = settings.llamacpp_model_path if force_local else None
                        iterator = stream_chat(sys_p, usr_p, provider=prov, model=mdl, username=username)
                        full_text = ""
                        first_token = True
                        while True:
                            token = await run.io_bound(next, iterator, None)
                            if token is None:
                                break
                            
                            # Seamless transition: Delete spinner row only when the first actual token arrives
                            if first_token:
                                try:
                                    spinner_row.delete()
                                except:
                                    pass
                                first_token = False
                                
                            full_text += token
                            
                            # Post-process to ensure all mentions of retrieved case IDs and party names are linked
                            temp_text = full_text
                            for hit in hits:
                                cno = hit["meta"].get("case_no")
                                if cno:
                                    if cno in temp_text and f"case={cno}" not in temp_text:
                                        temp_text = temp_text.replace(cno, f"[{cno.replace('_', ' ')}](/?case={cno})")
                                    # Link the short name of the parties
                                    m = case_meta(cno)
                                    if m and m.get("parties") and m["parties"] != NOT_AVAILABLE:
                                        parties = m["parties"]
                                        short_name = parties.split("v.")[0].strip()
                                        if len(short_name) > 5 and short_name in temp_text and f"case={cno}" not in temp_text:
                                            temp_text = temp_text.replace(short_name, f"[{short_name}](/?case={cno})")
                                            
                            bubble.content = temp_text
                            messages_container.run_method('scrollTo', 0, 1000000)
                            
                        # Save final assistant response to history
                        history = app.storage.user.get("chat_history", [])
                        history.append({"role": "assistant", "text": bubble.content})
                        app.storage.user["chat_history"] = history

                    except Exception as e:
                        try:
                            spinner_row.delete()
                        except:
                            pass
                        bubble.content = f"Error: {e}"
                
                input_field.on("keydown.enter", lambda: asyncio.create_task(submit_message()))
                ui.button(icon="send", on_click=lambda: asyncio.create_task(submit_message())).props("flat round color=primary size=sm").classes("w-8 h-8")
                
            if initial_query:
                prompt = f"Analyze the keyword/phrase '{initial_query.strip()}' across the case judgment corpus. Provide a summary of its legal context, key variations in how it is interpreted, and primary legal insights from the Supreme Court case law."
                # When using initial query, load chat history first, then submit prompt
                dialog.on("show", lambda: asyncio.create_task(submit_message(prompt)))
                # Trigger search execution immediately
                asyncio.create_task(submit_message(prompt))
        
        dialog.open()

    def refresh_bookmarks_ui():
        if not containers["bookmarks"]:
            return
        containers["bookmarks"].clear()
        username = _safe_username()
        saved = get_bookmarks(username)
        if not saved:
            return

        def remove_bookmark(case_no):
            toggle_bookmark(username, case_no)
            ui.notify("Removed from Bookmarks", color="grey-7")
            refresh_bookmarks_ui()
            if state.get("case") and state["case"][0] == case_no:
                render_breakdown()

        with containers["bookmarks"]:
            exp = ui.expansion().classes("w-full mb-1").props("dense header-class='bg-amber-100 text-primary text-xs font-semibold rounded-md' expand-icon-class='text-primary'")
            with exp.add_slot('header'):
                with ui.row().classes("items-center justify-between w-full py-1.5 px-2"):
                    ui.label("Saved Bookmarks").classes("text-xs font-bold text-primary")
                    ui.badge(str(len(saved)), color="amber-3", text_color="primary").classes("text-[10px] px-2 py-0.5 rounded-full")
            
            with exp:
                with ui.column().classes("w-full pl-2 gap-1 py-1"):
                    for b_cn in saved:
                        with ui.row().classes("w-full items-center justify-between py-1 px-2 hover:bg-gray-100 rounded cursor-pointer transition-colors duration-150").on("click", lambda c=b_cn: open_case(c)):
                            ui.label(b_cn).classes("text-[11px] font-medium text-gray-800")
                            with ui.button(icon="delete_outline").props("flat round dense size=xs color=grey-6").classes("hover:text-red-600").on("click.stop", lambda c=b_cn: remove_bookmark(c)):
                                ui.tooltip("Remove Bookmark")

    # ----------------------------- PDF ------------------------------------ #
    def render_pdf():
        pdf_pane.clear()
        with pdf_pane:
            st = state.get("statute")
            if st:
                ui.label(st.get("title") or st.get("statute_id") or "Statute").classes("text-sm font-bold case-title mb-1")
                sfn = st.get("filename") or ""
                if sfn and ((STATUTE_DIR / sfn).is_file() or (LANKALAW_CASES_DIR / sfn).is_file()):
                    iframe = ui.element("iframe").props(f'src="/statute/{sfn}"').classes("w-full flex-grow iframe-pdf").style(
                        "border:1px solid #e5e7eb;border-radius:6px;min-height:0;")
                    if left_drawer:
                        iframe.bind_visibility_from(left_drawer, 'value', backward=lambda v: not v)
                    
                    # Mobile Download Option Container
                    with ui.column().classes("mobile-download-btn w-full items-center justify-center gap-4 p-6 bg-slate-50 border rounded-lg").style("display: none;"):
                        ui.icon("picture_as_pdf", size="64px", color="primary")
                        ui.label("PDF Preview is disabled on mobile devices.").classes("text-sm font-semibold text-gray-600 text-center")
                        ui.button("Download PDF Document", icon="download", on_click=lambda: ui.download(f"/statute/{sfn}")).props("color=primary").classes("px-6 py-2 rounded-lg text-xs font-bold shadow")
                else:
                    with ui.column().classes("items-center w-full mt-10 gap-2"):
                        ui.label("Statute PDF not available locally.").classes("text-gray-400 text-center")
                        pass
                return
            if not state["case"]:
                ui.label("Open a case from the Library  →").classes("text-gray-400 mt-10 w-full text-center")
                return
            cn, fn = state["case"]
            ui.label(cn).classes("text-sm font-bold case-title mb-1")
            src = f"/pdf/{fn}" + (f"#page={state['page']}" if state["page"] else "")
            iframe = ui.element("iframe").props(f'src="{src}"').classes("w-full flex-grow iframe-pdf").style(
                "border:1px solid #e5e7eb;border-radius:6px;min-height:0;")
            if left_drawer:
                iframe.bind_visibility_from(left_drawer, 'value', backward=lambda v: not v)
                
            # Mobile Download Option Container
            with ui.column().classes("mobile-download-btn w-full items-center justify-center gap-4 p-6 bg-slate-50 border rounded-lg").style("display: none;"):
                ui.icon("picture_as_pdf", size="64px", color="primary")
                ui.label("PDF Preview is disabled on mobile devices.").classes("text-sm font-semibold text-gray-600 text-center")
                ui.button("Download PDF Document", icon="download", on_click=lambda: ui.download(f"/pdf/{fn}")).props("color=primary").classes("px-6 py-2 rounded-lg text-xs font-bold shadow")
    # --------------------------- Breakdown -------------------------------- #
    def sec(t):
        ui.label(t).classes("sec")

    def chips(items, color, text_color="white"):
        with ui.row().classes("flex-wrap gap-x-2.5 gap-y-3.5 mt-2 mb-3"):
            for it in items:
                if color == "amber-3":
                    ui.chip(it, on_click=lambda term=it: goto(term)).props("clickable").classes("cursor-pointer text-xs font-semibold").style(
                        "background-color: #fcd34d !important; color: #82212b !important; "
                        "height: 28px; padding: 0 12px; border-radius: 14px;"
                    )
                else:
                    ui.chip(it, on_click=lambda term=it: goto(term)).props(f"color={color} text-color={text_color} clickable").classes("cursor-pointer text-sm font-semibold px-3.5 py-1.5 rounded-full")

    def show_graph(cn):
        nodes = [{"id": cn, "label": cn, "color": "#1a73e8", "font": {"bold": True}, "size": 24}]
        edges = []
        
        bd = get_breakdown(cn)
        if bd and bd.get("precedent_index"):
            for p in bd["precedent_index"]:
                cited = p.get("cited_case")
                if cited and cited != NOT_AVAILABLE:
                    if not any(n["id"] == cited for n in nodes):
                        nodes.append({"id": cited, "label": cited, "color": "#00796b", "size": 16})
                    edges.append({"from": cn, "to": cited, "arrows": "to", "label": p.get("treatment", "")})
                    
        children = get_citing_cases(cn)
        for child in children:
            if not any(n["id"] == child for n in nodes):
                nodes.append({"id": child, "label": child, "color": "#d93025", "size": 16})
            edges.append({"from": child, "to": cn, "arrows": "to"})
            
        html_content = f"""
        <script type="text/javascript" src="https://cdnjs.cloudflare.com/ajax/libs/vis-network/9.1.2/standalone/umd/vis-network.min.js"></script>
        <div id="vis_graph" style="width:100%; height:440px; border:1px solid #e5e7eb; border-radius:8px; background-color:#f9fafb;"></div>
        <script>
            var nodes = new vis.DataSet({json.dumps(nodes)});
            var edges = new vis.DataSet({json.dumps(edges)});
            var container = document.getElementById('vis_graph');
            var data = {{ nodes: nodes, edges: edges }};
            var options = {{
                nodes: {{
                    shape: 'dot',
                    font: {{ size: 12, face: 'Plus Jakarta Sans', color: '#111827' }},
                    borderWidth: 2
                }},
                edges: {{
                    width: 1.5,
                    color: {{ color: '#9ca3af', highlight: '#1a73e8' }},
                    font: {{ size: 9, align: 'top', color: '#4b5563' }}
                }},
                physics: {{
                    stabilization: true,
                    barnesHut: {{
                        gravitationalConstant: -1500,
                        centralGravity: 0.3,
                        springLength: 95
                    }}
                }}
            }};
            var network = new vis.Network(container, data, options);
            network.on("doubleClick", function(params) {{
                if (params.nodes.length > 0) {{
                    var clickedNode = params.nodes[0];
                    var el = document.querySelector('.citation-graph-input input');
                    if (el) {{
                        el.value = clickedNode;
                        el.dispatchEvent(new Event('input'));
                    }}
                }}
            }});
        </script>
        """
        
        with ui.dialog() as dialog, ui.card().classes("w-[80vw] max-w-[1100px] h-[540px] p-4"):
            with ui.row().classes("w-full items-center justify-between no-wrap mb-2"):
                ui.label(f"Precedent Map — {cn}").classes("text-sm font-bold text-gray-800")
                ui.button(icon="close", on_click=dialog.close).props("flat round dense color=grey-7")
            
            def on_node_select(e):
                if e.value:
                    open_case(e.value)
                    dialog.close()
            graph_input = ui.input(on_change=on_node_select).classes("citation-graph-input hidden")
            
            ui.html(html_content).classes("w-full h-[450px]")
            
        dialog.open()

    async def gen_breakdown(cn):
        if state.get("bd_pending") == cn:
            return  # already generating this case (this session)
        if _BD_INFLIGHT.get(cn, 0) > time.monotonic() - 600:
            # Another session/device (or this phone before its tab reloaded) is
            # already running this case — adopt it; the poll timer renders it.
            state["bd_pending"] = cn
            ui.notify("This case is already being analysed — the result will appear "
                      "here automatically when it is ready.", type="info")
            return
        if not _ai_allowed():
            return
        if not _llm_acquire():
            ui.notify("The AI engine is busy with other requests — please try again in a minute.",
                      type="warning")
            return
        _BD_INFLIGHT[cn] = time.monotonic()
        breakdown_pane.clear()
        with breakdown_pane:
            ui.label("Analysis & Breakdown").classes("pane-head")
            with ui.column().classes("w-full gap-4 mt-4"):
                status_label = ui.html('<div class="text-xs font-semibold text-gray-700">Initialising smart legal analysis engine<span class="loading-dots"></span></div>')
                
                # Gemini-style flowing gradient skeleton placeholders
                with ui.column().classes("w-full gap-3 mt-2"):
                    # Paragraph 1
                    with ui.column().classes("w-full gap-2 mb-2"):
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-full")
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-11/12")
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-4/5")
                    
                    # Paragraph 2
                    with ui.column().classes("w-full gap-2 mb-2"):
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-full")
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-5/6")
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-2/3")
                        
                    # Paragraph 3
                    with ui.column().classes("w-full gap-2 mb-2"):
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-11/12")
                        ui.element("div").classes("h-3.5 gemini-skeleton-flow w-3/4")

        import asyncio
        loop = asyncio.get_running_loop()

        def update_ui(val: float, text: str, chunk: str | None = None):
            try:
                status_label.set_content(f'<div class="text-xs font-semibold text-gray-700">{text}<span class="loading-dots"></span></div>')
            except RuntimeError as e:
                if "The client this element belongs to has been deleted" not in str(e):
                    raise
            except Exception:
                pass

        def on_progress(val: float, text: str, chunk: str | None = None):
            loop.call_soon_threadsafe(update_ui, val, text, chunk)

        state["bd_pending"] = cn
        state.pop("bd_error", None)
        # Generate in a worker thread. analyze_case caches to the DB — and the cache
        # write completes even if THIS client's websocket drops during the wait (the
        # thread keeps running). _finish_breakdown renders now if we're still
        # connected; the _poll_breakdown timer is the backstop that renders from
        # cache after a drop, so the spinner can't get stuck.
        from src.analyze_smart import analyze_case
        state.pop("bd_fresh", None)
        try:
            provider = app.storage.user.get("llm_provider")
            model = app.storage.user.get("llm_model")
            api_key = app.storage.user.get("custom_api_key")
            username = app.storage.user.get("username")
            ca = await run.io_bound(analyze_case, cn, True, on_progress, provider, model, api_key, username)
            # Record deep analysis generated
            record_funnel_event("analyses_generated")
            q = ca.quality()
            # Hollow results aren't cached (so the next open re-attempts), so stash
            # this run in session state — render_breakdown shows it with a warning.
            state["bd_fresh"] = (cn, ca.model_dump(mode="json"), q)
        except Exception as e:  # noqa: BLE001
            print(f"Breakdown failed for {cn}: {e}")
            state["bd_error"] = (cn, str(e))
        finally:
            _llm_release()
            _BD_INFLIGHT.pop(cn, None)
        _finish_breakdown(cn)

    def _finish_breakdown(cn):
        if state.get("bd_pending") != cn:
            return  # already finished — avoid a double render from the timer
        state["bd_pending"] = None
        err = state.get("bd_error")
        if err and err[0] == cn:
            state.pop("bd_error", None)
            try:
                ui.notify(f"Breakdown failed: {err[1]}", type="negative")
            except Exception:
                pass
        if state.get("case") and state["case"][0] == cn:
            try:
                render_breakdown()
            except Exception as e:  # noqa: BLE001
                print(f"[gen_breakdown] render skipped: {e}")

    def _poll_breakdown():
        """Per-client 2 s timer: if the pending breakdown has finished (cached) or
        failed, render it — covers the case where this client dropped mid-wait so the
        awaited render above never ran."""
        cn = state.get("bd_pending")
        if not cn:
            return
        if (state.get("bd_error") or (None,))[0] == cn or get_breakdown(cn) \
                or (state.get("bd_fresh") or (None,))[0] == cn:
            _finish_breakdown(cn)

    # Room reserved (in tokens) for the running conversation when fitting the
    # judgment into the model context — without this, the judgment alone fills
    # n_ctx and the FIRST follow-up question overflows llama_decode.
    _CHAT_HISTORY_TOKENS = 1400
    _CHAT_HISTORY_MAX_MSGS = 6          # most recent turns kept verbatim
    _CHAT_HISTORY_MAX_CHARS = 4200      # ≈ _CHAT_HISTORY_TOKENS at ~3 chars/token

    def call_chatbot_api(cn, query):
        from src.config import REPO_ROOT, settings
        from src.parsing import extract_pages
        from src.analyze_smart import _chat, _fit_to_context
        import sqlite3

        # Check memory cache first
        case_texts = state.setdefault("case_texts", {})
        if cn not in case_texts:
            con = sqlite3.connect(settings.sqlite_path)
            row = con.execute("SELECT filename, local_path FROM judgements WHERE case_no=? LIMIT 1", (cn,)).fetchone()
            if not row:  # background renames can shift casing under an open tab
                row = con.execute("SELECT filename, local_path FROM judgements "
                                  "WHERE case_no=? COLLATE NOCASE LIMIT 1", (cn,)).fetchone()
            con.close()
            if not row:
                raise RuntimeError(f"Judgment file not found for case: {cn}")

            # Judgements come from two harvests — use the recorded path, then
            # fall back to both PDF folders (sc + lankalaw NLR/SLR reports).
            from pathlib import Path as _P
            pdf_path = _P(row[1]) if row[1] else _P("/nonexistent")
            if not pdf_path.exists():
                for _d in ("sc_judgements", "lankalaw_cases"):
                    _cand = REPO_ROOT / "data" / _d / row[0]
                    if _cand.exists():
                        pdf_path = _cand
                        break
            pages = extract_pages(str(pdf_path), ocr_langs=settings.tesseract_langs)
            text = "\n".join(pages)
            # Over-reserve by the history budget so judgment + history + output
            # always fit n_ctx together.
            fitted = _fit_to_context(text, max_output=4096 + _CHAT_HISTORY_TOKENS, provider="llamacpp")
            case_texts[cn] = fitted

        judgment_text = case_texts[cn]
        if state.get("uploaded_doc_text"):
            from src.analyze_smart import _fit_to_context
            # Re-fit judgment context to a smaller size (1024 tokens) when compared against an uploaded document
            judgment_text = _fit_to_context(judgment_text, max_output=6000, provider="llamacpp")

        system_text = (
            "You are a helpful, professional legal research assistant.\n"
            "You are provided with the text of a Supreme Court judgment below.\n"
            "Answer the user's questions about this judgment accurately, objectively, and based strictly on the judgment text.\n"
            "If the answer cannot be found or inferred from the text, state that you do not have enough information.\n\n"
            "Judgment Text:\n"
            f"{judgment_text}"
        )

        # History cap: most recent turns only, each clipped, total char-budgeted —
        # a long conversation can no longer overflow the context window.
        messages = state["chats"].get(cn, [])
        recent = messages[:-1][-_CHAT_HISTORY_MAX_MSGS:]
        trimmed = len(messages[:-1]) > len(recent)
        lines = []
        for msg in recent:
            role = "User" if msg["role"] == "user" else "Assistant"
            content = (msg["content"] or "").strip()
            if len(content) > 700:
                content = content[:700] + " …"
            lines.append(f"{role}: {content}")
        history_str = "\n".join(lines)
        while len(history_str) > _CHAT_HISTORY_MAX_CHARS and lines:
            lines.pop(0)
            trimmed = True
            history_str = "\n".join(lines)
        if trimmed and history_str:
            history_str = "[earlier conversation trimmed]\n" + history_str

        user_text = ""
        if history_str:
            user_text += f"Conversation history so far:\n{history_str}\n"
        user_text += f"Latest User Question: {query}"

        provider = app.storage.user.get("llm_provider")
        model = app.storage.user.get("llm_model")
        api_key = app.storage.user.get("custom_api_key")
        reply = _chat(system_text, user_text, provider=provider, model=model, api_key=api_key)
        return reply

    def stream_chatbot_api(cn, query, provider, model, api_key, username=None):
        # Record chatbot question submitted
        record_funnel_event("questions_submitted")
        from src.config import REPO_ROOT, settings
        from src.parsing import extract_pages
        from src.analyze_smart import stream_chat, _fit_to_context
        import sqlite3

        # Check memory cache first
        case_texts = state.setdefault("case_texts", {})
        if cn not in case_texts:
            con = sqlite3.connect(settings.sqlite_path)
            row = con.execute("SELECT filename, local_path FROM judgements WHERE case_no=? LIMIT 1", (cn,)).fetchone()
            if not row:  # background renames can shift casing under an open tab
                row = con.execute("SELECT filename, local_path FROM judgements "
                                  "WHERE case_no=? COLLATE NOCASE LIMIT 1", (cn,)).fetchone()
            con.close()
            if not row:
                raise RuntimeError(f"Judgment file not found for case: {cn}")

            # Judgements come from two harvests — use the recorded path, then
            # fall back to both PDF folders (sc + lankalaw NLR/SLR reports).
            from pathlib import Path as _P
            pdf_path = _P(row[1]) if row[1] else _P("/nonexistent")
            if not pdf_path.exists():
                for _d in ("sc_judgements", "lankalaw_cases"):
                    _cand = REPO_ROOT / "data" / _d / row[0]
                    if _cand.exists():
                        pdf_path = _cand
                        break
            pages = extract_pages(str(pdf_path), ocr_langs=settings.tesseract_langs)
            text = "\n".join(pages)
            fitted = _fit_to_context(text, max_output=4096 + _CHAT_HISTORY_TOKENS, provider="llamacpp")
            case_texts[cn] = fitted

        judgment_text = case_texts[cn]
        if state.get("uploaded_doc_text"):
            from src.analyze_smart import _fit_to_context
            # Re-fit judgment context to a smaller size (1024 tokens) when compared against an uploaded document
            judgment_text = _fit_to_context(judgment_text, max_output=6000, provider="llamacpp")

        system_text = (
            "You are a helpful, professional legal research assistant.\n"
            "You are provided with the text of a Supreme Court judgment below.\n"
            "Answer the user's questions about this judgment accurately, objectively, and based strictly on the judgment text.\n"
            "If the answer cannot be found or inferred from the text, state that you do not have enough information.\n\n"
            "Judgment Text:\n"
            f"{judgment_text}"
        )

        messages = state["chats"].get(cn, [])
        recent = messages[:-1][-_CHAT_HISTORY_MAX_MSGS:]
        trimmed = len(messages[:-1]) > len(recent)
        lines = []
        for msg in recent:
            role = "User" if msg["role"] == "user" else "Assistant"
            content = (msg["content"] or "").strip()
            if len(content) > 700:
                content = content[:700] + " …"
            lines.append(f"{role}: {content}")
        history_str = "\n".join(lines)
        while len(history_str) > _CHAT_HISTORY_MAX_CHARS and lines:
            lines.pop(0)
            trimmed = True
            history_str = "\n".join(lines)
        if trimmed and history_str:
            history_str = "[earlier conversation trimmed]\n" + history_str

        user_text = ""
        if history_str:
            user_text += f"Conversation history so far:\n{history_str}\n"
        user_text += f"Latest User Question: {query}"
        for token in stream_chat(system_text, user_text, provider=provider, model=model, api_key=api_key, username=username):
            yield token

    @ui.refreshable
    def chat_history_messages(cn: str, send_cb):
        messages = state.setdefault("chats", {}).setdefault(cn, [])
        if not messages:
            ui.label("Frequently Asked Questions:").classes("text-gray-700 text-xs font-bold mb-2 pl-2")
            
            bd = get_breakdown(cn)
            suggestions = []
            if bd:
                # 1. Legal Issues
                issues = bd.get("legal_issues") or []
                for issue in issues[:2]:
                    q_text = issue.get("question") if isinstance(issue, dict) else issue
                    if q_text and len(q_text) > 10:
                        q_text_clean = q_text.strip("• ").rstrip("?")
                        if len(q_text_clean) > 80:
                            q_text_clean = q_text_clean[:77] + "..."
                        suggestions.append(f"How did the court resolve the issue of {q_text_clean}?")
                
                # 2. Ratio decidendi
                if bd.get("ratio_decidendi"):
                    suggestions.append("Explain the ratio decidendi in simple terms.")
                
                # 3. Precedents
                precedents = bd.get("precedent_index") or []
                valid_precedents = [p.get("cited_case") for p in precedents if p.get("cited_case")]
                if valid_precedents:
                    suggestions.append(f"How was the case {valid_precedents[0]} treated or distinguished?")
            
            # Fallbacks to ensure we always have 3-4 suggestions
            if len(suggestions) < 3:
                suggestions.append("What is the procedural history of this case?")
            if len(suggestions) < 4:
                suggestions.append("What specific relief or remedy was granted?")
            if len(suggestions) < 4:
                suggestions.append("Was the decision unanimous, and who wrote the judgment?")
            if len(suggestions) < 4:
                suggestions.append("What are the practical implications of this decision?")
            
            suggestions = suggestions[:4]  # Keep max 4
            for sugg in suggestions:
                def make_click(s=sugg):
                    return lambda: send_cb(s)
                ui.button(sugg, on_click=make_click()).props("outline dense no-caps").classes("text-left text-[11px] text-primary bg-white border border-red-100 hover:bg-red-50/50 mb-2 w-full p-2").style("border-radius: 8px; line-height: 1.3;")
        else:
            for msg in messages:
                if msg["role"] == "user":
                    with ui.chat_message(sent=True, name="You").classes("text-xs user-message-bubble"):
                        ui.markdown(msg["content"])
                else:
                    if msg.get("content"):
                        with ui.chat_message(sent=False, name="ROS AI").classes("text-xs assistant-message-bubble"):
                            ui.markdown(msg["content"])
            
        if state.get("chat_loading") == cn and (not messages or messages[-1]["role"] != "assistant" or not messages[-1]["content"]):
            with ui.chat_message(sent=False, name="ROS AI").classes("text-xs assistant-message-bubble"):
                with ui.row().classes("items-center gap-2 no-wrap"):
                    ui.spinner(size="xs", color="primary")
                    ui.label("Thinking...").classes("text-[10px] text-gray-400 font-bold uppercase tracking-wider")

    @ui.refreshable
    def floating_chat_widget():
        if not state["case"]:
            return
            
        cn = state["case"][0]
        btn_class = "ask-ross-btn-close" if state["chat_open"] else "ask-ross-btn-glowing"
        
        if state["chat_open"]:
            ui.button("CLOSE", icon="close", on_click=lambda: (state.update({"chat_open": False}), floating_chat_widget.refresh())).props(
                "rounded"
            ).classes(btn_class).style(
                "position: fixed; bottom: 50px; right: 24px; z-index: 9999; padding: 0 20px; height: 48px; font-size: 13px; font-weight: bold; text-transform: uppercase; letter-spacing: 0.05em; border-radius: 24px; "
                "background-color: #ffffff !important; border: 1.5px solid #82212b !important; color: #82212b !important;"
            )
        else:
            ui.button("ASK ROS", icon="auto_awesome", on_click=lambda: (state.update({"chat_open": True}), floating_chat_widget.refresh())).props(
                "color=primary rounded"
            ).classes(btn_class).style(
                "position: fixed; bottom: 50px; right: 24px; z-index: 9999; padding: 0 20px; height: 48px; font-size: 13px; font-weight: bold; text-transform: uppercase; letter-spacing: 0.05em; border-radius: 24px;"
            )
        # Floating Chat Dialog Card (opens above the button)
        if state["chat_open"]:
            with ui.card().style(
                "position: fixed; bottom: 120px; right: 24px; width: 360px; height: 480px; z-index: 9999; "
                "border-radius: 16px; border: 1px solid #dadce0; box-shadow: 0 8px 32px rgba(0,0,0,0.15); "
                "background: #ffffff; display: flex; flex-direction: column; overflow: hidden;"
            ).classes("p-0"):
                # Header row
                with ui.row().classes("w-full items-center justify-between bg-primary text-white p-3").style("flex-shrink: 0;"):
                    with ui.row().classes("items-center gap-2"):
                        ui.icon("chat_bubble_outline", size="20px")
                        with ui.column().classes("gap-1"):
                            ui.label("Ask ROS about case:").classes("text-[10px] opacity-90 uppercase tracking-wider font-bold leading-none")
                            ui.label(cn).classes("text-xs font-bold leading-none text-amber-300 truncate").style("max-width: 220px;")
                    ui.button(icon="close", on_click=lambda: (state.update({"chat_open": False}), floating_chat_widget.refresh())).props("flat round dense color=white").classes("text-xs")
                
                # Chat History Area
                messages = state.setdefault("chats", {}).setdefault(cn, [])
                
                async def send_suggested_question(text: str):
                    if state.get("chat_loading"):
                        return
                    cooldown = 20.0 if demo else 5.0
                    since = time.monotonic() - state.get("chat_last_ts", 0.0)
                    if since < cooldown:
                        ui.notify(f"Please wait {int(cooldown - since) + 1}s between questions.",
                                  type="warning")
                        return
                    if not _ai_allowed():
                        return
                    if not _llm_acquire():
                        ui.notify("The AI engine is busy with other requests — try again shortly.",
                                  type="warning")
                        return
                    state["chat_last_ts"] = time.monotonic()
                    messages.append({"role": "user", "content": text})
                    state["chat_loading"] = cn
                    chat_history_messages.refresh(cn, send_suggested_question)
                    scroll.scroll_to(percent=1.0)

                    try:
                        messages.append({"role": "assistant", "content": ""})
                        last_refresh = time.monotonic()
                        provider = app.storage.user.get("llm_provider") or "llamacpp"
                        model = app.storage.user.get("llm_model") or settings.llamacpp_model_path
                        api_key = app.storage.user.get("custom_api_key")
                        username = app.storage.user.get("username")
                        if state.get("uploaded_doc_text"): text = f"Uploaded Doc Context:\n{state['uploaded_doc_text']}\n\nUser query: " + text
                        iterator = stream_chatbot_api(cn, text, provider, model, api_key, username)
                        while True:
                            chunk = await run.io_bound(next, iterator, None)
                            if chunk is None:
                                break
                            messages[-1]["content"] += chunk
                            now = time.monotonic()
                            if now - last_refresh > 0.08:
                                chat_history_messages.refresh(cn, send_suggested_question)
                                scroll.scroll_to(percent=1.0)
                                last_refresh = now
                    except Exception as e:
                        messages.append({"role": "assistant", "content": f"Error calling AI: {e}"})
                    finally:
                        _llm_release()
                        state["chat_loading"] = None
                        chat_history_messages.refresh(cn, send_suggested_question)
                        scroll.scroll_to(percent=1.0)

                scroll = ui.scroll_area().classes("flex-grow w-full p-3 bg-gray-50")
                with scroll:
                    chat_history_messages(cn, send_suggested_question)
                
                scroll.scroll_to(percent=1.0)
                
                # Bottom Input Area
                async def handle_widget_upload(e):
                    if not _ai_allowed():   # doc analysis hits the paid API
                        return
                    filename = Path(e.file.name).name   # NiceGUI 3.x API; basename only (no path tricks)
                    content_bytes = await e.file.read()
                    doc_text = ""
                    if filename.endswith(".pdf"):
                        temp_path = Path(f"data/parsed/temp_{filename}")
                        temp_path.write_bytes(content_bytes)
                        try:
                            from src.parsing import extract_pages
                            pages = extract_pages(str(temp_path))
                            doc_text = "\n".join(pages)
                        except Exception as ex:
                            doc_text = f"Error reading PDF: {ex}"
                        finally:
                            if temp_path.exists():
                                temp_path.unlink()
                    elif filename.lower().endswith((".doc", ".docx")):
                        tp = Path(f"data/parsed/temp_{filename}"); tp.write_bytes(content_bytes)
                        try:
                            from src.ingest import extract_document
                            doc_text = "\n".join(extract_document(str(tp)))
                        except Exception as ex:
                            doc_text = f"Error reading document: {ex}"
                        finally:
                            if tp.exists():
                                tp.unlink()
                    else:
                        doc_text = content_bytes.decode("utf-8", errors="ignore")
                    
                    state["uploaded_doc_text"] = doc_text
                    state["uploaded_doc_name"] = filename
                    
                    messages.append({"role": "assistant", "content": f"📎 Attached doc: {filename} ({len(content_bytes):,} bytes). Ask questions below!"})
                    chat_history_messages.refresh(cn, send_suggested_question)
                    scroll.scroll_to(percent=1.0)
                    ui.notify(f"Attached document: {filename}", type="positive")

                # Bottom Input Area
                with ui.row().classes("w-full items-center gap-2 p-2 bg-white border-t no-wrap").style("border-color: #dadce0; flex-shrink: 0;"):
                    widget_uploader = ui.upload(on_upload=handle_widget_upload, auto_upload=True).props('accept=".pdf,.txt,.doc,.docx"').classes("hidden").style("display: none;")
                    ui.button(icon="add", on_click=lambda: widget_uploader.run_method("pickFiles")).props("round flat color=primary").classes("flex-shrink-0").style("width: 36px; height: 36px; min-height: 36px; margin: 0;").tooltip("Attach Document")
                    chat_input = ui.input(placeholder="Ask a question about this case...").props("outlined dense rounded").classes("flex-grow text-xs")
                    async def on_send():
                        val = chat_input.value.strip()
                        if not val:
                            return
                        if state.get("chat_loading"):
                            return  # one question at a time per session
                        # Cooldown between questions — stricter on the public demo —
                        # plus the global gate so visitors can't queue up the model.
                        cooldown = 20.0 if demo else 5.0
                        since = time.monotonic() - state.get("chat_last_ts", 0.0)
                        if since < cooldown:
                            ui.notify(f"Please wait {int(cooldown - since) + 1}s between questions.",
                                      type="warning")
                            return
                        if not _ai_allowed():
                            return
                        if not _llm_acquire():
                            ui.notify("The AI engine is busy with other requests — try again shortly.",
                                      type="warning")
                            return
                        state["chat_last_ts"] = time.monotonic()
                        chat_input.value = ""
                        messages.append({"role": "user", "content": val})
                        state["chat_loading"] = cn
                        chat_history_messages.refresh(cn, send_suggested_question)
                        scroll.scroll_to(percent=1.0)

                        try:
                            messages.append({"role": "assistant", "content": ""})
                            last_refresh = time.monotonic()
                            provider = app.storage.user.get("llm_provider") or "llamacpp"
                            model = app.storage.user.get("llm_model") or settings.llamacpp_model_path
                            api_key = app.storage.user.get("custom_api_key")
                            username = app.storage.user.get("username")
                            iterator = stream_chatbot_api(cn, val, provider, model, api_key, username)
                            while True:
                                chunk = await run.io_bound(next, iterator, None)
                                if chunk is None:
                                    break
                                messages[-1]["content"] += chunk
                                now = time.monotonic()
                                if now - last_refresh > 0.08:
                                    chat_history_messages.refresh(cn, send_suggested_question)
                                    scroll.scroll_to(percent=1.0)
                                    last_refresh = now
                        except Exception as e:
                            messages.append({"role": "assistant", "content": f"Error calling AI: {e}"})
                        finally:
                            _llm_release()
                            state["chat_loading"] = None
                            chat_history_messages.refresh(cn, send_suggested_question)
                            scroll.scroll_to(percent=1.0)
                            
                    chat_input.on("keydown.enter", on_send)
                    ui.button(icon="send", on_click=on_send).props("flat round dense color=primary").classes("hover:bg-blue-50")
                    
                    def clear_history():
                        state["chats"][cn] = []
                        chat_history_messages.refresh(cn, send_suggested_question)
                        scroll.scroll_to(percent=1.0)
                    ui.button(icon="delete_outline", on_click=clear_history).props("flat round dense color=grey-7").classes("hover:bg-gray-100")

    def render_statute_info(st):
        """Statute companion panel — statutes carry no judicial breakdown, so this
        shows the Act's identity + source while its full text fills the PDF pane."""
        title = st.get("title") or st.get("statute_id") or "Statute"
        with ui.card().classes("w-full p-4 mb-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none; background: #ffffff;"):
            ui.label("Statute / Legislation").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
            ui.label(title).classes("text-sm font-bold text-gray-800 case-title")
            meta_bits = []
            if st.get("act_no"):
                meta_bits.append(f"No. {st['act_no']}")
            if st.get("year"):
                meta_bits.append(str(st["year"]))
            if meta_bits:
                ui.label(" · ".join(meta_bits)).classes("text-xs text-gray-500 mt-0.5")
            pass

        sid = st.get("statute_id") or st.get("id")
        an = get_breakdown(sid) if sid else None
        if an and an.get("doc_kind") == "statute":
            _render_statute_analysis(an, sid)
        elif not demo:
            with ui.card().classes("w-full p-4 mt-2").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                if _BD_INFLIGHT.get(sid, 0) > time.monotonic() - 600:
                    with ui.row().classes("items-center gap-2"):
                        ui.spinner(size="18px", color="primary")
                        ui.label("Analysing the Act — appears here automatically when ready.").classes("text-xs text-gray-600")
                    state["bd_pending"] = sid
                else:
                    ui.button("⚡ Generate AI analysis", on_click=lambda s=sid: gen_statute_analysis(s)).props("color=primary rounded")
                    ui.label("Extracts the Act's purpose, key sections, definitions, scope, and amendments.").classes("text-xs text-gray-400 mt-2")

    def _render_statute_analysis(an, sid):
        with ui.column().classes("w-full gap-1"):
            if an.get("long_title"):
                sec("Long Title")
                ui.label(an["long_title"]).classes("body-text mb-2").style("font-style: italic;")
            if an.get("purpose"):
                sec("Purpose")
                ui.label(an["purpose"]).classes("body-text mb-3").style("white-space: pre-line;")
            provs = an.get("key_provisions") or []
            if provs:
                sec("Key Provisions")
                for p in provs:
                    if isinstance(p, dict):
                        s, e = p.get("section", ""), p.get("effect", "")
                        with ui.row().classes("items-baseline gap-2 no-wrap w-full pl-1 mb-1"):
                            if s:
                                ui.label(s).classes("text-[11px] font-bold text-primary flex-shrink-0").style("min-width: 62px;")
                            ui.label(e).classes("body-text").style("flex-grow:1; min-width:0;")
                    else:
                        ui.label(f"• {p}").classes("body-text pl-2 mb-1")
            defs = an.get("definitions") or []
            if defs:
                sec("Defined Terms")
                for d in defs:
                    if isinstance(d, dict) and d.get("term"):
                        with ui.row().classes("items-baseline gap-2 no-wrap w-full pl-1 mb-1"):
                            ui.label(d["term"]).classes("text-[11px] font-bold text-gray-700 flex-shrink-0").style("min-width: 62px;")
                            ui.label(d.get("meaning", "")).classes("body-text").style("flex-grow:1; min-width:0;")
            if an.get("scope"):
                sec("Scope & Application")
                ui.label(an["scope"]).classes("body-text mb-2")
            if an.get("amendments"):
                sec("Amendments / Repeals")
                ui.label(an["amendments"]).classes("body-text mb-2")
            if not demo:
                with ui.row().classes("w-full justify-end items-center"):
                    ui.button("Regenerate", on_click=lambda s=sid: gen_statute_analysis(s)).props("flat dense color=primary icon=refresh").classes("text-xs")

    async def gen_statute_analysis(sid):
        if state.get("bd_pending") == sid:
            return
        if not _ai_allowed():
            return
        if not _llm_acquire():
            ui.notify("The AI engine is busy — please try again in a minute.", type="warning")
            return
        _BD_INFLIGHT[sid] = time.monotonic()
        state["bd_pending"] = sid
        render_breakdown()  # show the spinner state
        try:
            from src.analyze_smart import analyze_statute
            provider = app.storage.user.get("llm_provider")
            model = app.storage.user.get("llm_model")
            api_key = app.storage.user.get("custom_api_key")
            await run.io_bound(analyze_statute, sid, True, None, provider, model, api_key)
        except Exception as e:  # noqa: BLE001
            print(f"Statute analysis failed for {sid}: {e}")
            try:
                ui.notify(f"Analysis failed: {e}", type="negative")
            except Exception:
                pass
        finally:
            _llm_release()
            _BD_INFLIGHT.pop(sid, None)
            state["bd_pending"] = None
        if state.get("statute") and (state["statute"].get("statute_id") or state["statute"].get("id")) == sid:
            render_breakdown()

    def render_breakdown():
        breakdown_pane.clear()
        with breakdown_pane:
            st = state.get("statute")
            if st:
                ui.label("Legislation").classes("pane-head")
                render_statute_info(st)
                return
            ui.label("Analysis & Breakdown").classes("pane-head")
            if not state["case"]:
                ui.label("Select a case from the library to view its analysis.").classes("text-gray-400 mt-4 text-sm")
                return
            
            cn = state["case"][0]
            m = case_meta(cn)

            # Resolve the breakdown FIRST — the header card is intentionally
            # minimal before generation (just Judgement-by + Parties) and only
            # expands to the full record once an analysis exists. Prefer a
            # just-generated result (incl. an uncached hollow one), else cache.
            fresh = state.get("bd_fresh")
            quality = None
            if fresh and fresh[0] == cn:
                bd, quality = fresh[1], fresh[2]
            else:
                bd = get_breakdown(cn)
            generated = bd is not None

            authoring_judges = m.get("judges") or []
            decided_date = m.get("date") or "Date not available"
            authoring_str = (", ".join(authoring_judges) if authoring_judges
                             else (author_for(cn) or "Authoring judge not specified"))

            username = _safe_username()
            bookmarked = is_bookmarked(username, cn)

            def render_parties_block():
                """Parties as a structured list (petitioners → respondents), falling
                back to the raw scrape paragraph when unparseable."""
                plist = parties_for(cn)
                raw = m.get("parties")
                if not plist and (not raw or raw == NOT_AVAILABLE):
                    return
                with ui.column().classes("gap-0.5 w-full"):
                    ui.label("Parties").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                    if not plist:
                        ui.label(raw).classes("text-xs font-medium text-gray-700 leading-normal case-title")
                        return
                    _SIDE_HEAD = {"petitioner": "Petitioners / Appellants",
                                  "respondent": "Respondents", "": "Named parties"}
                    for side in ("petitioner", "respondent", ""):
                        group = [e for e in plist if e.get("side", "") == side]
                        if not group:
                            continue
                        ui.label(_SIDE_HEAD[side]).classes("text-[10px] font-bold text-secondary uppercase tracking-wide mt-1")
                        for i, e in enumerate(group, 1):
                            with ui.row().classes("items-baseline gap-1.5 no-wrap w-full pl-1"):
                                ui.label(f"{i}.").classes("text-[11px] text-gray-400 flex-shrink-0")
                                with ui.column().classes("gap-0"):
                                    ui.label(e["name"]).classes("text-xs font-medium text-gray-800 leading-snug")
                                    if e.get("role"):
                                        ui.label(e["role"]).classes("text-[10px] text-gray-500 italic leading-none")

            # --- Case-information header card ---
            with ui.card().classes("w-full p-4 mb-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none; background: #ffffff;"):
                # Bookmark Toggle Row
                with ui.row().classes("w-full items-center justify-between no-wrap mb-2"):
                    ui.label("Case Information").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                    with ui.row().classes("items-center gap-1.5"):
                        # Download PDF judgment button (especially useful on mobile)
                        fn = m.get("filename")
                        if fn:
                            ui.button(on_click=lambda: ui.download(f"/pdf/{fn}")).props("flat round dense icon=download color=primary").classes("text-sm").tooltip("Download PDF judgment")
                        
                        def on_bookmark():
                            status = toggle_bookmark(username, cn)
                            bookmark_btn.props(f"icon={'bookmark' if status else 'bookmark_border'}")
                            ui.notify("Added to Bookmarks" if status else "Removed from Bookmarks", color="primary" if status else "grey-7")
                            refresh_bookmarks_ui()
                        bookmark_btn = ui.button(on_click=on_bookmark).props(f"flat round dense icon={'bookmark' if bookmarked else 'bookmark_border'} color=primary").classes("text-sm").tooltip("Bookmark case")

                if generated:
                    # Bench (Coram) — only after generation (full record view)
                    bench = bench_for(cn, m)
                    bench_str = ", ".join(bench) if bench else "Bench information not available in source"
                    with ui.row().classes("items-start gap-2 mt-1 w-full no-wrap"):
                        ui.icon("gavel", size="18px", color="primary").classes("mt-0.5 w-5 text-center flex-shrink-0")
                        with ui.column().classes("gap-0.5"):
                            ui.label("Bench (Coram)").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                            ui.label(bench_str).classes("text-xs font-semibold text-gray-800")

                # Authoring Judge — always visible (pre- and post-generation)
                with ui.row().classes("items-start gap-2 mt-2 w-full no-wrap"):
                    ui.icon("edit_note", size="18px", color="secondary").classes("mt-0.5 w-5 text-center flex-shrink-0")
                    with ui.column().classes("gap-0.5"):
                        ui.label("Judgement Delivered By").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                        ui.label(authoring_str).classes("text-xs font-semibold text-gray-800")

                if generated:
                    # Decided Date — full record view only
                    with ui.row().classes("items-start gap-2 mt-2 w-full no-wrap"):
                        ui.icon("calendar_today", size="18px", color="grey-6").classes("mt-0.5 w-5 text-center flex-shrink-0")
                        with ui.column().classes("gap-0.5"):
                            ui.label("Decided On").classes("text-[10px] uppercase tracking-wider font-bold text-gray-400")
                            ui.label(decided_date).classes("text-xs font-semibold text-gray-800")

                ui.separator().classes("my-3")
                # Parties — always visible, as a structured list
                render_parties_block()

            if generated and m.get("keywords"):
                ui.label("Keywords").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mt-2 pl-1")
                chips(m["keywords"][:14], "amber-3", "primary")
            if quality and quality.get("hollow"):
                with ui.card().classes("w-full p-3 mt-2 mb-1").style("border-radius: 10px; border: 1px solid #f0c36d; background: #fff8e6; box-shadow: none;"):
                    with ui.row().classes("items-center gap-2 no-wrap"):
                        ui.icon("warning_amber", color="orange-9", size="20px")
                        with ui.column().classes("gap-0.5"):
                            ui.label("Low-confidence analysis").classes("text-xs font-bold text-orange-9")
                            ui.label(f"The model returned mostly placeholders ({quality['filled']}/{quality['total']} sections filled) — it likely could not digest this judgment. This result was NOT saved; try Regenerate.").classes("text-[11px] text-gray-600 leading-snug")
                    if not demo:
                        ui.button("Regenerate Analysis", on_click=lambda: gen_breakdown(cn)).props("flat dense color=orange-9 icon=refresh").classes("text-xs mt-1")
            if not bd:
                # Pre-generation: the header above already shows Judgement-by +
                # Parties — this card only carries the generate action. Bench,
                # dates, keywords, legislation etc. appear after generation.
                with ui.card().classes("w-full p-4 mt-2").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                    if demo:
                        ui.label("🔒 Full AI analysis (facts · issues · ratio · precedents) is available after login.").classes("text-sm text-gray-500 mb-2")
                        ui.button("Log in to Unlock", on_click=lambda: ui.navigate.to("/login")).props("color=primary rounded unevaluated")
                    elif _BD_INFLIGHT.get(cn, 0) > time.monotonic() - 600:
                        # A run is in flight (phone tab reloads mid-generation land here).
                        with ui.row().classes("items-center gap-2"):
                            ui.spinner(size="18px", color="primary")
                            ui.label("Analysis is being generated — it appears here automatically "
                                     "when ready (~2 min). Keep the screen on, or check back."
                                     ).classes("text-xs text-gray-600")
                        state["bd_pending"] = cn  # let the poll timer render it on arrival
                    else:
                        ui.button("⚡ Generate AI analysis", on_click=lambda: gen_breakdown(cn)).props("color=primary rounded")
                        ui.label("Extracts facts, ratio, citations, and legislation using the local GGUF model.").classes("text-xs text-gray-400 mt-2")
                return

            if bd.get("topics_discussed"):
                ui.label("Topics Discussed").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mt-2 pl-1")
                chips(bd["topics_discussed"][:10], "amber-3", "primary")
            
            # Detailed AI Analysis blocks styled as neat cards
            with ui.card().classes("w-full p-4 mt-3").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                sec("Facts / Factual Matrix")
                ui.label(bd.get("factual_matrix") or "—").classes("body-text mb-3").style(
                    "white-space: pre-line;")  # keep the narrative's paragraph breaks
                
                if bd.get("legal_issues"):
                    sec("Legal Issues")
                    for li in bd["legal_issues"][:6]:
                        ui.label(f"• {li.get('question') if isinstance(li, dict) else li}").classes("body-text pl-2 mb-1")
                
                sec("Ratio Decidendi")
                ui.label(bd.get("ratio_decidendi") or "—").classes("body-text mb-3")
                
                if bd.get("deciding_factors"):
                    sec("Deciding Factors")
                    for d in bd["deciding_factors"][:8]:
                        ui.label(f"• {d}").classes("body-text pl-2 mb-1")
                
                if bd.get("precedent_index"):
                    sec("Citations & Distinctions")
                    for p in bd["precedent_index"][:12]:
                        cited, tr = p.get("cited_case", ""), p.get("treatment", "")
                        if not cited:
                            continue
                        target = find_case(cited, active_case_no=cn)  # resolves to a case in the repository, if present
                        with ui.row().classes("items-center gap-2 py-1 no-wrap pl-2"):
                            if tr and tr != NOT_AVAILABLE:
                                ui.badge(tr, color="amber-3", text_color="grey-9").classes("text-[10px] px-2 py-0.5 rounded")
                            if target:
                                ui.link(cited, "#").classes("text-xs font-semibold no-underline text-primary").on("click", lambda t=target: open_case(t))
                                ui.icon("open_in_new", size="13px").classes("text-primary flex-shrink-0").tooltip("Open this judgement in the library")
                            else:
                                ui.label(cited).classes("text-xs font-semibold text-gray-500").tooltip("Not in corpus")
                        note = (p.get("note") or "").strip()
                        if note and note != NOT_AVAILABLE:
                            ui.label(note).classes("text-[11px] text-gray-500 italic pl-8 -mt-1 mb-1 leading-snug")

                raw_leg = m.get("legislation") or []
                if isinstance(raw_leg, str):
                    raw_leg = [raw_leg]
                else:
                    raw_leg = list(raw_leg)
                    
                bd_leg = bd.get("legislation_cited") or []
                bd_leg = list(bd_leg)
                
                # Retrieve cited legislation from connected cases in the corpus to be thorough
                connected_cases = get_connected_cases_in_corpus(cn)
                for ccn in connected_cases:
                    con = _con()
                    cc_row = con.execute("SELECT legislation FROM judgements WHERE case_no=? LIMIT 1", (ccn,)).fetchone()
                    cc_bd_row = con.execute("SELECT json FROM analyses WHERE case_no=? LIMIT 1", (ccn,)).fetchone()
                    con.close()
                    
                    if cc_row and cc_row[0]:
                        try:
                            import json
                            cc_leg = json.loads(cc_row[0]) if isinstance(cc_row[0], str) and cc_row[0].startswith('[') else [cc_row[0]]
                            raw_leg.extend(cc_leg)
                        except Exception:
                            raw_leg.append(cc_row[0])
                            
                    if cc_bd_row and cc_bd_row[0]:
                        try:
                            import json
                            cc_bd = json.loads(cc_bd_row[0])
                            if isinstance(cc_bd, dict) and cc_bd.get("legislation_cited"):
                                bd_leg.extend(cc_bd["legislation_cited"])
                        except Exception:
                            pass
                
                normalized_leg = []
                seen_keys = set()
                
                for item in raw_leg:
                    if isinstance(item, str) and item.strip():
                        key = item.strip().lower()
                        if key not in seen_keys:
                            seen_keys.add(key)
                            normalized_leg.append({"statute": item.strip(), "section": "", "interpretation": ""})
                            
                for item in bd_leg:
                    if isinstance(item, dict):
                        statute = (item.get("statute") or "").strip()
                        section = (item.get("section") or "").strip()
                        interpretation = (item.get("interpretation") or "").strip()
                    elif isinstance(item, str):
                        statute = item.strip()
                        section = ""
                        interpretation = ""
                    else:
                        continue
                        
                    if statute:
                        key = f"{statute} {section}".strip().lower()
                        if key not in seen_keys:
                            seen_keys.add(key)
                            normalized_leg.append({"statute": statute, "section": section, "interpretation": interpretation})
                            
                if normalized_leg:
                    sec("Legislation Cited")
                    for leg_item in normalized_leg[:12]:
                        statute = leg_item["statute"]
                        sec_num = leg_item["section"]

                        if sec_num:
                            display_text = f"{sec_num} of {statute}"
                            search_term = f"{statute} {sec_num}"
                        else:
                            display_text = statute
                            search_term = statute

                        # Resolve to a corpus Act, mirroring how precedents resolve to
                        # a corpus case via find_case() above. resolve_statute() matches
                        # on Act identity (name / No.+year / contained Act name), so it
                        # handles "Section 304 of the Penal Code" -> "Penal Code".
                        statute_id = resolve_statute(statute)
                        with ui.row().classes("items-center gap-1.5 py-0.5 no-wrap pl-2"):
                            if statute_id:
                                # In corpus → open the actual statute text on click.
                                ui.icon("article", size="15px").classes("text-primary flex-shrink-0")
                                ui.link(display_text, "#").classes(
                                    "text-xs font-medium text-primary no-underline hover:underline"
                                ).on("click", lambda s=statute_id: open_statute(s))
                                ui.icon("open_in_new", size="13px").classes("text-primary flex-shrink-0").tooltip("Open this statute")
                            else:
                                ui.icon("article", size="15px").classes("text-gray-400 flex-shrink-0")
                                ui.label(display_text).classes("text-xs font-medium text-gray-500").tooltip("Not in corpus")
                                ui.button(icon="search", on_click=lambda t=search_term: goto(t)).props(
                                    "flat round dense size=xs color=grey-6"
                                ).classes("hover:text-primary").tooltip("Search corpus for cases citing this")
                
                sec("Final Order")
                ui.label(bd.get("final_order") or "—").classes("body-text mb-2")
                
                if not demo and bd.get("academic_synthesis") and bd["academic_synthesis"] != NOT_AVAILABLE:
                    from src.analyze_smart import clean_academic_synthesis
                    sec("What ROS says")
                    synthesis_text = clean_academic_synthesis(bd["academic_synthesis"])
                    ui.markdown(synthesis_text).classes("synthesis-md mb-2")
                if not demo:
                    # --- Document Reference: compare an uploaded document against this case ---
                    sec("Document Reference")
                    ui.label("Upload a document (a brief, draft pleading, or facts summary) and ROS will "
                             "compare its facts with this judgment and assess how the ruling applies."
                             ).classes("text-xs text-gray-500 mb-1")
                    docref_out = ui.column().classes("w-full")
                    docref_key = f"docref2::{cn}"

                    def render_docref(data, _out=docref_out):
                        """Compact verdict: % badge + side-by-side table, no essay."""
                        with _out:
                            pct = max(0, min(100, int(data.get("match_percent") or 0)))
                            related = bool(data.get("related")) and pct >= 15
                            if not related:
                                pct = min(pct, 15)
                            color, tcolor = (("green-2", "green-9") if pct >= 70 else
                                             ("amber-2", "amber-9") if pct >= 40 else
                                             ("orange-2", "orange-10") if pct >= 15 else
                                             ("grey-3", "grey-8"))
                            with ui.row().classes("items-center gap-2 mb-1 no-wrap"):
                                ui.badge(f"{pct}% match", color=color, text_color=tcolor).classes(
                                    "text-sm font-bold px-3 py-1").style("border-radius: 6px;")
                                if not related:
                                    ui.badge("No meaningful relation", color="grey-3", text_color="grey-8"
                                             ).classes("text-xs px-2 py-1").style("border-radius: 6px;")
                            verdict = str(data.get("verdict") or "").strip()
                            if verdict:
                                ui.label(verdict).classes("text-xs text-gray-700 mb-1 leading-snug")
                            rows = data.get("comparison") or []
                            if rows and related:
                                def _cell(v):
                                    return str(v or "—").replace("|", "/").strip()
                                md = ("| Aspect | This judgment | Your document | Match |\n"
                                      "|---|---|---|---|\n")
                                for r in rows[:8]:
                                    md += (f"| **{_cell(r.get('aspect'))}** | {_cell(r.get('case'))} "
                                           f"| {_cell(r.get('document'))} | {_cell(r.get('match'))} |\n")
                                ui.markdown(md).classes("docref-table mb-1").style("font-size: 0.8rem;")
                            applic = str(data.get("applicability") or "").strip()
                            if applic and related:
                                ui.label(f"Applicability: {applic}").classes(
                                    "text-xs text-gray-600 italic mb-2 leading-snug")

                    if isinstance(state.get(docref_key), dict):
                        render_docref(state[docref_key])

                    async def on_docref_upload(e, _cn=cn, _bd=bd, _out=docref_out, _key=docref_key):
                        name, data = e.file.name, await e.file.read()
                        if not _ai_allowed():
                            return
                        if not _llm_acquire():
                            ui.notify("The AI engine is busy with other requests — try again shortly.",
                                      type="warning")
                            return
                        _out.clear()
                        with _out:
                            with ui.row().classes("items-center gap-2 my-2"):
                                ui.spinner(size="18px", color="primary")
                                ui.label(f"Reading “{name}” and comparing with this judgment…"
                                         ).classes("text-xs text-gray-500")
                        try:
                            provider = app.storage.user.get("llm_provider")
                            model = app.storage.user.get("llm_model")
                            api_key = app.storage.user.get("custom_api_key")
                            def _compare() -> dict:
                                if name.lower().endswith((".txt", ".md")):
                                    doc_text = data.decode("utf-8", errors="replace")
                                else:
                                    from src.parsing import pages_from_bytes
                                    doc_text = "\n".join(pages_from_bytes(name, data))
                                doc_text = doc_text.strip()[:6000]
                                if len(doc_text) < 80:
                                    return {"match_percent": 0, "related": False,
                                            "verdict": "Could not read enough text from the uploaded file — "
                                                       "if it is a scanned image PDF, try a text PDF or .txt summary."}
                                issues = "; ".join(
                                    (i.get("question") if isinstance(i, dict) else str(i))
                                    for i in (_bd.get("legal_issues") or [])[:6])
                                sys_p = (
                                    "You are ROS, a Sri Lankan legal research assistant doing a strict, "
                                    "grounded comparison. Return ONLY a JSON object, no prose:\n"
                                    '{"match_percent": <int 0-100>, "related": <bool>, '
                                    '"verdict": "<plain-language conclusion, max 25 words>", '
                                    '"applicability": "<how the ruling applies to the document, max 30 words>", '
                                    '"comparison": [{"aspect": "...", "case": "...", "document": "...", '
                                    '"match": "High|Partial|None"}]}\n'
                                    "Rules:\n"
                                    "- 4-7 comparison rows over: parties/relationship, cause of action, key facts, "
                                    "legal issues, statute involved, remedy sought. Each cell max 12 words, drawn "
                                    "ONLY from the provided texts; write 'Not present' when a side lacks it.\n"
                                    "- match_percent = overlap of MATERIAL facts and legal issues: 0-15 unrelated, "
                                    "15-40 weak, 40-70 partial, 70-100 strong.\n"
                                    "- If the document has no meaningful factual or legal connection to the case: "
                                    "related=false, match_percent<=15, verdict must say plainly there is no "
                                    "meaningful relation, applicability='None'. NEVER invent or stretch "
                                    "similarities; an honest low score is the correct answer.")
                                usr = (
                                    f"=== DECIDED CASE: {_cn} ===\n"
                                    f"FACTS:\n{(_bd.get('factual_matrix') or '')[:2500]}\n\n"
                                    f"LEGAL ISSUES:\n{issues}\n\n"
                                    f"RATIO DECIDENDI:\n{(_bd.get('ratio_decidendi') or '')[:1500]}\n\n"
                                    f"FINAL ORDER:\n{(_bd.get('final_order') or '')[:600]}\n\n"
                                    f"=== UPLOADED DOCUMENT ({name}) ===\n{doc_text}\n\n"
                                    "Compare the two strictly per the rules. Return ONLY the JSON object.")
                                from src.analyze_smart import _chat, _extract_json
                                return _extract_json(_chat(sys_p, usr, max_tokens=900, json_mode=True, provider=provider, model=model, api_key=api_key))

                            result = await run.io_bound(_compare)
                            state[_key] = result
                            _out.clear()
                            render_docref(result)
                        except Exception as ex:  # noqa: BLE001
                            _out.clear()
                            ui.notify(f"Could not compare the document: {ex}", type="negative")
                        finally:
                            _llm_release()

                    ui.upload(label="Upload document to compare (.pdf / .docx / .txt)", auto_upload=True,
                              max_file_size=15_000_000, on_upload=on_docref_upload
                              ).props('accept=".pdf,.docx,.txt,.md" flat bordered color=primary'
                              ).classes("w-full mb-2")

                    with ui.row().classes("w-full justify-start items-center mt-2 gap-2"):
                        ui.button("Regenerate Analysis", on_click=lambda: gen_breakdown(cn)).props("flat dense color=primary icon=refresh").classes("text-xs font-semibold").style("margin-left: -8px;")
                        ts = get_breakdown_metadata(cn)
                        if ts:
                            try:
                                from datetime import datetime
                                dt = datetime.fromisoformat(ts)
                                ts_str = dt.strftime("%Y-%m-%d %H:%M")
                            except Exception:
                                ts_str = ts
                            ui.label(f"Last generated: {ts_str}").classes("text-[10px] text-gray-400 font-medium")
                            ui.icon("check_circle", color="green-4", size="12px").classes("opacity-60")

    # ----------------------------- navigation ----------------------------- #
    def open_case(case_no, page=None, is_back=False):
        if not is_back:
            stack = state.setdefault("history_stack", [])
            current = None
            if state.get("case"):
                current = {"type": "case", "id": state["case"][0], "page": state.get("page")}
            elif state.get("statute"):
                current = {"type": "statute", "id": state["statute"]["id"]}
            if current:
                if not stack or stack[-1] != current:
                    stack.append(current)
            if back_btn:
                back_btn.set_visibility(len(stack) > 0)

        con = _con()
        row = con.execute("SELECT case_no, filename FROM judgements WHERE case_no=? LIMIT 1", (case_no,)).fetchone()
        if not row:  # background renames can shift names under a cached tree
            row = con.execute("SELECT case_no, filename FROM judgements "
                              "WHERE case_no=? COLLATE NOCASE LIMIT 1", (case_no,)).fetchone()
        con.close()
        if not row:
            # Consolidated Acts / statutes sometimes surface via a case link —
            # route them to the legislation view rather than failing.
            if get_statute(case_no):
                open_statute(case_no, is_back=is_back)
                return
            ui.notify(f"Case not found: {case_no}", type="warning")
            return
        state["case"], state["page"] = row, page
        state["statute"] = None
        state["workspace_active"] = True
        update_workspace_visibility()
        if left_drawer:
            left_drawer.hide()
        render_pdf()
        render_breakdown()
        floating_chat_widget.refresh()
        set_active("bd")  # mobile: jump to the analysis when a case opens

    def open_statute(statute_id, is_back=False):
        """Open an in-corpus statute (the legislation analogue of open_case): load
        the Act's PDF into the document pane and its identity card into the centre
        pane. Statutes carry no judicial breakdown, so no analysis is generated."""
        if not is_back:
            stack = state.setdefault("history_stack", [])
            current = None
            if state.get("case"):
                current = {"type": "case", "id": state["case"][0], "page": state.get("page")}
            elif state.get("statute"):
                current = {"type": "statute", "id": state["statute"]["id"]}
            if current:
                if not stack or stack[-1] != current:
                    stack.append(current)
            if back_btn:
                back_btn.set_visibility(len(stack) > 0)

        st = get_statute(statute_id)
        if not st:
            ui.notify(f"Statute not found: {statute_id}", type="warning")
            return
        state["statute"] = st
        state["case"], state["page"] = None, None
        state["workspace_active"] = True
        update_workspace_visibility()
        if left_drawer:
            left_drawer.hide()
        render_pdf()
        render_breakdown()
        floating_chat_widget.refresh()  # judgment-only chat hides while a statute is open
        set_active("pdf")  # the statute text itself is the content

    def go_back():
        stack = state.setdefault("history_stack", [])
        if not stack:
            return
        prev = stack.pop()
        if back_btn:
            back_btn.set_visibility(len(stack) > 0)
        if prev["type"] == "case":
            open_case(prev["id"], prev.get("page"), is_back=True)
        elif prev["type"] == "statute":
            open_statute(prev["id"], is_back=True)

    # ----------------------------- Library -------------------------------- #
    def make_year(y, months_dict, on_leaf=None):
        on_leaf = on_leaf or open_case
        total_cases = sum(len(cases) for cases in months_dict.values())
        exp = ui.expansion().classes("w-full mb-1").props("dense header-class='bg-gray-50 text-gray-700 text-xs font-semibold rounded-md' expand-icon-class='text-gray-400'")
        with exp.add_slot('header'):
            with ui.row().classes("items-center justify-between w-full py-1.5 px-2"):
                ui.label(y).classes("text-xs font-bold text-gray-700")
                ui.badge(str(total_cases), color="grey-3", text_color="grey-8").classes("text-[10px] px-2 py-0.5 rounded-full")

        loaded = {"v": False}

        def load():
            if loaded["v"] or not exp.value:
                return
            loaded["v"] = True
            with exp:
                with ui.column().classes("w-full pl-2 gap-1 py-1"):
                    def render_case_rows(cases):
                        for cn, _date in sorted(cases, key=lambda c: c[1], reverse=True):
                            with ui.row().classes("w-full items-start justify-between py-1.5 px-2 hover:bg-gray-100 rounded cursor-pointer transition-colors duration-150 gap-2 no-wrap").on("click", lambda c=cn: on_leaf(c)):
                                ui.label(cn).classes("text-[11px] font-medium text-gray-800").style("word-break: break-word; flex-grow: 1; min-width: 0;")
                                if _date:
                                    ui.label(_date).classes("text-[9px] text-gray-400 flex-shrink-0")

                    def month_sort_key(m_name):
                        try:
                            return _MONTH_NAMES.index(m_name)
                        except ValueError:
                            return 12

                    sorted_months = [m for m in sorted(months_dict.keys(), key=month_sort_key)
                                     if months_dict[m]]
                    # A single sub-group ('80 NLR', 'SLR 2021', 'Digest entries',
                    # 'Unknown Month') is pure nesting noise — flatten it away.
                    if len(sorted_months) == 1:
                        render_case_rows(months_dict[sorted_months[0]])
                        return
                    for m_name in sorted_months:
                        cases = months_dict[m_name]
                        # Sub-expansion only when a year truly has multiple groups
                        m_exp = ui.expansion().classes("w-full pl-1 mb-0.5").props("dense header-class='text-gray-600 text-[11px] font-medium' expand-icon-class='text-gray-400'")
                        with m_exp.add_slot('header'):
                            with ui.row().classes("items-center justify-between w-full py-1 px-1"):
                                ui.label(m_name).classes("text-[11px] font-semibold text-gray-600")
                                ui.badge(str(len(cases)), color="grey-2", text_color="grey-6").classes("text-[9px] px-1.5 py-0.2 rounded-full")

                        with m_exp:
                            with ui.column().classes("w-full pl-2 gap-1 py-1"):
                                render_case_rows(cases)

        exp.on_value_change(load)

    def show_tree():
        results.clear()
        with results:
            ui.label("Browse the Library").classes("text-xs font-semibold text-gray-500 mb-1 mt-1")
            mode = {"v": "Courts"}
            counts = {s: sum(len(cs) for yr in judgements_by_report()[s].values() for cs in yr.values())
                      for s in ("NLR", "SLR", "Digest")}
            act_count = sum(len(cs) for yr in statutes_by_year().values() for cs in yr.values())
            
            buttons = {}
            def select_mode(val):
                mode["v"] = val
                for k, btn in buttons.items():
                    if k == val:
                        btn.props("color=primary text-color=white unelevated")
                    else:
                        btn.props("color=grey-2 text-color=grey-8 unelevated")
                render()

            with ui.column().classes("w-full gap-2 mb-3 mt-1"):
                # Row 1: Courts
                with ui.row().classes("w-full"):
                    buttons["Courts"] = ui.button("Supreme Court", on_click=lambda: select_mode("Courts")) \
                        .classes("w-full no-caps text-xs py-2 font-medium")
                # Row 2: NLR & SLR
                with ui.row().classes("w-full no-wrap gap-2"):
                    buttons["NLR"] = ui.button(f"NLR ({counts['NLR']:,})", on_click=lambda: select_mode("NLR")) \
                        .classes("w-1/2 no-caps text-xs py-2 font-medium")
                    buttons["SLR"] = ui.button(f"SLR ({counts['SLR']:,})", on_click=lambda: select_mode("SLR")) \
                        .classes("w-1/2 no-caps text-xs py-2 font-medium")
                # Row 3: Digests & Acts
                with ui.row().classes("w-full no-wrap gap-2"):
                    buttons["Digest"] = ui.button(f"Digests ({counts['Digest']:,})", on_click=lambda: select_mode("Digest")) \
                        .classes("w-1/2 no-caps text-xs py-2 font-medium")
                    buttons["Acts"] = ui.button(f"Acts ({act_count:,})", on_click=lambda: select_mode("Acts")) \
                        .classes("w-1/2 no-caps text-xs py-2 font-medium")

            body = ui.column().classes("w-full")

            def render():
                body.clear()
                with body:
                    if mode["v"] == "Acts":
                        by = statutes_by_year()
                        years = sorted((y for y in by if y != "Undated"), reverse=True)
                        if "Undated" in by:
                            years.append("Undated")
                        if not years:
                            ui.label("No Acts indexed yet.").classes("text-sm text-gray-500 mt-2")
                        for y in years:
                            make_year(y, by[y], on_leaf=open_statute)
                    elif mode["v"] in ("NLR", "SLR", "Digest"):
                        # Report shelf: year -> volume ('76 NLR' under 1976) -> cases;
                        # the Digest shelf is alphabetical (A-Z letters instead of years).
                        series = judgements_by_report().get(mode["v"], {})
                        if not series:
                            ui.label(f"No {mode['v']} cases indexed yet — the ingest may still be running.").classes(
                                "text-sm text-gray-500 mt-2")
                        for y in sorted(series, reverse=(mode["v"] != "Digest")):
                            make_year(y, series[y])
                    else:
                        by = judgements_by_year_month()
                        years = sorted((y for y in by if y != "Undated"), reverse=True)
                        if "Undated" in by:
                            years.append("Undated")
                        for y in years:
                            make_year(y, by[y])

            # Initialize colors
            select_mode("Courts")

    @ui.refreshable
    def render_results_ui():
        hits = state.get("search_results", [])
        page = state.get("search_page", 1)
        pretty_label = state.get("search_label", "")
        
        PAGE_SIZE = 50
        total = len(hits)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        
        if page > total_pages:
            page = total_pages
        if page < 1:
            page = 1
        state["search_page"] = page

        start = (page - 1) * PAGE_SIZE
        end = start + PAGE_SIZE
        page_hits = hits[start:end]

        with ui.row().classes("items-center justify-between w-full mb-2"):
            ui.label(f"{pretty_label}  ·  {total} cases").classes("text-xs font-semibold text-gray-500")
            ui.button("Reset", on_click=reset).props("flat dense color=primary icon=restart_alt").classes("text-xs")
            
        if not hits:
            ui.label("No judgements match these criteria.").classes("text-sm text-gray-500 mt-2")
            return

        for h in page_hits:
            cn = h["case_no"]
            dt = h.get("date", "")
            snip = h.get("snippet", "")
            why = h.get("why", "")
            page_anchor = h.get("page")
            with ui.card().classes("w-full p-3 mb-2 cursor-pointer hover:shadow-md transition-shadow duration-200").on("click", lambda c=cn, pg=page_anchor: open_case(c, page=pg)).style("border-radius: 8px; border: 1px solid #e8eaed; box-shadow: none;"):
                with ui.row().classes("justify-between items-start w-full no-wrap gap-2"):
                    ui.label(cn).classes("text-xs font-bold text-primary").style("word-break: break-word; flex-grow: 1; min-width: 0;")
                    with ui.row().classes("items-center gap-1 no-wrap"):
                        if why == "semantic":
                            ui.badge("✦ AI", color="purple-1", text_color="deep-purple").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;").tooltip("Semantically related — the exact words may not appear")
                        elif why == "broad":
                            ui.badge("partial", color="grey-3", text_color="grey-8").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;").tooltip("Matches some of your search terms")
                        if page_anchor:
                            ui.badge(f"p.{page_anchor}", color="grey-2", text_color="grey-8").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;").tooltip(f"Opens the PDF at page {page_anchor}, where your search matched")
                        if dt:
                            ui.badge(dt[:4], color="blue-1", text_color="primary").classes("text-[10px] px-2 py-0.5").style("border-radius: 4px; box-shadow: none;")
                if snip:
                    ui.label(snip).classes("text-[11px] text-gray-600 mt-1 line-clamp-2")

        if total_pages > 1:
            with ui.row().classes("w-full items-center justify-center gap-2 mt-4 pb-4"):
                prev_btn = ui.button(icon="chevron_left", on_click=lambda: change_page(-1)).props("flat round dense color=primary")
                if page <= 1:
                    prev_btn.props("disabled")
                
                ui.label(f"{page} / {total_pages}").classes("text-xs font-bold text-gray-600 px-2")
                
                next_btn = ui.button(icon="chevron_right", on_click=lambda: change_page(1)).props("flat round dense color=primary")
                if page >= total_pages:
                    next_btn.props("disabled")

    def change_page(delta):
        state["search_page"] = state.get("search_page", 1) + delta
        render_results_ui.refresh()

    def show_results(hits, label):
        state["search_results"] = hits
        state["search_page"] = 1
        state["search_label"] = label
        results.clear()
        with results:
            render_results_ui()

    # ---- combinable filters: Justice · legal area · year · month · search ---- #
    _FILTER_ICON = {"judge": "gavel", "area": "category", "year": "event",
                    "month": "calendar_month", "query": "search"}

    def current_filters() -> dict:
        """Live value of every facet control (None when unset)."""
        return {
            "judge": (judge_sel.value or None),
            "area": (area_sel.value or None),
            "year": (year_sel.value or None),
            "month": (month_sel.value or None),
            "query": ((q.value or "").strip() or None),
        }

    def render_active_filters(active):
        active_filters.clear()
        if not active:
            return
        with active_filters:
            ui.label("Filters:").classes("text-[10px] uppercase font-bold tracking-wider text-gray-400 mr-1 self-center")
            for k, v in active:
                disp = f'"{v}"' if k == "query" else (" · ".join(v) if isinstance(v, list) else str(v))
                with ui.row().classes("items-center gap-0.5 bg-blue-1 rounded-full pl-2 pr-0.5 py-0.5 no-wrap"):
                    ui.icon(_FILTER_ICON.get(k, "filter_alt"), size="13px").classes("text-primary")
                    ui.label(disp).classes("text-[10px] font-semibold text-primary truncate").style("max-width: 110px;")
                    ui.button(icon="close", on_click=lambda key=k: clear_one(key)).props("flat round dense size=xs color=primary")

    def clear_one(key):
        state["_resetting"] = True
        if key == "query":
            q.value = ""
        elif key == "judge":
            judge_sel.value = []          # multi-select clears to an empty list
        else:
            {"area": area_sel, "year": year_sel, "month": month_sel}[key].value = None
        state["_resetting"] = False
        apply_filters()

    def apply_filters(deep: bool = False):
        """Re-run the AND-combined facet query; repaint results + active-filter chips.
        deep=True (Enter / ✦ button) also merges AI semantic matches once the
        embedder is warm — typing stays on the fast FTS path."""
        if state.get("_resetting"):
            return
        f = current_filters()
        active = [(k, v) for k, v in f.items() if v]
        render_active_filters(active)
        if not active:
            show_tree()
            return
            
        deep = deep and bool(f.get("query")) and embedder_ready()
        hits = combined_search(**f, semantic=deep, limit=10000)
        
        state["workspace_active"] = True
        update_workspace_visibility()
        if left_drawer:
            left_drawer.show()
            
        pretty = " · ".join((f'"{v}"' if k == "query" else (" / ".join(v) if isinstance(v, list) else str(v))) for k, v in active)
        if deep:
            pretty += "  ·  ✦ deep"
        show_results(hits, f"{pretty}  ·  {len(hits)} cases")

    def reset_controls_silently():
        state["_resetting"] = True
        judge_sel.value = []             # multi-select resets to an empty list
        for ctrl in (area_sel, year_sel, month_sel):
            ctrl.value = None
        q.value = ""
        state["_resetting"] = False



    def goto(term):
        """Jump the Library to a topic / keyword / Act clicked in the breakdown."""
        reset_controls_silently()
        q.value = term or ""
        apply_filters()
        set_active("library")  # mobile: show the related results

    def reset():
        reset_controls_silently()
        render_active_filters([])
        show_tree()

    # ------------------------------ layout -------------------------------- #
    # Ported ROS Batch Ingestion and Metadata Extractor Portal (Dialog)
    def open_extractor_dialog():
        from metadata_extractor.extractor import run_extraction
        from metadata_extractor.models import JudgmentMetadata
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        import asyncio
        
        dialog = ui.dialog()
        with dialog, ui.card().style("width: 1200px; max-width: 95vw; height: 80vh; padding: 0; display: flex; flex-direction: column; background-color: #f8fafc; border-radius: 16px; overflow: hidden; box-shadow: 0 20px 25px -5px rgba(0,0,0,0.1);"):
            # Dialog Header: Slate Navy (#0f172a) with Gold Accent bottom border
            with ui.row().classes("w-full bg-[#0f172a] text-white p-4 items-center justify-between no-wrap").style("flex-shrink: 0; border-bottom: 2px solid #b8965a;"):
                with ui.row().classes("items-center gap-2"):
                    ui.icon("document_scanner", color="amber").classes("text-xl")
                    with ui.column().classes("gap-0"):
                        ui.label("Batch Ingestion & Metadata Extractor Portal").classes("text-sm font-bold tracking-wider text-white uppercase")
                        ui.label("Upload PDF judgments, run AI metadata extraction, and load cases to active search index").classes("text-[10px] text-slate-400")
                ui.button(icon="close", on_click=dialog.close).props("flat round dense color=white").classes("hover:bg-white/10")

            # Local Page State inside dialog closure
            queue = {}
            queue_bytes = {}
            table_rows = []
            results = []
            upload_state = {"count": 0, "timer": None}

            def refresh_table():
                results_table.rows = list(table_rows)
                results_table.update()

            def build_excel_bytes(extracted_records: list[dict]) -> bytes:
                wb = Workbook()
                ws = wb.active
                ws.title = "Registry Overview"
                ws.views.sheetView[0].showGridLines = True
                
                headers = ["Case Number", "Date of Judgment", "Appellants / Petitioners", "Respondents", "Judges", "Legislation Cited", "Keywords"]
                ws.append(headers)
                
                font_family = "Segoe UI"
                header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
                
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[1].height = 28
                
                data_font = Font(name=font_family, size=10)
                thin_border = Border(
                    left=Side(style='thin', color='DADCE0'), right=Side(style='thin', color='DADCE0'),
                    top=Side(style='thin', color='DADCE0'), bottom=Side(style='thin', color='DADCE0')
                )
                align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
                align_center = Alignment(horizontal="center", vertical="center")
                fill_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                for row_idx, record in enumerate(extracted_records, 2):
                    meta = record.get("metadata", {})
                    parties = meta.get("parties", {})
                    app_list = parties.get("appellants_petitioners", []) or []
                    resp_list = parties.get("respondents", []) or []
                    if abbreviate_parties_switch.value:
                        appellants = app_list[0] + " and other" if len(app_list) > 1 else ", ".join(app_list)
                        respondents = resp_list[0] + " and other" if len(resp_list) > 1 else ", ".join(resp_list)
                    else:
                        appellants = ", ".join(app_list) if isinstance(app_list, list) else ""
                        respondents = ", ".join(resp_list) if isinstance(resp_list, list) else ""
                    judges = ", ".join(meta.get("judges", [])) if isinstance(meta.get("judges"), list) else ""
                    leg_list = meta.get("legislation_cited", [])
                    if isinstance(leg_list, list):
                        leg_strs = []
                        for item in leg_list:
                            if isinstance(item, dict):
                                sec_num = item.get("section") or ""
                                statute = item.get("statute") or ""
                                leg_strs.append(f"{sec_num} of {statute}" if sec_num else statute)
                            else:
                                leg_strs.append(str(item))
                        legislation = ", ".join(leg_strs)
                    else:
                        legislation = ""
                    keywords = ", ".join(meta.get("keywords", [])) if isinstance(meta.get("keywords"), list) else ""
                    
                    row_data = [meta.get("case_number", ""), meta.get("date_of_judgment", ""), appellants, respondents, judges, legislation, keywords]
                    ws.append(row_data)
                    ws.row_dimensions[row_idx].height = 24
                    
                    is_even = (row_idx % 2 == 0)
                    for col_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = data_font
                        cell.border = thin_border
                        if is_even:
                            cell.fill = fill_even
                        if col_idx in (1, 2):
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left
                            
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        val = str(cell.value or "")
                        lines = val.split('\n')
                        for line in lines:
                            if len(line) > max_len:
                                max_len = len(line)
                    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
                    
                bytes_io = io.BytesIO()
                wb.save(bytes_io)
                return bytes_io.getvalue()

            # Main dialog layout grid
            with ui.row().classes("w-full no-wrap gap-4 p-4 items-start overflow-auto").style("flex-grow: 1; height: 100%;"):
                # Left settings + upload panel (35%)
                with ui.column().classes("w-[35%] gap-4 flex-shrink-0"):
                    # Settings card
                    with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                        ui.label("AI Settings").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                        
                        env_provider = "anthropic" if settings.anthropic_api_key and "anthropic" in settings.llm_provider.lower() else "llamacpp"
                        provider_sel = ui.select(
                            {"llamacpp": "Local GGUF (Llama.cpp)", "openai": "OpenAI / Groq API", "anthropic": "Anthropic Claude API"},
                            value=env_provider,
                            label="LLM Provider"
                        ).classes("w-full")
                        
                        env_model = settings.anthropic_model if env_provider == "anthropic" else settings.llamacpp_model_path
                        model_input = ui.input("Model or GGUF Path", value=env_model).classes("w-full")
                        
                        env_base_url = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                        base_url_input = ui.input("Custom API Base URL (Optional)", value=env_base_url, placeholder="e.g. https://api.groq.com/openai/v1").classes("w-full").props('autocomplete="off"')
                        
                        is_cur_admin = (app.storage.user.get("username") or "").lower() == "admin"
                        if is_cur_admin:
                            env_api_key = settings.anthropic_api_key if env_provider == "anthropic" else settings.openai_api_key if env_provider == "openai" else ""
                            if not env_api_key:
                                env_api_key = os.getenv("ANTHROPIC_API_KEY") or os.getenv("OPENAI_API_KEY") or ""
                        else:
                            env_api_key = ""
                        apikey_input = ui.input("API Key (Required for RO)", value=env_api_key, password=True).classes("w-full").props('autocomplete="new-password"')
                        
                        ui.label("Parallel Processing Threads").classes("text-[10px] text-gray-500 font-semibold mt-2")
                        workers_slider = ui.slider(min=1, max=10, value=3).props("label label-always")
                        workers_hint = ui.label("").classes("text-[10px] text-gray-400 mt-0.5")

                        def sync_workers_ui(provider):
                            if provider == "llamacpp":
                                workers_slider.disable()
                                workers_hint.set_text("Local GGUF runs one document at a time.")
                            else:
                                workers_slider.enable()
                                workers_hint.set_text("Processes up to N judgments in parallel.")

                        def on_provider_change(e):
                            sync_workers_ui(e.value)
                            if e.value == "llamacpp":
                                model_input.value = settings.llamacpp_model_path
                                base_url_input.value = ""
                                apikey_input.value = ""
                                base_url_input.disable()
                                apikey_input.disable()
                            elif e.value == "openai":
                                model_input.value = settings.llm_model if "openai" in settings.llm_provider.lower() else "gpt-4o-mini"
                                base_url_input.value = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                                base_url_input.enable()
                                apikey_input.enable()
                                is_cur_admin = (app.storage.user.get("username") or "").lower() == "admin"
                                apikey_input.value = (settings.openai_api_key if "openai" in settings.llm_provider.lower() else os.getenv("OPENAI_API_KEY") or "") if is_cur_admin else ""
                            else:  # anthropic
                                model_input.value = settings.anthropic_model or "claude-sonnet-5"
                                base_url_input.value = ""
                                base_url_input.disable()
                                apikey_input.enable()
                                is_cur_admin = (app.storage.user.get("username") or "").lower() == "admin"
                                apikey_input.value = (settings.anthropic_api_key if "anthropic" in settings.llm_provider.lower() else os.getenv("ANTHROPIC_API_KEY") or "") if is_cur_admin else ""
                                
                        provider_sel.on_value_change(on_provider_change)
                        sync_workers_ui(env_provider)
                        if env_provider == "llamacpp":
                            base_url_input.disable()
                            apikey_input.disable()

                        abbreviate_parties_switch = ui.switch("Abbreviate Parties (e.g. 'X and other')").classes("mt-2 text-xs")

                    # Upload card
                    with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                        ui.label("Upload Documents").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                        
                        async def handle_upload(e):
                            name = e.file.name
                            if any(r["filename"] == name for r in table_rows):
                                ui.notify(f"File already in queue: {name}", color="warning")
                                return
                            content_bytes = await e.file.read()
                            try:
                                if name.lower().endswith(".pdf"):
                                    import fitz
                                    doc = fitz.open(stream=content_bytes, filetype="pdf")
                                    text = ""
                                    for idx, page in enumerate(doc):
                                        text += f"\n===== Page {idx+1} =====\n" + page.get_text()
                                else:
                                    text = content_bytes.decode("utf-8", errors="ignore")
                                
                                queue[name] = text
                                queue_bytes[name] = content_bytes
                                table_rows.append({"filename": name, "case_number": "—", "date": "—", "status": "Pending", "actions": ""})
                                refresh_table()
                                
                                import asyncio
                                client = ui.context.client
                                upload_state["count"] += 1
                                if upload_state["timer"]:
                                    upload_state["timer"].cancel()
                                    
                                async def show_summary():
                                    await asyncio.sleep(0.5)
                                    with client:
                                        count = upload_state["count"]
                                        if count > 0:
                                            ui.notify(f"Added {count} file{'s' if count > 1 else ''} to the queue.", color="positive")
                                            upload_state["count"] = 0
                                    upload_state["timer"] = None
                                    
                                upload_state["timer"] = asyncio.create_task(show_summary())
                            except Exception as err:
                                ui.notify(f"Failed to read {name}: {err}", color="negative")
                                
                        ui.upload(multiple=True, label="Drag & Drop PDFs / Texts", auto_upload=True, on_upload=handle_upload).classes("w-full").props('accept=".pdf,.txt"')

                # Right queue + actions panel (65%)
                with ui.column().classes("w-[65%] gap-4"):
                    # Actions Card
                    with ui.row().classes("w-full items-center justify-between p-4 bg-white border").style("border-radius: 12px; border-color: #dadce0;"):
                        with ui.row().classes("gap-2"):
                            async def start_ingestion():
                                if not queue:
                                    ui.notify("Queue is empty. Upload some files first.", color="warning")
                                    return
                                run_files = []
                                for fname, txt in queue.items():
                                    row = next(r for r in table_rows if r["filename"] == fname)
                                    if row["status"] in ("Queued", "Pending", "Failed") or "Failed" in row["status"]:
                                        run_files.append((fname, txt))
                                if not run_files:
                                    ui.notify("No new files to process.", color="info")
                                    return
                                ingest_btn.props("disable")
                                clear_btn.props("disable")
                                export_btn.props("disable")
                                add_all_btn.props("disable")
                                for fname, _ in run_files:
                                    row = next(r for r in table_rows if r["filename"] == fname)
                                    row["status"] = "Queued"
                                refresh_table()
                                progress_container.visible = True
                                progress_label.set_text(f"Processing 0 of {len(run_files)} files...")
                                progress_bar.set_value(0.0)
                                
                                max_workers = 1 if provider_sel.value == "llamacpp" else int(workers_slider.value)
                                sem = asyncio.Semaphore(max_workers)
                                completed_tasks = 0
                                total_tasks = len(run_files)
                                success_count = 0
                                fail_count = 0
                                
                                async def process_file(fname, txt):
                                    nonlocal completed_tasks, success_count, fail_count
                                    row = next(r for r in table_rows if r["filename"] == fname)
                                    async with sem:
                                        row["status"] = "Extracting..."
                                        refresh_table()
                                        
                                        # Resolve fallback keys for API-driven providers
                                        active_key = apikey_input.value
                                        is_cur_admin = (app.storage.user.get("username") or "").lower() == "admin"
                                        if not active_key:
                                            if is_cur_admin:
                                                if provider_sel.value == "anthropic":
                                                    active_key = settings.anthropic_api_key
                                                elif provider_sel.value == "openai":
                                                    active_key = settings.openai_api_key
                                            else:
                                                raise RuntimeError("API Key is required for non-admin accounts.")

                                        try:
                                            meta = await run.io_bound(
                                                run_extraction,
                                                text=txt,
                                                provider=provider_sel.value,
                                                model_or_path=model_input.value,
                                                api_key=active_key or None,
                                                base_url=base_url_input.value or None
                                            )
                                            row["case_number"] = meta.case_number
                                            row["date"] = meta.date_of_judgment
                                            row["status"] = "Completed"
                                            
                                            results[:] = [r for r in results if r["filepath"] != fname]
                                            results.append({
                                                "filepath": fname,
                                                "metadata": meta.model_dump(mode="json")
                                            })
                                            success_count += 1
                                            return True
                                        except Exception as exc:
                                            row["status"] = f"Failed: {exc}"
                                            fail_count += 1
                                            return False
                                        finally:
                                            completed_tasks += 1
                                            progress_label.set_text(f"Processing {completed_tasks} of {total_tasks} files...")
                                            progress_bar.set_value(completed_tasks / total_tasks)
                                            refresh_table()
                                            
                                tasks = [process_file(fname, txt) for fname, txt in run_files]
                                await asyncio.gather(*tasks)
                                ui.notify(f"Completed: {success_count} succeeded, {fail_count} failed.", color="positive" if fail_count == 0 else "warning")
                                await asyncio.sleep(2.0)
                                progress_container.visible = False
                                ingest_btn.props(remove="disable")
                                clear_btn.props(remove="disable")
                                export_btn.props(remove="disable")
                                add_all_btn.props(remove="disable")

                            ingest_btn = ui.button("Extract Metadata", icon="play_arrow", on_click=start_ingestion).props("color=primary rounded")
                            
                            def clear_queue():
                                queue.clear()
                                queue_bytes.clear()
                                table_rows.clear()
                                results.clear()
                                refresh_table()
                                ui.notify("Queue cleared", color="grey-7")
                            clear_btn = ui.button("Clear Queue", icon="clear_all", on_click=clear_queue).props("flat rounded color=grey-7")

                        with ui.row().classes("gap-2"):
                            async def add_all_to_corpus():
                                completed_records = [r for r in results]
                                if not completed_records:
                                    ui.notify("No completed extractions to add. Run AI extraction first.", color="warning")
                                    return
                                add_all_btn.props("disable")
                                count = 0
                                for rec in completed_records:
                                    fname = rec["filepath"]
                                    row = next((r for r in table_rows if r["filename"] == fname), None)
                                    if row and row["status"] == "Completed":
                                        success = await do_add_to_corpus_impl(fname, rec["metadata"], row)
                                        if success:
                                            count += 1
                                ui.notify(f"Successfully added {count} cases to the active corpus!", color="green")
                                show_tree()
                                add_all_btn.props(remove="disable")

                            add_all_btn = ui.button("Add All to Corpus", icon="library_add", on_click=add_all_to_corpus).props("color=primary rounded").tooltip("Save and index all completed files into search database")

                            def download_excel():
                                if not results:
                                    ui.notify("No completed extractions to export.", color="warning")
                                    return
                                try:
                                    xlsx_data = build_excel_bytes(results)
                                    ui.download(xlsx_data, "metadata_registry.xlsx")
                                except Exception as err:
                                    ui.notify(f"Excel generation failed: {err}", color="negative")
                            export_btn = ui.button("Export Excel", icon="grid_on", on_click=download_excel).props("color=green rounded")

                    # Progress Container
                    with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;") as progress_container:
                        progress_container.visible = False
                        progress_label = ui.label("Ingesting documents...").classes("text-xs font-semibold text-gray-600 mb-2")
                        progress_bar = ui.linear_progress(value=0.0).props("stripe lstrip").classes("w-full")

                    # Table Card
                    with ui.card().classes("w-full p-4 overflow-auto").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none; flex-grow: 1; height: 100%;"):
                        ui.label("Extraction Queue").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                        columns = [
                            {"name": "filename", "label": "File Name", "field": "filename", "align": "left"},
                            {"name": "case_number", "label": "Case Number", "field": "case_number", "align": "center"},
                            {"name": "date", "label": "Date of Judgment", "field": "date", "align": "center"},
                            {"name": "status", "label": "Status", "field": "status", "align": "center"},
                            {"name": "actions", "label": "Actions", "field": "actions", "align": "center"}
                        ]
                        results_table = ui.table(columns=columns, rows=table_rows, row_key="filename").classes("w-full shadow-none border")
                        
                        results_table.add_slot('body-cell-actions', '''
                            <q-td :props="props">
                                <q-btn v-if="props.row.status === 'Completed'" flat round dense icon="cloud_upload" color="primary" @click="() => $parent.$emit('add_to_corpus', props.row.filename)">
                                    <q-tooltip>Save & Index Case into search corpus</q-tooltip>
                                </q-btn>
                                <q-icon v-else-if="props.row.status === 'Added'" name="check_circle" color="green" size="sm">
                                    <q-tooltip>Successfully added to corpus</q-tooltip>
                                </q-icon>
                                <span v-else>—</span>
                            </q-td>
                        ''')

                        # Core Single-file indexing implementation
                        async def do_add_to_corpus_impl(filename, meta, row):
                            case_no = meta.get("case_number") or filename.split(".")[0]
                            row["status"] = "Adding..."
                            refresh_table()
                            
                            def run_indexing():
                                from src import store
                                from src.parsing import parse_and_chunk
                                from src.store import fts_index_cases
                                
                                target_dir = REPO_ROOT / "data" / "sc_judgements"
                                target_dir.mkdir(parents=True, exist_ok=True)
                                target_path = target_dir / filename
                                target_path.write_bytes(queue_bytes[filename])
                                
                                chunks = parse_and_chunk(target_path, case_no)
                                store.add_chunks(chunks, extra_meta={"date": meta.get("date_of_judgment") or "", "filename": filename})
                                
                                con = store.init_db()
                                db_meta = {
                                    "filename": filename,
                                    "case_no": case_no,
                                    "local_path": str(target_path),
                                    "title": case_no.replace("_", " "),
                                    "court": "SUPREME COURT",
                                    "judges": meta.get("judges") or []
                                }
                                store.upsert_judgement(con, db_meta, len(chunks))
                                store.mark_indexed(con, filename, "judgment", len(chunks))
                                fts_index_cases([case_no])
                                con.close()
                                return case_no

                            try:
                                await asyncio.get_event_loop().run_in_executor(None, run_indexing)
                                row["status"] = "Added"
                                refresh_table()
                                return True
                            except Exception as ex:
                                row["status"] = "Failed Ingest"
                                refresh_table()
                                ui.notify(f"Ingestion failed for {filename}: {ex}", color="red")
                                return False

                        async def handle_table_corpus_add(msg):
                            filename = msg.args
                            row = next((r for r in table_rows if r["filename"] == filename), None)
                            record = next((r for r in results if r["filepath"] == filename), None)
                            if row and record:
                                success = await do_add_to_corpus_impl(filename, record["metadata"], row)
                                if success:
                                    ui.notify(f"Successfully added case {row['case_number']} to search corpus!", color="green")
                                    show_tree()

                        results_table.on('add_to_corpus', handle_table_corpus_add)

        dialog.open()

    # 1. Setup left collapsible drawer first so it is available to the header toggle
    left_drawer = ui.left_drawer(value=False).props("width=400 behavior=mobile").classes("p-0 border-r smart-drawer drawer-closed")
    
    # Hidden input for receiving Javascript-intercepted case loading clicks without full page reloads
    js_case_input = ui.input(on_change=lambda e: open_case(e.value) if e.value else None).classes("hidden").style("display: none;").props('id="js-case-input"')
    
    def handle_drawer_change(e):
        if e.value:
            left_drawer.classes(add="drawer-open").classes(remove="drawer-closed")
        else:
            left_drawer.classes(add="drawer-closed").classes(remove="drawer-open")
            
    left_drawer.on_value_change(handle_drawer_change)
    with left_drawer:
        drawer_container = ui.column().classes("w-full h-full p-2 overflow-auto gap-3")
        with drawer_container:
            library = ui.column().classes("pane w-full h-full overflow-auto p-4 bg-white border rounded shadow-sm gap-3")

    with ui.header().classes("items-center justify-between text-white shadow-none px-4").style("background: #16273f; border-bottom: 3px solid #82212b; height: 56px; "):
        # Left side containing Menu Hamburger button, Back button, Logo, Wordmark, Extract and AI Status
        with ui.row().classes("items-center gap-2 no-wrap"):
            menu_btn = ui.button(icon="menu", on_click=lambda: left_drawer.toggle()).props("flat round dense color=white").classes("hover:bg-white/10")
            menu_btn.set_visibility(True)
            back_btn = ui.button(icon="arrow_back", on_click=go_back).props("flat round dense color=white").classes("hover:bg-white/10")
            back_btn.set_visibility(False)
            with back_btn:
                ui.tooltip("Go Back")
            
            # Clickable emblem logo and wordmark to navigate home
            with ui.row().classes("items-center gap-2 no-wrap cursor-pointer").style("user-select: none;").on("click", lambda: go_home()):
                ui.image("/logo/logo_emblem_trans.png").classes("w-7 h-7").style("object-fit: contain;")
                ui.label("ROS").classes("text-2xl font-bold text-white").style("font-family: 'Lora', Georgia, serif; letter-spacing: 0.1em;")
            
            # Extract and AI status badge
            if not demo:
                ui.badge("A.I. ACTIVE", color="green") \
                    .classes("px-3 text-[11px] font-bold ml-3 hide-narrow flex items-center justify-center") \
                    .style("border-radius: 9999px !important; box-shadow: none; height: 30px; min-width: 105px; text-transform: uppercase; letter-spacing: 0.05em; background-color: #22c55e !important;") \
                    .props('id="ai-status-badge"')
            if not demo:
                is_admin = (app.storage.user.get("username") or "").lower() == "admin"
                is_ro = (app.storage.user.get("username") or "").lower() in _RO_USERS
                if is_admin or is_ro:
                    with ui.button("EXTRACT", icon="document_scanner", on_click=open_extractor_dialog) \
                            .props("unelevated dense") \
                            .classes("text-[11px] font-bold ml-2 hide-narrow") \
                            .style("border-radius: 9999px !important; height: 30px; min-width: 105px; border: 1.5px solid rgba(255,255,255,0.3) !important; background-color: rgba(255,255,255,0.12) !important; color: white !important; text-transform: uppercase; letter-spacing: 0.05em;"):
                        ui.tooltip("Extract, chunk, and index a new PDF judgment instantly")
                else:
                    with ui.button("EXTRACT", icon="document_scanner") \
                            .props("disable unelevated dense") \
                            .classes("text-[11px] font-bold ml-2 hide-narrow") \
                            .style("border-radius: 9999px !important; height: 30px; min-width: 105px; border: 1.5px solid rgba(255,255,255,0.3) !important; background-color: rgba(255,255,255,0.12) !important; color: white !important; text-transform: uppercase; letter-spacing: 0.05em;"):
                        ui.tooltip("Extraction portal is restricted to Administrator / RO")
            if demo:
                ui.badge("DEMO · read-only", color="orange").classes("px-3 py-1 text-xs font-semibold ml-3 hide-narrow").style("border-radius: 2px; box-shadow: none;")
        # Centered Full-size Central Search Bar with FTS case suggestions
        with ui.row().classes("items-center no-wrap justify-center").style("position: absolute; left: 50%; transform: translateX(-50%); height: 100%; pointer-events: auto;"):
            header_search = ui.input(placeholder="Search cases, statutes, or ask ROS...") \
                .props("outlined rounded dense dark label-color=white text-color=white debounce=200") \
                .classes("w-[380px] text-xs hide-narrow") \
                .style("background-color: rgba(255,255,255,0.08); border-radius: 20px; font-size: 11px;")
            
            with header_search:
                with header_search.add_slot('prepend'):
                    ui.icon("search", color="white").classes("text-white")
                with header_search.add_slot('append'):
                    ui.icon("auto_awesome", color="white").classes("text-white cursor-pointer hover:opacity-80").on("click", lambda: librarian_dialog(header_search.value, force_local=True)).props('title="Ask ROS (A.I. Assistant)"')
                
                suggestions_menu = ui.menu().props("fit no-parent-event anchor='bottom left' self='top left'")

            def handle_search_change(e):
                val = ""
                if hasattr(e, 'args') and e.args:
                    val = str(e.args[0]).strip()
                else:
                    val = (e.value or "").strip()
                
                if len(val) < 2:
                    suggestions_menu.close()
                    return
                # Record portal search hit
                record_funnel_event("portal_hits")
                try:
                    con = _con()
                    matches = con.execute(
                        "SELECT case_no, parties, report_cite FROM judgements WHERE case_no LIKE ? OR parties LIKE ? OR report_cite LIKE ? LIMIT 5",
                        (f"%{val}%", f"%{val}%", f"%{val}%")
                    ).fetchall()
                    con.close()
                except Exception:
                    matches = []
                
                suggestions_menu.clear()
                with suggestions_menu:
                    # Ask ROS built-in action inside search dropdown - forces local model
                    def ask_ros_hdr(query=val):
                        header_search.value = ""
                        suggestions_menu.close()
                        librarian_dialog(query, force_local=True)
                    ui.menu_item(f'✨ Ask ROS about "{val}"', on_click=ask_ros_hdr).classes("text-xs py-1.5 px-3 font-semibold text-primary")
                    for case_no, parties, cite in matches:
                        display_name = parties if parties else case_no
                        if cite:
                            display_name = f"{display_name} ({cite})"
                        if len(display_name) > 60:
                            display_name = display_name[:57] + "..."
                            
                        def select_suggestion(c_no=case_no):
                            header_search.value = ""
                            suggestions_menu.close()
                            state["case"] = c_no
                            state["workspace_active"] = True
                            update_workspace_visibility()
                            show_tree()
                            q.value = ""
                            
                        ui.menu_item(display_name, on_click=select_suggestion).classes("text-xs py-1.5 px-3 max-w-[380px] overflow-hidden text-ellipsis")
                
                suggestions_menu.open()

            def perform_header_search():
                val = (header_search.value or "").strip()
                if not val:
                    return
                q.value = val
                apply_filters(deep=True)
                header_search.value = ""
                suggestions_menu.close()
                if left_drawer:
                    left_drawer.show()
                
            header_search.on("keydown.enter", perform_header_search)
            header_search.on("update:model-value", handle_search_change)

            def mobile_search_dialog():
                with ui.dialog() as dialog, ui.card().classes("w-full max-w-[400px] p-4"):
                    ui.label("Search Cases & Statutes").classes("text-sm font-bold text-primary mb-2")
                    search_input = ui.input(placeholder="Search cases, statutes, or ask ROS...").props("outlined dense autofocus").classes("w-full text-xs")
                    
                    # Suggestions container
                    suggestions_container = ui.column().classes("w-full gap-1 mt-2")
                    
                    def handle_mobile_change(e):
                        val = search_input.value.strip()
                        suggestions_container.clear()
                        if len(val) < 2:
                            return
                        try:
                            con = _con()
                            matches = con.execute(
                                "SELECT case_no, parties, citation FROM judgements WHERE case_no LIKE ? OR parties LIKE ? OR citation LIKE ? LIMIT 5",
                                (f"%{val}%", f"%{val}%", f"%{val}%")
                            ).fetchall()
                            con.close()
                            
                            with suggestions_container:
                                def ask_ros_cb(q=val):
                                    dialog.close()
                                    librarian_dialog(q, force_local=True)
                                ui.button(f'✨ Ask ROS about "{val}"', on_click=ask_ros_cb).props("flat align=left").classes("w-full text-xs text-primary font-semibold text-left normal-case")
                                
                                for case_no, parties, cite in matches:
                                    display_name = parties if parties else case_no
                                    if cite:
                                        display_name = f"{display_name} ({cite})"
                                    if len(display_name) > 60:
                                        display_name = display_name[:57] + "..."
                                        
                                    def make_select_cb(c_no=case_no):
                                        dialog.close()
                                        state["case"] = c_no
                                        state["workspace_active"] = True
                                        update_workspace_visibility()
                                        show_tree()
                                        q.value = ""
                                    ui.button(display_name, on_click=make_select_cb).props("flat align=left").classes("w-full text-xs text-left overflow-hidden text-ellipsis normal-case")
                        except Exception:
                            pass
                            
                    def perform_mobile_search():
                        val = search_input.value.strip()
                        if not val:
                            return
                        q.value = val
                        apply_filters(deep=True)
                        dialog.close()
                        if left_drawer:
                            left_drawer.show()
                            
                    search_input.on("update:model-value", handle_mobile_change)
                    search_input.on("keydown.enter", perform_mobile_search)
                    
                    with ui.row().classes("w-full justify-end mt-4"):
                        ui.button("Close", on_click=dialog.close).props("flat").classes("text-xs")
                        ui.button("Search", on_click=perform_mobile_search).props("unelevated").classes("text-xs text-white bg-primary")
                        
                dialog.open()

        # Right side containing admin console, settings, and logout
        with ui.row().classes("items-center gap-1 no-wrap").style("height: 100%;"):
            is_admin = (app.storage.user.get("username") or "").lower() == "admin"
            if is_admin:
                admin_btn = ui.button(icon="admin_panel_settings", on_click=admin_dialog).props("flat round dense color=white").classes("hover:bg-white/10 w-9 h-9")
                with admin_btn:
                    ui.tooltip("Admin Console")
            
            # Mobile search button
            ui.button(icon="search", on_click=mobile_search_dialog).props("flat round dense color=white").classes("show-narrow hover:bg-white/10 w-9 h-9")
            settings_btn = ui.button(icon="settings", on_click=settings_dialog).props("flat round dense color=white").classes("hover:bg-white/10 w-9 h-9")
            with settings_btn:
                ui.tooltip("Settings")
                
            logout_btn = ui.button(icon="logout", on_click=lambda: (app.storage.user.clear(), ui.navigate.to("/login"))).props("flat round dense color=white").classes("hover:bg-white/10 w-9 h-9")
            with logout_btn:
                ui.tooltip("Logout")
    
    # mobile-only tab switcher removed (PDF download button is now inline in breakdown card)
    tab_btns = {}

    ui.add_head_html('<style>.hidden-separator > .q-splitter__separator { display: none !important; }</style>')
    ui.add_head_html('''
    <script>
    document.addEventListener('click', function(e) {
        let anchor = e.target.closest('a');
        if (anchor && anchor.getAttribute('href')) {
            let href = anchor.getAttribute('href');
            if (href.startsWith('open://') || href.includes('case=')) {
                e.preventDefault();
                let caseNo = '';
                if (href.startsWith('open://')) {
                    caseNo = href.replace('open://', '');
                } else {
                    let search = href.split('?')[1] || '';
                    let params = new URLSearchParams(search);
                    caseNo = params.get('case');
                }
                if (caseNo) {
                    let input = document.getElementById('js-case-input');
                    if (input) {
                        input.value = caseNo;
                        input.dispatchEvent(new Event('input'));
                    }
                }
            }
        }
    });
    </script>
    ''')
    ui.add_head_html('''
    <style>
        /* Uniform Rounding System */
        .pane {
            border-radius: 16px !important;
        }
        .q-card {
            border-radius: 16px !important;
        }
        .q-dialog .q-card {
            border-radius: 24px !important;
        }
        .q-expansion-item {
            border-radius: 12px !important;
            overflow: hidden;
        }
        .q-expansion-item__container {
            border-radius: 12px !important;
        }
        
        /* Modern Rounded Chat Bubbles */
        .q-message-text {
            border-radius: 18px !important;
            padding: 10px 14px !important;
            font-size: 12px !important;
            line-height: 1.4 !important;
        }
        .q-message-text-received {
            border-top-left-radius: 2px !important;
            background-color: #fffbeb !important; /* light gold/amber tint */
            border: 1px solid #fde68a !important;
            color: #1e293b !important;
        }
        .q-message-text-sent {
            border-top-right-radius: 2px !important;
            background-color: #82212b !important; /* deep red */
            color: #ffffff !important;
            border: 1px solid #82212b !important;
        }
        
        /* Input & Button Capsular Roundness */
        .q-field--outlined .q-field__control {
            border-radius: 10px !important;
        }
        .q-field--rounded .q-field__control {
            border-radius: 24px !important;
        }
        .q-btn {
            border-radius: 8px !important;
        }
        .q-btn--round {
            border-radius: 50% !important;
        }
        .q-btn--outline {
            border-radius: 10px !important;
        }
        .q-chip {
            border-radius: 16px !important;
        }
        
        /* Scrollbar roundness */
        ::-webkit-scrollbar {
            width: 8px;
            height: 8px;
        }
        ::-webkit-scrollbar-track {
            background: transparent;
        }
        ::-webkit-scrollbar-thumb {
            background: #cbd5e1;
            border-radius: 4px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background: #94a3b8;
        }
        
        /* Mosaic Ambient Background & Plain Design */
        body {
            transition: background 0.3s ease;
        }
        body:not(.workspace-mode) {
            background: radial-gradient(circle at 10% 10%, rgba(130, 33, 43, 0.04) 0%, transparent 60%),
                        radial-gradient(circle at 90% 90%, rgba(30, 41, 59, 0.04) 0%, transparent 60%),
                        linear-gradient(135deg, #f8fafc 0%, #f1f5f9 50%, #e2e8f0 100%) !important;
        }
        body:not(.workspace-mode) .panes-row {
            background-color: transparent !important;
            height: auto !important;
            min-height: calc(100vh - 64px - 30px) !important;
            align-items: flex-start !important;
            /* Home hub keeps NiceGUI's default header padding, so only a small
               top margin is needed here (the 76px is workspace-only). */
            margin: 12px 16px 16px 16px !important;
            width: calc(100% - 32px) !important;
        }
        body.workspace-mode {
            background-color: #f1f5f9 !important;
        }
        
        /* Auto-scrolling Infinite Marquee CSS */
        @keyframes marquee-left {
            0% { transform: translateX(0); }
            100% { transform: translateX(-50%); }
        }
        @keyframes marquee-right {
            0% { transform: translateX(-50%); }
            100% { transform: translateX(0); }
        }

        .marquee-container {
            overflow: hidden;
            width: 100%;
            display: flex;
            position: relative;
            padding: 8px 0;
            mask-image: linear-gradient(to right, transparent, white 15%, white 85%, transparent);
            -webkit-mask-image: linear-gradient(to right, transparent, white 15%, white 85%, transparent);
        }

        .marquee-content-left {
            display: flex;
            gap: 16px;
            animation: marquee-left 35s linear infinite;
            width: max-content;
        }
        .marquee-content-right {
            display: flex;
            gap: 16px;
            animation: marquee-right 145s linear infinite;
            width: max-content;
        }

        .marquee-container:hover .marquee-content-left,
        .marquee-container:hover .marquee-content-right {
            animation-play-state: paused;
        }

        /* Make the attach upload input button compact and perfectly center-middle align its text label */
        .attach-upload {
            height: 36px !important;
            min-height: 36px !important;
            overflow: hidden;
            border-radius: 24px !important;
        }
        .attach-upload .q-uploader__header {
            height: 100% !important;
            padding: 0 !important;
            background-color: transparent !important;
        }
        .attach-upload .q-uploader__header-content {
            height: 100% !important;
            padding: 0 !important;
            display: flex !important;
            align-items: center !important;
            justify-content: center !important;
            text-align: center !important;
        }
        .attach-upload .q-uploader__title {
            font-size: 11px !important;
            font-weight: 700 !important;
            line-height: 1 !important;
            margin: 0 !important;
            text-align: center !important;
            width: 100% !important;
            color: var(--q-primary) !important;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }
        .attach-upload .q-uploader__list {
            display: none !important;
        }
        
        @media (min-width: 901px) {
            .show-narrow {
                display: none !important;
            }
        }

        @media (max-width: 900px) {
            .show-narrow {
                display: inline-flex !important;
            }
            .mobile-hide-infographics {
                display: none !important;
            }
        }
    </style>
    ''')
    panes_row = ui.row().classes("panes-row w-full no-wrap gap-0").style("padding: 0px !important; gap: 24px !important; background: transparent !important;")
    with panes_row:
        welcome_container = ui.column().classes("w-full items-center justify-start p-0")
        splitter_main = ui.splitter(value=50).classes("w-full h-full").style("background: transparent;")
        with splitter_main.before:
            pdf_pane = ui.column().classes("pane w-full h-full overflow-hidden p-4 bg-white border rounded shadow-sm").props('id="pdf-pane"')
        with splitter_main.after:
            breakdown_pane = ui.column().classes("pane w-full h-full overflow-auto p-4 bg-white border rounded shadow-sm").props('id="breakdown-pane"')
        with library:
            # Welcome infographics section (visible only in welcome state)
            welcome_infographics_container = ui.column().classes("w-full items-center gap-4 mb-4 p-2")
            with welcome_infographics_container:
                # Title with transparent emblem only (no text next to it)
                with ui.row().classes("items-center justify-center mb-2"):
                    ui.image("/logo/logo_emblem_trans.png").classes("w-20 h-20").style("object-fit: contain;")
                
                ui.label("Case Index Platform - Fast, Accurate, Secure..").classes("text-[13px] text-gray-600 mb-4 uppercase tracking-widest font-bold text-center w-full")
                
                # Fetch statistics
                total_cases, total_precedents, total_legislation = get_db_stats()
                
                # Static Infographics Row restored back to the main card
                with ui.row().classes("w-full justify-center gap-3 mt-2 flex-wrap mobile-hide-infographics"):
                    metrics_config = [
                        ("Cases & Reports", f"{total_cases:,}"),
                        ("Acts & Statutes", f"{total_legislation:,}"),
                        ("Precedent Links", f"{total_precedents:,}"),
                        ("NLR Cases", "8,866"),
                        ("SLR Cases", "3,580"),
                        ("Case Digests", "1,055"),
                        ("SC Minutes", "3,374")
                    ]
                    for label, val in metrics_config:
                        with ui.card().classes("p-3 items-center justify-center border hover:shadow-md transition-all duration-200").style("width: 130px; height: 85px; border-radius: 12px; border-color: #cbd5e1; background-color: #ffffff; box-shadow: none;"):
                            ui.label(val).classes("text-lg font-bold text-primary").style("line-height: 1;")
                            ui.label(label).classes("text-[9px] uppercase font-bold text-gray-500 text-center tracking-wider mt-1")



                # In addition, render the diagonal background scroll loops behind the card!
                bg_scroll_html = _get_diagonal_scroll_bg_html()
                ui.html(bg_scroll_html + f'''
                <style>
                .login-bg-pattern {{
                    position: fixed;
                    top: 0;
                    left: 0;
                    width: 100vw;
                    height: 100vh;
                    overflow: hidden;
                    background-color: #f1f5f9;
                    z-index: -1;
                    display: grid;
                    grid-template-columns: repeat(14, 1fr);
                    gap: 12px;
                    pointer-events: none;
                    padding: 5px;
                }}
                .diagonal-strip {{
                    display: flex;
                    flex-direction: column;
                    transform: rotate(-15deg) translateY(-20%);
                    animation: scrollDiagonal 45s linear infinite;
                }}
                .diagonal-strip.reverse {{
                    animation: scrollDiagonalReverse 45s linear infinite;
                }}
                @keyframes scrollDiagonal {{
                    0% {{ transform: rotate(-15deg) translateY(-50%); }}
                    100% {{ transform: rotate(-15deg) translateY(0%); }}
                }}
                @keyframes scrollDiagonalReverse {{
                    0% {{ transform: rotate(-15deg) translateY(0%); }}
                    100% {{ transform: rotate(-15deg) translateY(-50%); }}
                }}
                .a4-sheet {{
                    background: #ffffff;
                    border: 1px solid #cbd5e1;
                    box-shadow: 0 4px 12px rgba(0,0,0,0.02), 0 1px 2px rgba(0,0,0,0.01);
                    border-radius: 4px;
                    display: flex;
                    flex-direction: column;
                    gap: 6px;
                    opacity: 0.55;
                }}
                .a4-header {{
                    font-weight: 700;
                    text-align: center;
                    color: #475569;
                    text-transform: uppercase;
                    letter-spacing: 0.05em;
                    border-bottom: 1.5px solid #cbd5e1;
                }}
                </style>
                ''').classes("mobile-hide-infographics")
            
            with ui.row().classes("w-full items-center justify-between no-wrap"):
                drawer_close_btn = ui.button(icon="close", on_click=lambda: left_drawer.hide()).props("flat round dense color=grey-7").classes("hover:bg-gray-100")
                drawer_close_btn.set_visibility(False)
                library_title = ui.label("Library").classes("pane-head")
                home_btn = ui.button(icon="home", on_click=lambda: go_home()).props("flat round dense color=grey-7").classes("hover:bg-gray-100")
            
            def account_dialog():
                from src.analyze_smart import LLM_METRICS
                username = app.storage.user.get("username") or "Guest"
                is_admin = username.lower() == "admin"
                acct_type = "Administrator" if is_admin else "Standard User"
                
                with ui.dialog() as dialog, ui.card().style("width: 380px; max-width: 90vw; border-radius: 16px; padding: 0; overflow: hidden; border: 1px solid #e2e8f0;"):
                    # Header: Slate Navy
                    with ui.row().classes("w-full bg-[#0f172a] text-white p-4 items-center justify-between no-wrap").style("border-bottom: 2px solid #b8965a;"):
                        with ui.row().classes("items-center gap-2"):
                            ui.icon("account_circle", color="amber").classes("text-xl")
                            ui.label("Account Info").classes("text-xs font-bold tracking-wider text-white uppercase")
                        ui.button(icon="close", on_click=dialog.close).props("flat round dense color=white")
                    
                    # Content panel
                    with ui.column().classes("w-full p-4 gap-2 bg-slate-50"):
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Username:").classes("text-[11px] font-semibold text-gray-500")
                            ui.label(username).classes("text-[11px] font-bold text-slate-800")
                            
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Account Type:").classes("text-[11px] font-semibold text-gray-500")
                            ui.label(acct_type).classes("text-[11px] font-bold text-slate-800")
                            
                        user_spec_metrics = get_user_metrics(username)
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Total A.I. Queries:").classes("text-[11px] font-semibold text-gray-500")
                            ui.label(str(user_spec_metrics['query_count'])).classes("text-[11px] font-bold text-slate-800")
                            
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Tokens Processed:").classes("text-[11px] font-semibold text-gray-500")
                            ui.label(f"{user_spec_metrics['total_tokens']:,}").classes("text-[11px] font-bold text-slate-800")
                    
                    # Footer Close
                    with ui.row().classes("w-full justify-end p-3 bg-white border-t"):
                        ui.button("Close", on_click=dialog.close).props("outline color=primary").classes("px-4 text-xs font-semibold rounded-lg")
                dialog.open()

            def contact_dialog():
                with ui.dialog() as dialog, ui.card().style("width: 380px; max-width: 90vw; border-radius: 16px; padding: 0; overflow: hidden; border: 1px solid #e2e8f0;"):
                    # Header: Slate Navy
                    with ui.row().classes("w-full bg-[#0f172a] text-white p-4 items-center justify-between no-wrap").style("border-bottom: 2px solid #b8965a;"):
                        with ui.row().classes("items-center gap-2"):
                            ui.icon("contact_mail", color="amber").classes("text-xl")
                            ui.label("Contact Info").classes("text-xs font-bold tracking-wider text-white uppercase")
                        ui.button(icon="close", on_click=dialog.close).props("flat round dense color=white")
                    
                    # Content panel
                    with ui.column().classes("w-full p-4 gap-2 bg-slate-50"):
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Email:").classes("text-[11px] font-semibold text-gray-500")
                            ui.link("smslakshman@gmail.com", "mailto:smslakshman@gmail.com").classes("text-[11px] font-bold text-primary hover:underline")
                            
                        with ui.row().classes("w-full justify-between items-center border-b pb-2 border-slate-200"):
                            ui.label("Phone:").classes("text-[11px] font-semibold text-gray-500")
                            ui.link("0767292161", "tel:0767292161").classes("text-[11px] font-bold text-primary hover:underline")
                    # Footer Close
                    with ui.row().classes("w-full justify-end p-3 bg-white border-t"):
                        ui.button("Close", on_click=dialog.close).props("outline color=primary").classes("px-4 text-xs font-semibold rounded-lg")
                dialog.open()

            # Placeholders for later: Payment, Contact Us, Account (visible only in landing screen)
            with ui.row().classes("w-full justify-center gap-4 mb-2 flex-col sm:flex-row items-center").style("margin-top: -8px;") as home_buttons_row:
                # PAYMENT: Solid Gold/Yellow background with Bold Crimson text and Crimson border outline
                ui.button("Payment", on_click=lambda: ui.notify("Payment coming soon..."))\
                    .classes("text-xs font-bold py-1.5 rounded-lg w-full max-w-[200px] sm:w-[160px]")\
                    .style("background-color: #ffd54f !important; color: #82212b !important; border: 2px solid #82212b;")
                
                # CONTACT US: Solid Gold/Yellow background with Bold Crimson text and Crimson border outline
                ui.button("Contact Us", on_click=contact_dialog)\
                    .classes("text-xs font-bold py-1.5 rounded-lg w-full max-w-[200px] sm:w-[160px]")\
                    .style("background-color: #ffd54f !important; color: #82212b !important; border: 2px solid #82212b;")
                
                # ACCOUNT: Solid Gold/Yellow background with Bold Crimson text and Crimson border outline
                ui.button("Account", on_click=account_dialog)\
                    .classes("text-xs font-bold py-1.5 rounded-lg w-full max-w-[200px] sm:w-[160px]")\
                    .style("background-color: #ffd54f !important; color: #82212b !important; border: 2px solid #82212b;")
            
            # Library indicator label
            home_library_signifier = ui.label("LIBRARY BROWSER").classes("w-full text-center text-[10px] text-gray-400 font-bold uppercase tracking-widest mt-1 mb-2")
            home_library_signifier.set_visibility(False)
            home_buttons_row.set_visibility(False)

            containers["bookmarks"] = ui.column().classes("w-full mb-1")
            
            # Google Search Input styled as an invisible mock element so filter functions can read/write its value
            q = ui.input().style("display: none;")
            

            
            judge_sel = ui.select(distinct_justices(), label="By Justice (select one or more)", with_input=True,
                                  multiple=True, clearable=True,
                                  on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("w-full")
            area_sel = ui.select(legal_areas(), label="By legal area", with_input=True, clearable=True,
                                 on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("w-full")
            with ui.row().classes("w-full no-wrap gap-2"):
                year_sel = ui.select(_available_years(), label="Year", with_input=True, clearable=True,
                                     on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("flex-1 min-w-0")
                month_sel = ui.select(_MONTH_NAMES, label="Month", clearable=True,
                                      on_change=lambda e: apply_filters()).props("outlined dense rounded").classes("flex-1 min-w-0")
            active_filters = ui.row().classes("w-full items-center gap-1 mt-1").style("flex-wrap: wrap;")
            results = ui.column().classes("w-full")

    with ui.footer().classes("items-center justify-center bg-white border-t p-1 shadow-none").style("height: 30px; border-color: #dadce0;"):
        ui.label("built by SMS").classes("text-[10px] text-gray-500 font-semibold uppercase tracking-wider")

    def go_home():
        state["history_stack"] = []
        if back_btn:
            back_btn.set_visibility(False)
        q.value = ""
        state["case"] = None
        state["statute"] = None
        state["workspace_active"] = False
        update_workspace_visibility()
        show_tree()
        floating_chat_widget.refresh()

    def update_workspace_visibility():
        active = state.get("workspace_active", False)
        
        pdf_pane.set_visibility(active)
        breakdown_pane.set_visibility(active)
        welcome_infographics_container.set_visibility(not active)
        splitter_main.set_visibility(active)
        welcome_container.set_visibility(not active)
        if menu_btn:
            menu_btn.set_visibility(active)
        
        if active:
            try:
                ui.run_javascript("document.body.classList.add('workspace-mode');")
            except Exception:
                pass
            # Move library to left drawer container
            if drawer_container and library:
                library.move(target_container=drawer_container)
            
            # Prevent double cards: strip library container styles inside the sidebar
            library.classes(remove="max-w-[1100px] mx-auto mt-3 p-6 shadow-md border rounded bg-white shadow-sm mt-0")
            library.style("background-color: transparent !important; border: none !important; box-shadow: none !important;")
            
            library_title.classes(remove="text-2xl text-center mb-4")
            library_title.classes(add="pane-head")
            library_title.set_visibility(True)
            home_buttons_row.set_visibility(False)
            home_library_signifier.set_visibility(False)
            home_btn.set_visibility(True)
            if drawer_close_btn:
                drawer_close_btn.set_visibility(True)
            if left_drawer:
                left_drawer.show()
        else:
            try:
                ui.run_javascript("document.body.classList.remove('workspace-mode');")
            except Exception:
                pass
            # Move library to welcome container on landing screen
            if welcome_container and library:
                library.move(target_container=welcome_container)
                
            # Home page library container layout & standard white card styling with translucent red outline
            library.classes(remove="pane-head")
            library.classes(add="w-full max-w-[1100px] mx-auto mt-0 p-6 shadow-md border rounded")
            library.style("background-color: #ffffff !important; border: 2px solid rgba(130, 33, 43, 0.35) !important; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05), 0 2px 4px -1px rgba(0,0,0,0.03) !important;")
            
            library_title.set_visibility(False)
            home_buttons_row.set_visibility(True)
            home_library_signifier.set_visibility(True)
            home_btn.set_visibility(False)
            if drawer_close_btn:
                drawer_close_btn.set_visibility(False)
            if left_drawer:
                left_drawer.hide()

    def set_active(name):
        if name == "library":
            if left_drawer:
                left_drawer.show()
            return
            
        if state.get("workspace_active", False):
            if left_drawer:
                left_drawer.hide()
                
        for k, pane in {"pdf": pdf_pane, "bd": breakdown_pane, "library": library}.items():
            pane.classes(add="active") if k == name else pane.classes(remove="active")
        for k, btn in tab_btns.items():
            if k in tab_btns:
                btn.props(f"color={'primary' if k == name else 'grey-6'}")
                btn.style("border-bottom-color: #82212b;" if k == name else "border-bottom-color: transparent;")
            
        if name == "pdf":
            splitter_main.classes(add="show-before").classes(remove="show-after")
        elif name == "bd":
            splitter_main.classes(add="show-after").classes(remove="show-before")
        else:
            splitter_main.classes(remove="show-before show-after")

    show_tree()
    refresh_bookmarks_ui()
    if initial_case:
        open_case(initial_case)
    render_pdf()
    render_breakdown()
    floating_chat_widget()
    update_workspace_visibility()
    set_active("bd")  # default visible pane on mobile (desktop shows all 3)
    ui.timer(2.0, _poll_breakdown)  # render a finished breakdown even if the client dropped mid-wait


@ui.page("/", title="ROS", favicon="data/logos/logo_emblem_trans.png")
def home(case: str = None):
    if not app.storage.user.get("authenticated", False):
        app.storage.user["referrer_path"] = f"/?case={case}" if case else "/"
        return RedirectResponse("/login")
    build_workspace(demo=False, initial_case=case)


@ui.page("/demo", title="ROS", favicon="data/logos/logo_emblem_trans.png")
def demo_page(case: str = None):
    app.storage.user["demo_visitor"] = True
    build_workspace(demo=True, initial_case=case)


@ui.page("/extractor", title="Ingestion Portal", favicon="data/logos/logo_emblem_trans.png")
def extractor_page():
    from metadata_extractor.extractor import run_extraction
    from metadata_extractor.models import JudgmentMetadata
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import logging
    
    logger = logging.getLogger("metadata_extractor_ui")

    # Check authorization first
    if not app.storage.user.get("authenticated", False):
        return RedirectResponse("/login")

    # Page state
    queue = {}
    table_rows = []
    results = []

    upload_state = {"count": 0, "timer": None}

    def refresh_table():
        results_table.rows = list(table_rows)
        results_table.update()

    def build_excel_bytes(extracted_records: list[dict]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Registry Overview"
        ws.views.sheetView[0].showGridLines = True
        
        headers = [
            "Case Number", 
            "Date of Judgment", 
            "Appellants / Petitioners", 
            "Respondents", 
            "Judges", 
            "Legislation Cited", 
            "Keywords"
        ]
        ws.append(headers)
        
        font_family = "Segoe UI"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        
        data_font = Font(name=font_family, size=10)
        thin_border = Border(
            left=Side(style='thin', color='DADCE0'),
            right=Side(style='thin', color='DADCE0'),
            top=Side(style='thin', color='DADCE0'),
            bottom=Side(style='thin', color='DADCE0')
        )
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")
        
        fill_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for row_idx, record in enumerate(extracted_records, 2):
            meta = record.get("metadata", {})
            parties = meta.get("parties", {})
            
            appellants = ", ".join(parties.get("appellants_petitioners", [])) if isinstance(parties.get("appellants_petitioners"), list) else ""
            respondents = ", ".join(parties.get("respondents", [])) if isinstance(parties.get("respondents"), list) else ""
            judges = ", ".join(meta.get("judges", [])) if isinstance(meta.get("judges"), list) else ""
            leg_list = meta.get("legislation_cited", [])
            if isinstance(leg_list, list):
                leg_strs = []
                for item in leg_list:
                    if isinstance(item, dict):
                        sec_num = item.get("section") or ""
                        statute = item.get("statute") or ""
                        leg_strs.append(f"{sec_num} of {statute}" if sec_num else statute)
                    else:
                        leg_strs.append(str(item))
                legislation = ", ".join(leg_strs)
            else:
                legislation = ""
            keywords = ", ".join(meta.get("keywords", [])) if isinstance(meta.get("keywords"), list) else ""
            
            row_data = [
                meta.get("case_number", ""),
                meta.get("date_of_judgment", ""),
                appellants,
                respondents,
                judges,
                legislation,
                keywords
            ]
            ws.append(row_data)
            ws.row_dimensions[row_idx].height = 24
            
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if is_even:
                    cell.fill = fill_even
                
                if col_idx in (1, 2):
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        return bytes_io.getvalue()

    # Dynamic styling helpers
    ui.add_head_html("""
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;500;600;700&family=Lora:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap" rel="stylesheet">
    <style>
      body {
        background: #f8f9fa;
        font-family: 'Plus Jakarta Sans', sans-serif;
      }
    </style>
    """)

    # Header Bar
    with ui.header().classes("items-center justify-between bg-white text-gray-900 border-b shadow-none").style("border-color: #dadce0; height: 56px;"):
        with ui.row().classes("items-center gap-3"):
            ui.button(icon="arrow_back", on_click=lambda: ui.navigate.to("/")).props("flat round dense color=grey-7").classes("hover:bg-gray-100")
            ui.label("Batch Ingestion Portal").classes("text-sm font-bold text-gray-800")
        with ui.row().classes("items-center gap-2 pr-2"):
            ui.image("/logo/logo_emblem_trans.png").classes("w-6 h-6").style("object-fit: contain;")
            ui.label("ROS").classes("text-lg font-bold").style("font-family: 'Lora', Georgia, serif; color: #82212b;")

    # Main layout grid (Left: Settings + Upload, Right: Progress & Table)
    with ui.row().classes("w-full no-wrap gap-4 p-4 items-start"):
        
        # Left Panel (Settings + Upload)
        with ui.column().classes("w-1/3 gap-4"):
            
            # Settings Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Settings").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                # Default provider selection based on workspace settings
                env_provider = "llamacpp"
                provider_sel = ui.select(
                    {"llamacpp": "Local GGUF (Llama.cpp)", "openai": "OpenAI / Groq API", "anthropic": "Anthropic Claude API"},
                    value=env_provider,
                    label="LLM Provider"
                ).classes("w-full")
                
                # Model or GGUF Path
                env_model = settings.llamacpp_model_path if env_provider == "llamacpp" else settings.llm_model
                model_input = ui.input("Model or GGUF Path", value=env_model).classes("w-full")
                
                # Custom Base URL (Groq/Ollama/OpenRouter)
                env_base_url = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                base_url_input = ui.input("Custom API Base URL (Optional)", value=env_base_url, placeholder="e.g. https://api.groq.com/openai/v1").classes("w-full")
                
                # API Key (hidden by default)
                apikey_input = ui.input("API Key (Optional)", password=True).classes("w-full")
                
                # Concurrency slider
                ui.label("Parallel Processing Threads").classes("text-[10px] text-gray-500 font-semibold mt-2")
                workers_slider = ui.slider(min=1, max=10, value=3).props("label label-always")

                # Auto-change settings helper
                def on_provider_change(e):
                    if e.value == "llamacpp":
                        model_input.value = settings.llamacpp_model_path
                        base_url_input.value = ""
                        apikey_input.value = ""
                        base_url_input.disable()
                        apikey_input.disable()
                    elif e.value == "openai":
                        model_input.value = settings.llm_model if "openai" in settings.llm_provider.lower() else "gpt-4o-mini"
                        base_url_input.value = settings.openai_base_url if "openai" in settings.llm_provider.lower() else ""
                        base_url_input.enable()
                        apikey_input.enable()
                    else:  # anthropic
                        model_input.value = settings.anthropic_model or "claude-sonnet-5"
                        base_url_input.value = ""
                        base_url_input.disable()
                        apikey_input.enable()
                        
                provider_sel.on_value_change(on_provider_change)
                # Run once initially
                if env_provider == "llamacpp":
                    base_url_input.disable()
                    apikey_input.disable()

            # Upload Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Upload Documents").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                async def handle_upload(e):
                    name = e.file.name
                    # Prevent duplicate file names in the active queue
                    if any(r["filename"] == name for r in table_rows):
                        ui.notify(f"File already in queue: {name}", color="warning")
                        return
                        
                    content_bytes = await e.file.read()
                    try:
                        if name.lower().endswith(".pdf"):
                            import fitz
                            doc = fitz.open(stream=content_bytes, filetype="pdf")
                            text = ""
                            for i, page in enumerate(doc):
                                text += f"\n===== Page {i+1} =====\n" + page.get_text()
                        else:
                            text = content_bytes.decode("utf-8", errors="ignore")
                        
                        queue[name] = text
                        table_rows.append({"filename": name, "case_number": "—", "date": "—", "status": "Pending"})
                        refresh_table()
                        logger.info(f"File uploaded successfully: {name}")
                        print(f"File uploaded successfully: {name}", flush=True)
                        
                        # Debounced notification logic
                        import asyncio
                        client = ui.context.client
                        upload_state["count"] += 1
                        if upload_state["timer"]:
                            upload_state["timer"].cancel()
                            
                        async def show_summary():
                            await asyncio.sleep(0.5)
                            with client:
                                count = upload_state["count"]
                                if count > 0:
                                    ui.notify(f"Successfully added {count} file{'s' if count > 1 else ''} to the queue.", color="positive")
                                    upload_state["count"] = 0
                            upload_state["timer"] = None
                            
                        upload_state["timer"] = asyncio.create_task(show_summary())
                    except Exception as err:
                        logger.error(f"Failed to read file {name}: {err}")
                        print(f"Failed to read file {name}: {err}", flush=True)
                        ui.notify(f"Failed to read {name}: {err}", color="negative")
                        
                ui.upload(multiple=True, label="Drag & Drop Folder or Files (.txt, .pdf)", auto_upload=True, on_upload=handle_upload).classes("w-full").props("webkitdirectory directory")

        # Right Panel (Queue & Progress Grid)
        with ui.column().classes("w-2/3 gap-4"):
            
            # Actions & Stats Row
            with ui.row().classes("w-full items-center justify-between p-4 bg-white border").style("border-radius: 12px; border-color: #dadce0;"):
                with ui.row().classes("gap-2"):
                    # Process Button
                    async def start_ingestion():
                        if not queue:
                            ui.notify("Queue is empty. Upload some files first.", color="warning")
                            return
                            
                        # Filter down to tasks that actually need processing
                        run_files = []
                        for fname, txt in queue.items():
                            row = next(r for r in table_rows if r["filename"] == fname)
                            if row["status"] in ("Queued", "Pending", "Failed") or "Failed" in row["status"]:
                                run_files.append((fname, txt))
                                
                        if not run_files:
                            ui.notify("No new files to process.", color="info")
                            return
                            
                        ingest_btn.props("disable")
                        clear_btn.props("disable")
                        export_btn.props("disable")
                        
                        # Set active state
                        for fname, _ in run_files:
                            row = next(r for r in table_rows if r["filename"] == fname)
                            row["status"] = "Queued"
                        refresh_table()
                        
                        # Show progress container
                        progress_container.visible = True
                        progress_label.set_text(f"Processing 0 of {len(run_files)} files...")
                        progress_bar.set_value(0.0)
                        
                        from nicegui import run
                        import asyncio
                        
                        max_workers = int(workers_slider.value)
                        sem = asyncio.Semaphore(max_workers)
                        
                        completed_tasks = 0
                        total_tasks = len(run_files)
                        success_count = 0
                        fail_count = 0
                        
                        async def process_file(fname, txt):
                            nonlocal completed_tasks, success_count, fail_count
                            row = next(r for r in table_rows if r["filename"] == fname)
                            
                            async with sem:
                                row["status"] = "Extracting..."
                                refresh_table()
                                
                                try:
                                    meta = await run.io_bound(
                                        run_extraction,
                                        text=txt,
                                        provider=provider_sel.value,
                                        model_or_path=model_input.value,
                                        api_key=apikey_input.value or None,
                                        base_url=base_url_input.value or None
                                    )
                                    row["case_number"] = meta.case_number
                                    row["date"] = meta.date_of_judgment
                                    row["status"] = "Completed"
                                    
                                    # Append to results list
                                    results[:] = [r for r in results if r["filepath"] != fname]
                                    results.append({
                                        "filepath": fname,
                                        "metadata": meta.model_dump(mode="json")
                                    })
                                    success_count += 1
                                    return True
                                except Exception as exc:
                                    row["status"] = f"Failed: {exc}"
                                    fail_count += 1
                                    return False
                                finally:
                                    completed_tasks += 1
                                    progress_label.set_text(f"Processing {completed_tasks} of {total_tasks} files...")
                                    progress_bar.set_value(completed_tasks / total_tasks)
                                    refresh_table()
                        
                        tasks = [process_file(fname, txt) for fname, txt in run_files]
                        await asyncio.gather(*tasks)
                        
                        ui.notify(f"Completed: {success_count} succeeded, {fail_count} failed.", color="positive" if fail_count == 0 else "warning")
                        
                        # Hide progress container after a short delay
                        await asyncio.sleep(2.0)
                        progress_container.visible = False
                        
                        ingest_btn.props(remove="disable")
                        clear_btn.props(remove="disable")
                        export_btn.props(remove="disable")
                        
                    ingest_btn = ui.button("Start Ingestion", icon="play_arrow", on_click=start_ingestion).props("color=primary rounded")
                    
                    # Clear Button
                    def clear_queue():
                        queue.clear()
                        table_rows.clear()
                        results.clear()
                        refresh_table()
                        ui.notify("Queue cleared", color="grey-7")
                        
                    clear_btn = ui.button("Clear Queue", icon="clear_all", on_click=clear_queue).props("flat rounded color=grey-7")
                
                # Export to Excel Button
                def download_excel():
                    if not results:
                        ui.notify("No completed extractions to export. Run ingestion first.", color="warning")
                        return
                    try:
                        xlsx_data = build_excel_bytes(results)
                        ui.download(xlsx_data, "metadata_registry.xlsx")
                        ui.notify("Excel spreadsheet generated successfully!", color="positive")
                    except Exception as err:
                        ui.notify(f"Excel generation failed: {err}", color="negative")
                        
                export_btn = ui.button("Export to Excel", icon="grid_on", on_click=download_excel).props("color=green rounded")

            # Progress Container Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;") as progress_container:
                progress_container.visible = False
                progress_label = ui.label("Ingesting documents...").classes("text-xs font-semibold text-gray-600 mb-2")
                progress_bar = ui.linear_progress(value=0.0).props("stripe lstrip").classes("w-full")

            # Table Card
            with ui.card().classes("w-full p-4").style("border-radius: 12px; border: 1px solid #dadce0; box-shadow: none;"):
                ui.label("Extraction Queue").classes("text-xs font-bold text-gray-400 uppercase tracking-wider mb-2")
                
                columns = [
                    {"name": "filename", "label": "File Name", "field": "filename", "align": "left"},
                    {"name": "case_number", "label": "Case Number", "field": "case_number", "align": "center"},
                    {"name": "date", "label": "Date of Judgment", "field": "date", "align": "center"},
                    {"name": "status", "label": "Status", "field": "status", "align": "center"}
                ]
                
                results_table = ui.table(columns=columns, rows=table_rows, row_key="filename").classes("w-full shadow-none border")


# Pre-warm the local model in the background at startup so (a) the first breakdown
# isn't slowed by a cold load and (b) two early requests can't race into the model
# loader at once (the race that could wedge the shared-context lock). Non-fatal.
def _prewarm_model():
    import threading
    def _load():
        if settings.llm_provider.lower() == "llamacpp":
            try:
                from src.analyze_smart import _get_llama
                _get_llama()
                print("[prewarm] local model ready")
            except Exception as e:  # noqa: BLE001
                print(f"[prewarm] model preload skipped: {e}")
        # Then the semantic embedder (bge-m3) so deep search is instant; the two
        # fit together in RAM (3B + bge-m3 — do NOT prewarm a 14B alongside).
        try:
            from src.store import warm_embedder
            if warm_embedder():
                print("[prewarm] semantic embedder ready")
        except Exception as e:  # noqa: BLE001
            print(f"[prewarm] embedder preload skipped: {e}")
    threading.Thread(target=_load, name="model-prewarm", daemon=True).start()


_prewarm_model()

ui.run(host=os.getenv("ROSCRIBE_HOST", "127.0.0.1"), port=int(os.getenv("ROSCRIBE_PORT", "8081")), show=False, reload=False,
       storage_secret=STORAGE_SECRET, title="ROS", favicon="data/logos/logo_emblem_trans.png",
       reconnect_timeout=60.0, ws_ping_interval=20, ws_ping_timeout=120,
       gzip_middleware_factory=None)